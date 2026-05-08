"""
Legajos digitales: metadatos en Firestore (empleados y documentos).
Los archivos viven en Firebase Storage bajo el prefijo legajos_uploads/ (ver web_chat).
"""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from typing import Any

LEGAJOS_EMPLEADOS_COLLECTION = "legajos_empleados"
LEGAJOS_DOCUMENTOS_COLLECTION = "legajos_documentos"
LEGAJOS_AUDITORIA_COLLECTION = "legajos_auditoria"
LEGAJOS_CONVENIOS_COLLECTION = "legajos_convenios"
LEGAJOS_CARPETAS_COLLECTION = "legajos_carpetas"

# Lista maestra de todos los tipos de documento posibles.
ALL_TIPOS_DOCUMENTO: list[tuple[str, str]] = [
    ("recibos", "Recibos de sueldo"),
    ("ropa_uniformes", "Ropa / Uniformes"),
    ("vacaciones", "Vacaciones"),
    ("datos_personales", "Datos personales"),
    ("sanciones", "Sanciones disciplinarias"),
    ("notificaciones", "Notificaciones"),
    ("cert_art", "Certificados ART"),
    ("cert_medico", "Certificados médicos"),
    ("cert_escolar", "Certificados escolares"),
    ("carpeta_medica", "Carpeta médica"),
    ("capacitaciones", "Capacitaciones"),
    ("afjp_afip", "AFJP / AFIP"),
    ("otro", "Otro"),
]

CONVENIOS: dict[str, str] = {
    "comercio": "Comercio",
    "camioneros": "Camioneros",
    "bancos": "Bancos",
    "seguridad": "Seguridad privada",
}

# Tipos de documento disponibles por convenio.
# Seguridad privada: sanciones a confirmar con la empresa (marcado con nota).
_TIPOS_COMUNES = [
    ("recibos", "Recibos de sueldo"),
    ("ropa_uniformes", "Ropa / Uniformes"),
    ("vacaciones", "Vacaciones"),
    ("datos_personales", "Datos personales"),
]
TIPOS_DOCUMENTO_POR_CONVENIO: dict[str, list[tuple[str, str]]] = {
    "comercio": [
        *_TIPOS_COMUNES,
        ("sanciones", "Sanciones disciplinarias"),
        ("notificaciones", "Notificaciones"),
        ("cert_art", "Certificados ART"),
        ("otro", "Otro"),
    ],
    "camioneros": [
        *_TIPOS_COMUNES,
        ("sanciones", "Sanciones disciplinarias"),
        ("notificaciones", "Notificaciones"),
        ("cert_medico", "Certificados médicos"),
        ("cert_art", "Certificados ART"),
        ("cert_escolar", "Certificados escolares"),
        ("capacitaciones", "Capacitaciones"),
        ("afjp_afip", "AFJP / AFIP"),
        ("otro", "Otro"),
    ],
    "bancos": [
        *_TIPOS_COMUNES,
        ("sanciones", "Sanciones disciplinarias"),
        ("notificaciones", "Notificaciones"),
        ("carpeta_medica", "Carpeta médica"),
        ("cert_escolar", "Certificados escolares"),
        ("otro", "Otro"),
    ],
    "seguridad": [
        *_TIPOS_COMUNES,
        ("cert_art", "Certificados ART"),
        ("otro", "Otro"),
    ],
    # Sin convenio asignado: tipos genéricos
    "": [
        *_TIPOS_COMUNES,
        ("sanciones", "Sanciones disciplinarias"),
        ("notificaciones", "Notificaciones"),
        ("cert_art", "Certificados ART"),
        ("cert_medico", "Certificados médicos"),
        ("cert_escolar", "Certificados escolares"),
        ("carpeta_medica", "Carpeta médica"),
        ("capacitaciones", "Capacitaciones"),
        ("afjp_afip", "AFJP / AFIP"),
        ("otro", "Otro"),
    ],
}

# Acciones registradas en auditoría (campo action)
LEGAJOS_AUDIT_EMPLEADO_CREAR = "empleado_crear"
LEGAJOS_AUDIT_EMPLEADO_EDITAR = "empleado_editar"
LEGAJOS_AUDIT_EMPLEADO_IMPORTAR = "empleado_importar"
LEGAJOS_AUDIT_EMPLEADO_ELIMINAR = "empleado_eliminar"
LEGAJOS_AUDIT_DOCUMENTO_SUBIR = "documento_subir"
LEGAJOS_AUDIT_DOCUMENTO_DESCARGAR = "documento_descargar"
LEGAJOS_AUDIT_DOCUMENTO_ELIMINAR = "documento_eliminar"

LEGAJOS_IMPORT_MAX_ROWS = 500


def _utc_now():
    return datetime.now(timezone.utc)


def normalize_dni(value: str) -> str:
    """Solo dígitos (para ID y validación)."""
    return "".join(c for c in str(value or "") if c.isdigit())


def make_empleado_doc_id(company_id: str, dni_raw: str) -> str | None:
    """
    ID de documento en Firestore: {empresa}_{dni} (ej. bacar_30123456).
    Requiere 6–16 dígitos en el DNI.
    """
    cid = str(company_id or "").strip().lower()
    digits = normalize_dni(dni_raw)
    if len(digits) < 6 or len(digits) > 16:
        return None
    if not cid:
        return None
    return f"{cid}_{digits}"


def _serialize_ts(value: Any) -> str | None:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    return str(value)


def list_empleados(
    db,
    company_id: str,
    search: str | None = None,
    activo: bool | None = True,
) -> list[dict]:
    """Lista empleados de una empresa. activo=True solo activos, False solo inactivos, None todos."""
    if not db or not company_id:
        return []
    cid = str(company_id).strip().lower()
    q = (search or "").strip().lower()
    out: list[dict] = []
    try:
        for snap in db.collection(LEGAJOS_EMPLEADOS_COLLECTION).where("company_id", "==", cid).stream():
            row = empleado_from_snap(snap)
            if activo is not None and row.get("activo", True) != activo:
                continue
            if q:
                leg = (row.get("legajo_numero") or "").lower()
                nom = (row.get("nombre_completo") or "").lower()
                dni_f = (row.get("dni") or "").lower()
                em = (row.get("email") or "").lower()
                if q not in leg and q not in nom and q not in dni_f and q not in em:
                    continue
            out.append(row)
    except Exception:
        return []
    out.sort(
        key=lambda r: (
            str(r.get("legajo_numero") or ""),
            str(r.get("nombre_completo") or "").lower(),
        )
    )
    return out


def empleado_from_snap(doc_snap) -> dict:
    data = doc_snap.to_dict() or {}
    return {
        "id": doc_snap.id,
        "company_id": data.get("company_id") or "",
        "legajo_numero": data.get("legajo_numero") or "",
        "nombre_completo": data.get("nombre_completo") or "",
        "sucursal": data.get("sucursal") or "",
        "area": data.get("area") or "",
        "notas": data.get("notas") or "",
        "dni": data.get("dni") or "",
        "email": data.get("email") or "",
        "telefono": data.get("telefono") or "",
        "convenio": data.get("convenio") or "",
        "activo": data.get("activo", True),
        "created_at": _serialize_ts(data.get("created_at")),
        "created_by": data.get("created_by") or "",
        "updated_at": _serialize_ts(data.get("updated_at")),
        "updated_by": data.get("updated_by") or "",
    }


def _empleado_row_matches_substring(q: str, emp: dict) -> bool:
    """Misma lógica que el filtro de list_empleados (legajo, nombre, DNI, email)."""
    qn = str(q or "").strip().lower()
    if not qn:
        return True
    leg = (emp.get("legajo_numero") or "").lower()
    nom = (emp.get("nombre_completo") or "").lower()
    dni_f = (emp.get("dni") or "").lower()
    em = (emp.get("email") or "").lower()
    return qn in leg or qn in nom or qn in dni_f or qn in em


def get_empleado(db, empleado_id: str) -> dict | None:
    if not db or not empleado_id:
        return None
    try:
        ref = db.collection(LEGAJOS_EMPLEADOS_COLLECTION).document(str(empleado_id))
        snap = ref.get()
        if not snap.exists:
            return None
        return empleado_from_snap(snap)
    except Exception:
        return None


def legajo_numero_exists(db, company_id: str, legajo_numero: str, exclude_id: str | None = None) -> bool:
    if not db or not company_id or not str(legajo_numero or "").strip():
        return False
    cid = str(company_id).strip().lower()
    leg = str(legajo_numero).strip()
    try:
        for snap in db.collection(LEGAJOS_EMPLEADOS_COLLECTION).where("company_id", "==", cid).stream():
            if exclude_id and snap.id == exclude_id:
                continue
            data = snap.to_dict() or {}
            if str(data.get("legajo_numero") or "").strip() == leg:
                return True
    except Exception:
        return False
    return False


def create_empleado(
    db,
    company_id: str,
    dni: str,
    nombre_completo: str,
    created_by: str,
    legajo_numero: str = "",
    sucursal: str = "",
    area: str = "",
    notas: str = "",
    email: str = "",
    convenio: str = "",
    telefono: str = "",
) -> tuple[bool, dict | None, str]:
    """Crea colaborador. ID en Firestore = {empresa}_{dni}. DNI obligatorio (6–16 dígitos). Legajo opcional."""
    if not db:
        return False, None, "Firestore no disponible."
    cid = str(company_id or "").strip().lower()
    leg = str(legajo_numero or "").strip()
    nom = str(nombre_completo or "").strip()
    dni_digits = normalize_dni(dni)
    if not cid:
        return False, None, "Falta empresa (company_id)."
    if not nom:
        return False, None, "Falta nombre completo."
    if len(dni_digits) < 6 or len(dni_digits) > 16:
        return False, None, "DNI inválido: usá entre 6 y 16 dígitos."
    doc_id = make_empleado_doc_id(cid, dni)
    if not doc_id:
        return False, None, "No se pudo generar el ID del colaborador."
    if leg and legajo_numero_exists(db, cid, leg):
        return False, None, "Ya existe un colaborador con ese legajo en esta empresa."
    now = _utc_now()
    payload = {
        "company_id": cid,
        "legajo_numero": leg,
        "nombre_completo": nom,
        "sucursal": str(sucursal or "").strip(),
        "area": str(area or "").strip(),
        "notas": str(notas or "").strip(),
        "dni": dni_digits,
        "email": str(email or "").strip().lower(),
        "telefono": str(telefono or "").strip(),
        "convenio": str(convenio or "").strip(),
        "activo": True,
        "created_at": now,
        "created_by": str(created_by or "").strip(),
        "updated_at": now,
    }
    try:
        ref = db.collection(LEGAJOS_EMPLEADOS_COLLECTION).document(doc_id)
        if ref.get().exists:
            return False, None, "Ya existe un colaborador con ese DNI en esta empresa."
        ref.set(payload)
        snap = ref.get()
        return True, empleado_from_snap(snap), ""
    except Exception as exc:
        return False, None, str(exc)


def update_empleado(
    db,
    empleado_id: str,
    legajo_numero: str,
    nombre_completo: str,
    updated_by: str,
    sucursal: str = "",
    area: str = "",
    notas: str = "",
    email: str = "",
    convenio: str = "",
    telefono: str = "",
) -> tuple[bool, dict | None, str]:
    """Actualiza datos. El DNI no se modifica (el ID del documento es fijo). Legajo vacío permitido."""
    if not db or not str(empleado_id or "").strip():
        return False, None, "Firestore no disponible o ID inválido."
    eid = str(empleado_id).strip()
    leg = str(legajo_numero or "").strip()
    nom = str(nombre_completo or "").strip()
    if not nom:
        return False, None, "Falta nombre completo."
    try:
        ref = db.collection(LEGAJOS_EMPLEADOS_COLLECTION).document(eid)
        snap = ref.get()
        if not snap.exists:
            return False, None, "Colaborador no encontrado."
        data = snap.to_dict() or {}
        cid = str(data.get("company_id") or "").strip().lower()
        if not cid:
            return False, None, "Registro sin empresa."
        if leg and legajo_numero_exists(db, cid, leg, exclude_id=eid):
            return False, None, "Ya existe otro colaborador con ese legajo en esta empresa."
        now = _utc_now()
        ref.update(
            {
                "legajo_numero": leg,
                "nombre_completo": nom,
                "sucursal": str(sucursal or "").strip(),
                "area": str(area or "").strip(),
                "notas": str(notas or "").strip(),
                "email": str(email or "").strip().lower(),
                "telefono": str(telefono or "").strip(),
                "convenio": str(convenio or "").strip(),
                "updated_at": now,
                "updated_by": str(updated_by or "").strip(),
            }
        )
        snap = ref.get()
        return True, empleado_from_snap(snap), ""
    except Exception as exc:
        return False, None, str(exc)


def set_empleado_activo(db, empleado_id: str, activo: bool) -> tuple[bool, dict | None, str]:
    """Activa o desactiva un colaborador (no elimina sus datos ni archivos)."""
    if not db or not str(empleado_id or "").strip():
        return False, None, "ID inválido."
    eid = str(empleado_id).strip()
    try:
        ref = db.collection(LEGAJOS_EMPLEADOS_COLLECTION).document(eid)
        snap = ref.get()
        if not snap.exists:
            return False, None, "Colaborador no encontrado."
        ref.update({"activo": bool(activo), "updated_at": _utc_now()})
        snap = ref.get()
        return True, empleado_from_snap(snap), ""
    except Exception as exc:
        return False, None, str(exc)


def empleado_tiene_documentos(db, empleado_id: str) -> bool:
    if not db or not empleado_id:
        return False
    eid = str(empleado_id).strip()
    try:
        snaps = list(db.collection(LEGAJOS_DOCUMENTOS_COLLECTION).where("empleado_id", "==", eid).limit(1).stream())
        return len(snaps) > 0
    except Exception:
        return False


def delete_empleado_completo(db, empleado_id: str) -> tuple[bool, list[str], str]:
    """Elimina el colaborador solo si no tiene documentos. Devuelve rutas en Storage a borrar."""
    if not db or not str(empleado_id or "").strip():
        return False, [], "ID inválido."
    eid = str(empleado_id).strip()
    try:
        if empleado_tiene_documentos(db, eid):
            return False, [], "No se puede eliminar: el colaborador tiene documentos. Desactivalo en cambio."
        ref = db.collection(LEGAJOS_EMPLEADOS_COLLECTION).document(eid)
        snap_e = ref.get()
        if not snap_e.exists:
            return False, [], "Colaborador no encontrado."
        ref.delete()
        return True, [], ""
    except Exception as exc:
        return False, [], str(exc)


def list_documentos(db, empleado_id: str) -> list[dict]:
    if not db or not empleado_id:
        return []
    eid = str(empleado_id).strip()
    out: list[dict] = []
    try:
        for snap in (
            db.collection(LEGAJOS_DOCUMENTOS_COLLECTION).where("empleado_id", "==", eid).stream()
        ):
            out.append(documento_from_snap(snap))
    except Exception:
        return []
    out.sort(key=lambda r: str(r.get("uploaded_at") or ""), reverse=True)
    return out


def documento_from_snap(doc_snap) -> dict:
    data = doc_snap.to_dict() or {}
    return {
        "id": doc_snap.id,
        "empleado_id": data.get("empleado_id") or "",
        "company_id": data.get("company_id") or "",
        "tipo_documento": data.get("tipo_documento") or "otro",
        "filename": data.get("filename") or "",
        "storage_path": data.get("storage_path") or "",
        "content_type": data.get("content_type") or "",
        "size_bytes": int(data.get("size_bytes") or 0),
        "uploaded_by": data.get("uploaded_by") or "",
        "uploaded_at": _serialize_ts(data.get("uploaded_at")),
    }


def create_documento(
    db,
    empleado_id: str,
    company_id: str,
    storage_path: str,
    filename: str,
    content_type: str,
    size_bytes: int,
    uploaded_by: str,
    tipo_documento: str = "otro",
) -> tuple[bool, dict | None, str]:
    if not db:
        return False, None, "Firestore no disponible."
    eid = str(empleado_id or "").strip()
    cid = str(company_id or "").strip().lower()
    path = str(storage_path or "").strip()
    if not eid or not cid or not path:
        return False, None, "Datos incompletos para registrar el documento."
    now = _utc_now()
    payload = {
        "empleado_id": eid,
        "company_id": cid,
        "tipo_documento": str(tipo_documento or "otro").strip() or "otro",
        "filename": str(filename or "").strip() or "archivo",
        "storage_path": path,
        "content_type": str(content_type or "").strip(),
        "size_bytes": int(size_bytes or 0),
        "uploaded_by": str(uploaded_by or "").strip(),
        "uploaded_at": now,
    }
    try:
        ref = db.collection(LEGAJOS_DOCUMENTOS_COLLECTION).document()
        ref.set(payload)
        snap = ref.get()
        return True, documento_from_snap(snap), ""
    except Exception as exc:
        return False, None, str(exc)


def get_documento(db, documento_id: str) -> dict | None:
    if not db or not documento_id:
        return None
    try:
        ref = db.collection(LEGAJOS_DOCUMENTOS_COLLECTION).document(str(documento_id))
        snap = ref.get()
        if not snap.exists:
            return None
        return documento_from_snap(snap)
    except Exception:
        return None


def delete_documento_record(db, documento_id: str) -> tuple[bool, dict | None, str]:
    """Elimina solo el documento de Firestore. El blob en Storage lo borra el caller."""
    if not db:
        return False, None, "Firestore no disponible."
    try:
        ref = db.collection(LEGAJOS_DOCUMENTOS_COLLECTION).document(str(documento_id))
        snap = ref.get()
        if not snap.exists:
            return False, None, "Documento no encontrado."
        data = documento_from_snap(snap)
        ref.delete()
        return True, data, ""
    except Exception as exc:
        return False, None, str(exc)


def append_auditoria(
    db,
    company_id: str,
    username: str,
    action: str,
    details: dict | None = None,
) -> None:
    """Registra un evento (best-effort; no lanza)."""
    if not db or not str(company_id or "").strip():
        return
    cid = str(company_id).strip().lower()
    payload = {
        "company_id": cid,
        "username": str(username or "").strip() or "desconocido",
        "action": str(action or "").strip(),
        "details": dict(details) if isinstance(details, dict) else {},
        "at": _utc_now(),
    }
    try:
        db.collection(LEGAJOS_AUDITORIA_COLLECTION).add(payload)
    except Exception:
        pass


def auditoria_from_snap(doc_snap) -> dict:
    data = doc_snap.to_dict() or {}
    det = data.get("details")
    if not isinstance(det, dict):
        det = {}
    return {
        "id": doc_snap.id,
        "company_id": data.get("company_id") or "",
        "username": data.get("username") or "",
        "action": data.get("action") or "",
        "details": det,
        "at": _serialize_ts(data.get("at")),
    }


def _event_at_utc(data: dict) -> datetime | None:
    """Convierte el campo at del documento de auditoría a datetime UTC."""
    v = (data or {}).get("at")
    if isinstance(v, datetime):
        if v.tzinfo is None:
            return v.replace(tzinfo=timezone.utc)
        return v.astimezone(timezone.utc)
    if hasattr(v, "timestamp"):
        try:
            return datetime.fromtimestamp(float(v.timestamp()), tz=timezone.utc)
        except Exception:
            return None
    return None


def _auditoria_details_flat(details: Any) -> str:
    if not isinstance(details, dict):
        return ""
    parts = [f"{k}: {v}" for k, v in details.items()]
    return " · ".join(parts)


def _auditoria_details_search_blob(details: Any) -> str:
    if not isinstance(details, dict):
        return ""
    parts: list[str] = []
    for k, v in details.items():
        parts.append(f"{k}:{v}")
    return " ".join(parts).lower()


def list_auditoria(
    db,
    company_id: str,
    limit: int = 80,
    username: str | None = None,
    action: str | None = None,
    q: str | None = None,
    at_from: datetime | None = None,
    at_to: datetime | None = None,
    max_limit: int = 500,
) -> list[dict]:
    """
    Eventos de una empresa (orden por fecha descendente).
    Filtros opcionales en memoria: fechas UTC (at_from / at_to inclusive),
    usuario (substring), acción (exacta), texto en detalle/username/acción.
    """
    if not db or not company_id:
        return []
    cid = str(company_id).strip().lower()
    cap = max(1, min(int(max_limit or 500), 8000))
    lim = max(1, min(int(limit or 80), cap))
    uname_f = (username or "").strip().lower()
    act_f = (action or "").strip()
    q_f = (q or "").strip().lower()
    try:
        snaps = list(
            db.collection(LEGAJOS_AUDITORIA_COLLECTION).where("company_id", "==", cid).stream()
        )
    except Exception:
        return []

    filtered: list = []
    for s in snaps:
        data = s.to_dict() or {}
        t = _event_at_utc(data)
        if at_from is not None:
            if t is None or t < at_from:
                continue
        if at_to is not None:
            if t is None or t > at_to:
                continue
        if uname_f:
            u = str(data.get("username") or "").lower()
            if uname_f not in u:
                continue
        if act_f and str(data.get("action") or "") != act_f:
            continue
        if q_f:
            blob = _auditoria_details_search_blob(data.get("details"))
            hay = q_f in blob
            hay = hay or q_f in str(data.get("username") or "").lower()
            hay = hay or q_f in str(data.get("action") or "").lower()
            if not hay:
                continue
        filtered.append(s)
    snaps = filtered

    def _sort_key(s):
        d = s.to_dict() or {}
        v = d.get("at")
        if isinstance(v, datetime):
            return v.timestamp()
        if hasattr(v, "timestamp"):
            try:
                return float(v.timestamp())
            except Exception:
                return 0.0
        return 0.0

    try:
        snaps.sort(key=_sort_key, reverse=True)
    except Exception:
        snaps.sort(key=lambda s: s.id, reverse=True)
    return [auditoria_from_snap(s) for s in snaps[:lim]]


def list_documentos_resumen_tipos(
    db,
    company_id: str,
    empleados_search: str | None = None,
) -> list[dict]:
    """
    Cuenta documentos por tipo (carpeta lógica) en la empresa.
    Si empleados_search está definido, solo cuenta archivos de colaboradores que
    coincidan con ese texto (mismo criterio que el buscador de colaboradores).
    Devuelve [{"tipo_documento": str, "count": int}, ...] ordenado por tipo.
    """
    if not db or not str(company_id or "").strip():
        return []
    cid = str(company_id).strip().lower()
    id_filter: set[str] | None = None
    es = str(empleados_search or "").strip()
    if es:
        emps = list_empleados(db, cid, search=es)
        id_filter = {str(r.get("id") or "").strip() for r in emps if r.get("id")}
        if not id_filter:
            return []
    counts: dict[str, int] = {}
    try:
        for snap in db.collection(LEGAJOS_DOCUMENTOS_COLLECTION).where("company_id", "==", cid).stream():
            row = documento_from_snap(snap)
            eid = str(row.get("empleado_id") or "").strip()
            if id_filter is not None and eid not in id_filter:
                continue
            t = str(row.get("tipo_documento") or "otro").strip() or "otro"
            counts[t] = counts.get(t, 0) + 1
    except Exception:
        return []
    out = [{"tipo_documento": k, "count": v} for k, v in counts.items()]
    out.sort(key=lambda x: (str(x.get("tipo_documento") or "").lower(),))
    return out


def search_documentos_empresa(
    db,
    company_id: str,
    q: str = "",
    limit: int = 40,
    tipo_documento: str | None = None,
    empleados_q: str | None = None,
) -> list[dict]:
    """
    Lista o busca documentos de la empresa. Filtros opcionales (al menos uno):
    - q: texto en nombre de archivo y/o en datos del colaborador (nombre, DNI, legajo, email)
    - tipo_documento: carpeta / tipo
    - empleados_q: solo documentos de colaboradores que coincidan con ese texto
      (igual que el buscador superior de la pantalla).
    Devuelve filas enriquecidas con nombre y DNI del colaborador.
    """
    if not db or not str(company_id or "").strip():
        return []
    cid = str(company_id).strip().lower()
    qn = str(q or "").strip().lower()
    tipo_f = str(tipo_documento or "").strip().lower() or None
    eq = str(empleados_q or "").strip()
    id_filter: set[str] | None = None
    if eq:
        emps = list_empleados(db, cid, search=eq)
        id_filter = {str(r.get("id") or "").strip() for r in emps if r.get("id")}
        if not id_filter:
            return []
    if not qn and not tipo_f and not eq:
        return []
    lim = max(1, min(int(limit or 40), 500))
    matches: list[dict] = []
    cache: dict[str, dict] = {}
    try:
        for snap in db.collection(LEGAJOS_DOCUMENTOS_COLLECTION).where("company_id", "==", cid).stream():
            row = documento_from_snap(snap)
            eid = str(row.get("empleado_id") or "").strip()
            if id_filter is not None and eid not in id_filter:
                continue
            rt = str(row.get("tipo_documento") or "otro").strip().lower() or "otro"
            if tipo_f and rt != tipo_f:
                continue
            if eid not in cache:
                cache[eid] = get_empleado(db, eid) or {}
            emp = cache[eid]
            if qn:
                fn_ok = qn in (row.get("filename") or "").lower()
                emp_ok = _empleado_row_matches_substring(qn, emp)
                if not fn_ok and not emp_ok:
                    continue
            matches.append(row)
    except Exception:
        return []
    matches.sort(key=lambda r: str(r.get("uploaded_at") or ""), reverse=True)
    matches = matches[:lim]
    out: list[dict] = []
    for row in matches:
        eid = str(row.get("empleado_id") or "").strip()
        emp = cache.get(eid) or get_empleado(db, eid) or {}
        if eid not in cache:
            cache[eid] = emp
        out.append(
            {
                **row,
                "empleado_nombre": emp.get("nombre_completo") or "",
                "empleado_dni": emp.get("dni") or "",
            }
        )
    return out


def _normalize_import_row(row: dict) -> tuple[str, str, str, str, str, str, str, str]:
    """Mapea encabezados flexibles a legajo, nombre, sucursal, área, notas, dni, email, telefono."""
    m: dict[str, str] = {}
    for k, v in (row or {}).items():
        if k is None:
            continue
        key = str(k).strip().lower().replace(" ", "_")
        m[key] = (str(v).strip() if v is not None else "")
    leg = (
        m.get("legajo_numero")
        or m.get("legajo")
        or m.get("nro_legajo")
        or m.get("numero_legajo")
        or m.get("número_legajo")
        or ""
    )
    nom = (
        m.get("nombre_completo")
        or m.get("nombre")
        or m.get("apellido_y_nombre")
        or m.get("apellido_nombre")
        or ""
    )
    suc = m.get("sucursal") or m.get("branch") or ""
    area = m.get("area") or m.get("área") or ""
    notas = m.get("notas") or m.get("observaciones") or ""
    dni = m.get("dni") or m.get("documento") or m.get("nro_documento") or ""
    email = m.get("email") or m.get("mail") or m.get("correo") or ""
    telefono = m.get("telefono") or m.get("teléfono") or m.get("tel") or m.get("phone") or ""
    return leg.strip(), nom.strip(), suc.strip(), area.strip(), notas.strip(), dni.strip(), email.strip(), telefono.strip()


def parse_legajos_import_file(filename: str, raw: bytes) -> tuple[list[tuple[int, dict]], str | None]:
    """
    Lee CSV (UTF-8) o Excel .xlsx y devuelve filas (nº línea, dict por encabezado)
    en el mismo formato que csv.DictReader, para import_empleados_desde_filas.
    """
    fn = (filename or "").lower().strip()
    if not raw:
        return [], "Archivo vacío."
    if fn.endswith(".csv"):
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            return [], "El CSV debe estar en UTF-8."
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return [], "CSV sin encabezados."
        filas: list[tuple[int, dict]] = []
        line_no = 1
        for row in reader:
            line_no += 1
            filas.append((line_no, dict(row or {})))
        return filas, None
    if fn.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return [], "Falta el paquete openpyxl en el servidor para importar Excel."
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                return [], "El Excel está vacío."
            headers: list[str] = []
            for h in header_row:
                if h is None:
                    headers.append("")
                else:
                    headers.append(str(h).strip())
            if not any(x for x in headers if x):
                return [], "Excel sin encabezados en la primera fila."
            filas = []
            line_no = 1
            for row in rows_iter:
                line_no += 1
                d: dict[str, str] = {}
                for i, key in enumerate(headers):
                    if not key:
                        continue
                    val = row[i] if i < len(row) else None
                    if val is None:
                        d[key] = ""
                    elif hasattr(val, "isoformat") and not isinstance(val, str):
                        try:
                            d[key] = val.isoformat()
                        except Exception:
                            d[key] = str(val).strip()
                    else:
                        d[key] = str(val).strip()
                if any(str(v).strip() for v in d.values()):
                    filas.append((line_no, d))
            return filas, None
        except Exception as exc:
            return [], f"No se pudo leer el Excel: {exc}"
    return [], "Formato no soportado. Usá .csv o .xlsx (Excel)."


def build_legajos_ejemplo_xlsx_bytes() -> tuple[bytes | None, str | None]:
    """Genera un .xlsx de ejemplo con columnas separadas (compatible con Excel)."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return None, "openpyxl no instalado."
    wb = Workbook()
    ws = wb.active
    ws.title = "Colaboradores"
    ws.append(
        ["dni", "nombre_completo", "email", "telefono", "legajo_numero", "sucursal", "area", "notas"]
    )
    ws.append(
        [
            30123456,
            "García Ana María",
            "ana@ejemplo.com",
            "5493513001234",
            "1001",
            "Córdoba",
            "Administración",
            "",
        ]
    )
    ws.append(
        [
            28999111,
            "López Juan Pablo",
            "juan@ejemplo.com",
            "5493513005678",
            "1002",
            "Córdoba",
            "Operaciones",
            "Alta 2025-03",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), None


def build_legajos_export_xlsx_bytes(items: list[dict]) -> tuple[bytes | None, str | None]:
    """
    Exporta colaboradores a .xlsx con las mismas columnas que el ejemplo de importación
    (dni, nombre_completo, email, legajo_numero, sucursal, area, notas).
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        return None, "openpyxl no instalado."
    wb = Workbook()
    ws = wb.active
    ws.title = "Colaboradores"
    ws.append(
        ["dni", "nombre_completo", "email", "telefono", "legajo_numero", "sucursal", "area", "notas"]
    )
    for row in items or []:
        ws.append(
            [
                str(row.get("dni") or ""),
                str(row.get("nombre_completo") or ""),
                str(row.get("email") or ""),
                str(row.get("telefono") or ""),
                str(row.get("legajo_numero") or ""),
                str(row.get("sucursal") or ""),
                str(row.get("area") or ""),
                str(row.get("notas") or ""),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), None


def build_auditoria_export_xlsx_bytes(eventos: list[dict]) -> tuple[bytes | None, str | None]:
    """Excel con columnas separadas para respaldo / control de auditoría de legajos."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return None, "openpyxl no instalado."
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoria legajos"
    ws.append(["fecha_hora_utc", "usuario", "accion", "detalle", "id_evento"])
    for ev in eventos or []:
        ws.append(
            [
                str(ev.get("at") or ""),
                str(ev.get("username") or ""),
                str(ev.get("action") or ""),
                _auditoria_details_flat(ev.get("details")),
                str(ev.get("id") or ""),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), None


def import_empleados_desde_filas(
    db,
    company_id: str,
    filas: list[tuple[int, dict]],
    created_by: str,
) -> dict:
    """
    filas: lista de (número_de_linea, dict como devuelve csv.DictReader).
    Máximo LEGAJOS_IMPORT_MAX_ROWS filas con datos.
    """
    if not db:
        return {"created": 0, "skipped_duplicate": 0, "errors": [{"line": 0, "error": "Firestore no disponible."}]}
    cid = str(company_id or "").strip().lower()
    if not cid:
        return {"created": 0, "skipped_duplicate": 0, "errors": [{"line": 0, "error": "Falta empresa."}]}

    created = 0
    skipped = 0
    errors: list[dict] = []
    processed = 0

    for line_no, raw in filas:
        leg, nom, suc, area, notas, dni, email, telefono = _normalize_import_row(raw)
        if not nom and not dni:
            continue
        processed += 1
        if processed > LEGAJOS_IMPORT_MAX_ROWS:
            errors.append({"line": line_no, "error": f"Límite de {LEGAJOS_IMPORT_MAX_ROWS} filas por importación."})
            break
        if not nom or not dni:
            errors.append({"line": line_no, "error": "Falta nombre o DNI."})
            continue
        ok, _, msg = create_empleado(
            db,
            company_id=cid,
            dni=dni,
            nombre_completo=nom,
            created_by=created_by,
            legajo_numero=leg,
            sucursal=suc,
            area=area,
            notas=notas,
            email=email,
            telefono=telefono,
        )
        if ok:
            created += 1
        else:
            low = (msg or "").lower()
            if "ya existe" in low:
                skipped += 1
            else:
                errors.append({"line": line_no, "error": msg or "Error al crear."})

    return {"created": created, "skipped_duplicate": skipped, "errors": errors}


# ---------------------------------------------------------------------------
# Convenios por empresa
# ---------------------------------------------------------------------------

def _convenio_from_snap(snap) -> dict:
    d = snap.to_dict() or {}
    return {
        "id": snap.id,
        "company_id": d.get("company_id", ""),
        "nombre": d.get("nombre", ""),
        "tipos_documento": d.get("tipos_documento", []),
    }


def list_convenios(db, company_id: str) -> list[dict]:
    if not db or not company_id:
        return []
    cid = str(company_id).strip().lower()
    out: list[dict] = []
    try:
        for snap in db.collection(LEGAJOS_CONVENIOS_COLLECTION).where("company_id", "==", cid).stream():
            out.append(_convenio_from_snap(snap))
    except Exception:
        pass
    return sorted(out, key=lambda x: x["nombre"].lower())


def get_convenio(db, convenio_id: str) -> dict | None:
    if not db or not convenio_id:
        return None
    try:
        snap = db.collection(LEGAJOS_CONVENIOS_COLLECTION).document(convenio_id).get()
        if snap.exists:
            return _convenio_from_snap(snap)
    except Exception:
        pass
    return None


def create_convenio(
    db,
    company_id: str,
    nombre: str,
    tipos_documento: list[str],
) -> tuple[bool, dict | None, str]:
    if not db:
        return False, None, "Sin base de datos."
    cid = str(company_id or "").strip().lower()
    nombre = str(nombre or "").strip()
    if not cid:
        return False, None, "Falta empresa."
    if not nombre:
        return False, None, "El nombre del convenio es obligatorio."
    tipos = [str(t).strip() for t in (tipos_documento or []) if str(t).strip()]
    try:
        ref = db.collection(LEGAJOS_CONVENIOS_COLLECTION).document()
        data = {"company_id": cid, "nombre": nombre, "tipos_documento": tipos, "created_at": _utc_now()}
        ref.set(data)
        row = {"id": ref.id, "company_id": cid, "nombre": nombre, "tipos_documento": tipos}
        return True, row, ""
    except Exception as e:
        return False, None, str(e)


def update_convenio(
    db,
    convenio_id: str,
    nombre: str,
    tipos_documento: list[str],
) -> tuple[bool, dict | None, str]:
    if not db or not convenio_id:
        return False, None, "Falta ID de convenio."
    nombre = str(nombre or "").strip()
    if not nombre:
        return False, None, "El nombre del convenio es obligatorio."
    tipos = [str(t).strip() for t in (tipos_documento or []) if str(t).strip()]
    try:
        ref = db.collection(LEGAJOS_CONVENIOS_COLLECTION).document(convenio_id)
        snap = ref.get()
        if not snap.exists:
            return False, None, "Convenio no encontrado."
        ref.update({"nombre": nombre, "tipos_documento": tipos})
        row = _convenio_from_snap(snap)
        row["nombre"] = nombre
        row["tipos_documento"] = tipos
        return True, row, ""
    except Exception as e:
        return False, None, str(e)


def delete_convenio(db, convenio_id: str) -> tuple[bool, str]:
    if not db or not convenio_id:
        return False, "Falta ID de convenio."
    try:
        ref = db.collection(LEGAJOS_CONVENIOS_COLLECTION).document(convenio_id)
        if not ref.get().exists:
            return False, "Convenio no encontrado."
        ref.delete()
        return True, ""
    except Exception as e:
        return False, str(e)


# ---------------------------------------------------------------------------
# Carpetas disponibles por empresa
# ---------------------------------------------------------------------------

def _slugify_label(label: str) -> str:
    s = label.lower().strip()
    for a, b in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),("ü","u"),("ñ","n")]:
        s = s.replace(a, b)
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return (s or "carpeta")[:50]


def _carpeta_from_snap(snap) -> dict:
    d = snap.to_dict() or {}
    return {
        "id": snap.id,
        "company_id": d.get("company_id", ""),
        "code": d.get("code", ""),
        "label": d.get("label", ""),
        "orden": d.get("orden", 0),
    }


def list_carpetas(db, company_id: str) -> list[dict]:
    """Carpetas configuradas para la empresa. Si no hay, devuelve la lista por defecto."""
    if not db or not company_id:
        return [{"id": "", "company_id": "", "code": c, "label": l, "orden": i} for i, (c, l) in enumerate(ALL_TIPOS_DOCUMENTO)]
    cid = str(company_id).strip().lower()
    out: list[dict] = []
    try:
        for snap in db.collection(LEGAJOS_CARPETAS_COLLECTION).where("company_id", "==", cid).stream():
            out.append(_carpeta_from_snap(snap))
    except Exception:
        pass
    if not out:
        return [{"id": "", "company_id": cid, "code": c, "label": l, "orden": i} for i, (c, l) in enumerate(ALL_TIPOS_DOCUMENTO)]
    return sorted(out, key=lambda x: (x["orden"], x["label"].lower()))


def has_custom_carpetas(db, company_id: str) -> bool:
    if not db or not company_id:
        return False
    cid = str(company_id).strip().lower()
    try:
        snaps = list(db.collection(LEGAJOS_CARPETAS_COLLECTION).where("company_id", "==", cid).limit(1).stream())
        return len(snaps) > 0
    except Exception:
        return False


def create_carpeta(db, company_id: str, label: str) -> tuple[bool, dict | None, str]:
    if not db:
        return False, None, "Sin base de datos."
    cid = str(company_id or "").strip().lower()
    label = str(label or "").strip()
    if not cid:
        return False, None, "Falta empresa."
    if not label:
        return False, None, "El nombre de la carpeta es obligatorio."
    code = _slugify_label(label)
    # Verificar que no exista ya ese code en la empresa
    try:
        existing = list(db.collection(LEGAJOS_CARPETAS_COLLECTION).where("company_id", "==", cid).where("code", "==", code).limit(1).stream())
        if existing:
            return False, None, f'Ya existe una carpeta con el código "{code}".'
        # Obtener el máximo orden actual
        all_snaps = list(db.collection(LEGAJOS_CARPETAS_COLLECTION).where("company_id", "==", cid).stream())
        orden = len(all_snaps)
        ref = db.collection(LEGAJOS_CARPETAS_COLLECTION).document()
        data = {"company_id": cid, "code": code, "label": label, "orden": orden, "created_at": _utc_now()}
        ref.set(data)
        row = {"id": ref.id, "company_id": cid, "code": code, "label": label, "orden": orden}
        return True, row, ""
    except Exception as e:
        return False, None, str(e)


def delete_carpeta(db, carpeta_id: str) -> tuple[bool, str]:
    if not db or not carpeta_id:
        return False, "Falta ID de carpeta."
    try:
        ref = db.collection(LEGAJOS_CARPETAS_COLLECTION).document(carpeta_id)
        if not ref.get().exists:
            return False, "Carpeta no encontrada."
        ref.delete()
        return True, ""
    except Exception as e:
        return False, str(e)


def init_carpetas_from_default(db, company_id: str) -> bool:
    """Inicializa las carpetas de una empresa con la lista por defecto."""
    if not db or not company_id:
        return False
    cid = str(company_id).strip().lower()
    if has_custom_carpetas(db, cid):
        return False
    try:
        for i, (code, label) in enumerate(ALL_TIPOS_DOCUMENTO):
            ref = db.collection(LEGAJOS_CARPETAS_COLLECTION).document()
            ref.set({"company_id": cid, "code": code, "label": label, "orden": i, "created_at": _utc_now()})
        return True
    except Exception:
        return False
