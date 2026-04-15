import os

# Cargar variables desde .env si existe python-dotenv (opcional)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import csv
import io
import logging
import re
import smtplib
import uuid
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from functools import wraps

from flask import (
    Flask,
    g,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)

import app as chatbot
import auth_rrhh
import legajos_service
import stats_service

# Conectar Firestore a auth_rrhh para persistencia de usuarios y roles
auth_rrhh.set_firestore_db(chatbot.db)

# Activar envío por WhatsApp vía Twilio si hay credenciales
try:
    from twilio_whatsapp import register_twilio_sender
    register_twilio_sender()
except ImportError:
    pass

# Web Push
try:
    import hashlib as _hashlib
    import threading as _threading
    import json as _json_mod
    import base64 as _b64mod
    from pywebpush import webpush as _webpush_send, WebPushException as _WebPushException
    from cryptography.hazmat.primitives.asymmetric import ec as _ec
    from cryptography.hazmat.primitives.asymmetric.ec import SECP256R1 as _SECP256R1
    from cryptography.hazmat.primitives import serialization as _crypto_serial
    _PUSH_AVAILABLE = True
except ImportError:
    _PUSH_AVAILABLE = False

_VAPID_PRIVATE_KEY: str | None = None
_VAPID_PUBLIC_KEY: str | None = None
_VAPID_INITIALIZED = False


def _vapid_pem_to_raw_b64(pem_str):
    """Convierte PEM EC private key a raw 32-byte base64url (formato requerido por pywebpush)."""
    try:
        k = _crypto_serial.load_pem_private_key(pem_str.encode(), password=None)
        raw = k.private_numbers().private_value.to_bytes(32, "big")
        return _b64mod.urlsafe_b64encode(raw).rstrip(b"=").decode("utf-8")
    except Exception:
        return None


def _ensure_vapid_keys():
    global _VAPID_PRIVATE_KEY, _VAPID_PUBLIC_KEY, _VAPID_INITIALIZED
    if _VAPID_INITIALIZED:
        return
    if not _PUSH_AVAILABLE:
        _VAPID_INITIALIZED = True
        return
    # 1. env vars
    priv = os.getenv("VAPID_PRIVATE_KEY", "").strip()
    pub = os.getenv("VAPID_PUBLIC_KEY", "").strip()
    if priv and pub:
        if priv.startswith("-----"):
            priv = _vapid_pem_to_raw_b64(priv) or priv
        _VAPID_PRIVATE_KEY, _VAPID_PUBLIC_KEY = priv, pub
        _VAPID_INITIALIZED = True
        return
    # 2. Firestore
    if chatbot.db:
        try:
            doc = chatbot.db.collection("rrhh_config").document("vapid").get()
            if doc.exists:
                data = doc.to_dict() or {}
                priv = (data.get("private_key") or "").strip()
                pub = (data.get("public_key") or "").strip()
                if priv and pub:
                    # Convertir de PEM a raw base64url si fue guardado en formato PEM
                    if priv.startswith("-----"):
                        converted = _vapid_pem_to_raw_b64(priv)
                        if converted:
                            priv = converted
                            try:
                                chatbot.db.collection("rrhh_config").document("vapid").update({"private_key": priv})
                            except Exception:
                                pass
                    if priv:
                        _VAPID_PRIVATE_KEY = priv
                        _VAPID_PUBLIC_KEY = pub
                        _VAPID_INITIALIZED = True
                        return
        except Exception as exc:
            logging.warning(f"VAPID Firestore read failed: {exc}")
    # 3. Generate new key in raw base64url format (required by pywebpush)
    try:
        priv_key = _ec.generate_private_key(_SECP256R1())
        priv_raw = priv_key.private_numbers().private_value.to_bytes(32, "big")
        priv_b64 = _b64mod.urlsafe_b64encode(priv_raw).rstrip(b"=").decode("utf-8")
        pub_bytes = priv_key.public_key().public_bytes(
            _crypto_serial.Encoding.X962,
            _crypto_serial.PublicFormat.UncompressedPoint,
        )
        pub_b64 = _b64mod.urlsafe_b64encode(pub_bytes).rstrip(b"=").decode("utf-8")
        _VAPID_PRIVATE_KEY = priv_b64
        _VAPID_PUBLIC_KEY = pub_b64
        _VAPID_INITIALIZED = True
        if chatbot.db:
            try:
                chatbot.db.collection("rrhh_config").document("vapid").set({
                    "private_key": priv_b64,
                    "public_key": pub_b64,
                })
            except Exception as exc:
                logging.warning(f"VAPID Firestore write failed: {exc}")
    except Exception as exc:
        logging.warning(f"VAPID key generation failed: {exc}")


def _send_push_all(title: str, body: str, url: str = "/rrhh"):
    if not _PUSH_AVAILABLE or not chatbot.db:
        return
    _ensure_vapid_keys()
    if not _VAPID_PRIVATE_KEY:
        return
    payload = _json_mod.dumps({"title": title, "body": body[:200], "url": url, "tag": "rrhh-push"})
    vapid_private = _VAPID_PRIVATE_KEY

    def _do():
        try:
            docs = list(chatbot.db.collection("push_subscriptions").stream())
            expired = []
            for doc in docs:
                sub_info = (doc.to_dict() or {}).get("subscription")
                if not sub_info:
                    continue
                try:
                    _webpush_send(
                        subscription_info=sub_info,
                        data=payload,
                        vapid_private_key=vapid_private,
                        vapid_claims={"sub": "mailto:noreply@debo-chat.web.app"},
                    )
                except _WebPushException as exc:
                    resp = getattr(exc, "response", None)
                    if resp is not None and resp.status_code in (404, 410):
                        expired.append(doc.id)
                    else:
                        logging.warning(f"Push error: {exc}")
                except Exception as exc:
                    logging.warning(f"Push error: {exc}")
            for eid in expired:
                try:
                    chatbot.db.collection("push_subscriptions").document(eid).delete()
                except Exception:
                    pass
        except Exception as exc:
            logging.warning(f"Push send error: {exc}")

    _threading.Thread(target=_do, daemon=True).start()


def _notify_handoff_via_n8n(handoff_payload: dict, company: dict):
    """Envía email de notificación cuando se crea un nuevo handoff."""
    # Para WA: buscar el email del número específico que recibió el mensaje
    notify_email = ""
    wa_from_number = str(handoff_payload.get("whatsapp_from_number") or "").strip()
    wa_numbers = company.get("whatsapp_numbers") or []
    if wa_from_number:
        for line in wa_numbers:
            p = str(line.get("phone") or "").strip()
            if p and _normalize_phone_for_match(p) == _normalize_phone_for_match(wa_from_number):
                notify_email = str(line.get("notify_email") or "").strip()
                break
    # Si no hubo match exacto y wa_from_number parece un ID numérico de Meta
    # (ej. 1078605635336424), intentar con el primer número que tenga notify_email
    if not notify_email and wa_from_number and wa_from_number.isdigit() and len(wa_from_number) > 10:
        for line in wa_numbers:
            candidate = str(line.get("notify_email") or "").strip()
            if candidate:
                notify_email = candidate
                break
    # Fallback: email general de la empresa
    if not notify_email:
        notify_email = str(company.get("handoff_notify_email") or "").strip()
    if not notify_email:
        return

    def _do():
        try:
            # Solo notificar por email si no hay agentes activos en el panel.
            # Cuando el panel está abierto los agentes ven el handoff en tiempo real.
            # Usamos TTL corto (60s): si no hubo heartbeat en el último minuto → panel cerrado.
            _cid = company.get("company_id")
            _active = _list_active_agents(ttl_seconds=60, company_id=_cid)
            if _active:
                logging.info(
                    f"Handoff notify: omitiendo email, hay {len(_active)} agente(s) activo(s) en el panel"
                )
                return

            company_name = handoff_payload.get("company_name") or handoff_payload.get("company_id") or "Sin empresa"
            colaborador = handoff_payload.get("colaborador_nombre") or handoff_payload.get("colaborador_telefono") or "Colaborador"
            consulta = handoff_payload.get("ultima_consulta") or "Solicitud de atención"
            area = handoff_payload.get("area") or ""
            branch = handoff_payload.get("branch") or ""
            canal = "WhatsApp" if handoff_payload.get("channel") == "whatsapp" else "Web"

            subject = f"[{company_name}] Nueva consulta de {colaborador}"

            lineas = [
                f"Nueva conversación iniciada en el ChatBot RRHH.",
                "",
                f"Empresa:     {company_name}",
                f"Colaborador: {colaborador}",
                f"Canal:       {canal}",
            ]
            if area:
                lineas.append(f"Área:        {area}")
            if branch:
                lineas.append(f"Sucursal:    {branch}")
            resumen = handoff_payload.get("resumen_conversacion") or ""
            lineas += [
                "",
                "Último mensaje:",
                consulta,
            ]
            if resumen:
                lineas += ["", "Resumen de la conversación:", resumen]
            lineas += ["", "Ver en el panel: https://debo-chat.web.app/?m=rrhh"]
            body = "\n".join(lineas)

            ok, err = _send_email(notify_email, subject, body)
            if not ok:
                logging.warning(f"Handoff notify email error: {err}")
        except Exception as exc:
            logging.warning(f"Handoff notify email error: {exc}")

    _threading.Thread(target=_do, daemon=True).start()


logger = logging.getLogger(__name__)
flask_app = Flask(__name__)
flask_app.config["SECRET_KEY"] = os.getenv("CHATBOT_WEB_SECRET", "dev-chatbot-secret")
BOT_NAME = os.getenv("CHATBOT_BOT_NAME", "Debo")
# Firebase Hosting preserves the "__session" cookie across rewrites to Cloud Run.
flask_app.config["SESSION_COOKIE_NAME"] = os.getenv("CHATBOT_SESSION_COOKIE_NAME", "__session")
# Timeout de inactividad: 8 horas (en segundos)
RRHH_SESSION_TIMEOUT_SECONDS = int(os.getenv("CHATBOT_SESSION_TIMEOUT", 8 * 3600))
SERVER_BOOT_AT = datetime.now(timezone.utc).isoformat(timespec="seconds")

BOOTSTRAP_COMPANY_NAME = (
    str(os.getenv("CHATBOT_COMPANY_NAME", getattr(chatbot, "COMPANY_NAME", "Bacar"))).strip()
    or "Bacar"
)
BOOTSTRAP_HR_TEAM_NAME = (
    str(os.getenv("CHATBOT_HR_TEAM_NAME", getattr(chatbot, "HR_TEAM_NAME", "Atención"))).strip()
    or "Atención"
)
BOOTSTRAP_HR_CONTACT = (
    str(os.getenv("CHATBOT_HR_CONTACT", getattr(chatbot, "HR_CONTACT", "interno 104"))).strip()
    or "interno 104"
)

HANDOFF_STATUS_PENDING = "pendiente"
HANDOFF_STATUS_ACTIVE = "en_atencion"
HANDOFF_STATUS_CLOSED = "cerrada"

HANDOFF_END_COMMANDS = {
    "__cerrar_rrhh__",
    "cerrar rrhh",
    "finalizar rrhh",
    "terminar rrhh",
    "volver al bot",
}
HANDOFF_POLL_COMMANDS = {"__poll_rrhh__", "actualizar rrhh", "actualizar"}

IN_MEMORY_HANDOFFS = {}
IN_MEMORY_CHAT_HISTORY = []
IN_MEMORY_ACTIVE_AGENTS = {}
IN_MEMORY_GENERAL_SETTINGS = {}
IN_MEMORY_COMPANIES = {}

# Caché TTL simple para lecturas frecuentes de Firestore
import time as _time_mod
_TTL_CACHE: dict = {}
_TTL_CACHE_SECONDS = 30  # segundos

def _cache_get(key):
    entry = _TTL_CACHE.get(key)
    if entry and (_time_mod.time() - entry["ts"]) < _TTL_CACHE_SECONDS:
        return entry["data"]
    return None

def _cache_set(key, data):
    _TTL_CACHE[key] = {"data": data, "ts": _time_mod.time()}

def _cache_del(*keys):
    for k in keys:
        _TTL_CACHE.pop(k, None)

GENERAL_SETTINGS_COLLECTION = "chatbot_config"
GENERAL_SETTINGS_DOC = "general"
ACTIVE_AGENTS_COLLECTION = "rrhh_agentes"
COMPANIES_COLLECTION = "chatbot_empresas"
# Clave Firestore-safe para "Todas (ámbito empresa)" en areas_by_branch (Firestore no permite "" como nombre de campo).
AREA_BRANCH_KEY_EMPRESA = "__e"

try:
    ACTIVE_AGENT_TTL_SECONDS = max(
        60, int(str(os.getenv("RRHH_AGENT_ACTIVE_TTL_SECONDS", "180")).strip())
    )
except Exception:
    ACTIVE_AGENT_TTL_SECONDS = 180

AUTO_CLOSE_MIN_MINUTES = 5
AUTO_CLOSE_MAX_MINUTES = 7 * 24 * 60
AUTO_CLOSE_DEFAULT_MINUTES = 0

# Sesión por número de WhatsApp cuando el mensaje llega por webhook (colaborador escribe por WA).
WHATSAPP_SESSIONS = {}

WHATSAPP_CONTEXT_COLLECTION = "whatsapp_chat_context"


def _load_whatsapp_chat_context(phone):
    """Carga contexto de chat (empresa/sucursal/área/step) desde Firestore para este número."""
    if not chatbot.db or not phone:
        return
    norm = _normalize_phone_for_match(phone)
    if not norm:
        return
    try:
        doc = chatbot.db.collection(WHATSAPP_CONTEXT_COLLECTION).document(norm).get()
        if not doc.exists:
            return
        data = doc.to_dict() or {}
    except Exception:
        return
    sess = getattr(g, "whatsapp_session", None)
    if not sess:
        return
    for key in ("chat_context_step", "chat_context_company_id", "chat_context_branch", "chat_context_area", "company_id", "company_name", "wa_empleado_id", "wa_convenio", "wa_nombre", "chat_session_id", "handoff_conversation_id"):
        if key in data and data[key] is not None:
            sess[key] = data[key]


def _reset_whatsapp_chat_context(phone):
    """Al cerrar un handoff: limpia solo el estado de handoff pero preserva empresa/sucursal/área."""
    if not phone:
        return
    # Preservar los campos de contexto que el colaborador ya eligió
    _KEEP_KEYS = {
        "chat_context_step", "chat_context_company_id", "chat_context_branch",
        "chat_context_area", "company_id", "company_name",
        "wa_empleado_id", "wa_convenio", "wa_nombre",
        "chat_session_id", "meta_phone_number_id",
    }
    if phone in WHATSAPP_SESSIONS:
        sess = WHATSAPP_SESSIONS[phone]
        keys_to_remove = [k for k in list(sess.keys()) if k not in _KEEP_KEYS]
        for k in keys_to_remove:
            sess.pop(k, None)
    # En Firestore: limpiar handoff_conversation_id y asegurarse de que el step queda en "ready".
    # Un handoff solo puede estar activo cuando el step era "ready", así que al cerrarlo
    # siempre corresponde volver a "ready". Esto evita que la siguiente sesión arranque
    # en step "company" si el doc de Firestore no tenía el step guardado.
    norm = _normalize_phone_for_match(phone)
    if norm and chatbot.db:
        try:
            # Leer step actual del doc para no pisarlo si ya tenía algo válido distinto de "ready"
            _fs_step = CHAT_CONTEXT_STEP_READY
            if phone in WHATSAPP_SESSIONS and WHATSAPP_SESSIONS[phone].get("chat_context_step"):
                _fs_step = WHATSAPP_SESSIONS[phone]["chat_context_step"]
            chatbot.db.collection(WHATSAPP_CONTEXT_COLLECTION).document(norm).set(
                {
                    "handoff_conversation_id": None,
                    "chat_context_step": _fs_step,
                    "updated_at": _utc_now(),
                },
                merge=True,
            )
        except Exception as e:
            logger.warning("_reset_whatsapp_chat_context: error para %s: %s", norm, e)


def _save_whatsapp_chat_context(phone):
    """Guarda en Firestore el contexto de chat de la sesión WhatsApp actual."""
    if not chatbot.db or not phone:
        return
    norm = _normalize_phone_for_match(phone)
    if not norm:
        return
    sess = getattr(g, "whatsapp_session", None)
    if not sess:
        return
    # No guardar si la sesión no tiene datos útiles
    if not sess.get("chat_context_step") and not sess.get("company_id"):
        return
    payload = {
        "chat_context_step": sess.get("chat_context_step"),
        "chat_context_company_id": sess.get("chat_context_company_id"),
        "chat_context_branch": sess.get("chat_context_branch"),
        "chat_context_area": sess.get("chat_context_area"),
        "company_id": sess.get("company_id"),
        "company_name": sess.get("company_name"),
        "wa_empleado_id": sess.get("wa_empleado_id"),
        "wa_convenio": sess.get("wa_convenio"),
        "wa_nombre": sess.get("wa_nombre"),
        "chat_session_id": sess.get("chat_session_id"),
        "handoff_conversation_id": sess.get("handoff_conversation_id"),
        "updated_at": _utc_now(),
    }
    payload = {k: v for k, v in payload.items() if v is not None}
    try:
        chatbot.db.collection(WHATSAPP_CONTEXT_COLLECTION).document(norm).set(payload, merge=True)
    except Exception as e:
        logger.warning("_save_whatsapp_chat_context: error guardando sesión para %s: %s", norm, e)


def _get_whatsapp_identity(phone):
    """Devuelve la identidad vinculada a este número WA (empleado_id, convenio, empresa_id, nombre) o None."""
    if not chatbot.db or not phone:
        return None
    norm = _normalize_phone_for_match(phone)
    if not norm:
        return None
    try:
        doc = chatbot.db.collection(WA_IDENTITIES_COLLECTION).document(norm).get()
        return doc.to_dict() if doc.exists else None
    except Exception as e:
        logger.warning("_get_whatsapp_identity: error para %s: %s", norm, e)
        return None


def _save_whatsapp_identity(phone, empleado_id, convenio, empresa_id, nombre):
    """Guarda la vinculación número WA ↔ empleado en Firestore."""
    if not chatbot.db or not phone:
        return
    norm = _normalize_phone_for_match(phone)
    if not norm:
        return
    try:
        chatbot.db.collection(WA_IDENTITIES_COLLECTION).document(norm).set({
            "empleado_id": empleado_id,
            "convenio": convenio or "",
            "empresa_id": empresa_id or "",
            "nombre": nombre or "",
            "updated_at": _utc_now(),
        }, merge=True)
    except Exception as e:
        logger.warning("_save_whatsapp_identity: error para %s: %s", norm, e)


def _sess():
    """Sesión efectiva: por WhatsApp (g.whatsapp_session) o sesión web (session)."""
    if getattr(g, "whatsapp_session", None) is not None:
        return g.whatsapp_session
    return session


def _accion(label, value, variant="default"):
    return {"label": label, "value": value, "variant": variant}


# Contexto de chat: empresa, sucursal y área elegidas por el colaborador (el asistente pregunta en ese orden).
CHAT_CONTEXT_STEP_DNI = "dni"      # identificación por DNI (solo WhatsApp, primer contacto)
CHAT_CONTEXT_STEP_COMPANY = "company"
CHAT_CONTEXT_STEP_BRANCH = "branch"
CHAT_CONTEXT_STEP_AREA = "area"
CHAT_CONTEXT_STEP_READY = "ready"

# Colección Firestore para persistir la vinculación número WA ↔ empleado
WA_IDENTITIES_COLLECTION = "whatsapp_identities"


def _chat_context_step():
    return _sess().get("chat_context_step") or CHAT_CONTEXT_STEP_COMPANY


def _set_chat_context_company(company_id):
    key = _normalize_company_id(company_id)
    if key:
        _sess()["chat_context_company_id"] = key
        _set_company_session(key)
        company = _get_company(key, include_inactive=False)
        branches = _normalize_branches((company or {}).get("branches"))
        if branches:
            _sess()["chat_context_step"] = CHAT_CONTEXT_STEP_BRANCH
        else:
            _sess()["chat_context_step"] = CHAT_CONTEXT_STEP_AREA
            _sess().pop("chat_context_branch", None)


def _set_chat_context_branch(branch_name):
    branch_clean = str(branch_name or "").strip()
    if branch_clean:
        _sess()["chat_context_branch"] = branch_clean
        _sess()["chat_context_step"] = CHAT_CONTEXT_STEP_AREA


def _set_chat_context_area(area_name):
    area_clean = str(area_name or "").strip()
    if area_clean:
        _sess()["chat_context_area"] = area_clean
        _sess()["chat_context_step"] = CHAT_CONTEXT_STEP_READY


def _clear_chat_context():
    _sess().pop("chat_context_step", None)
    _sess().pop("chat_context_company_id", None)
    _sess().pop("chat_context_branch", None)
    _sess().pop("chat_context_area", None)


def _parse_menu_number(mensaje):
    """Si el mensaje es un número de menú (ej. '1', '1.', '2)', 'opción 2'), devuelve el índice 1-based o None.
    Solo reconoce el número si el mensaje es básicamente solo un número (con puntuación opcional o
    prefijo corto tipo 'opción'/'sucursal'), para evitar extraer dígitos de nombres como 'test 3'."""
    s = (mensaje or "").strip()
    if not s:
        return None
    # Caso exacto: solo dígito(s), opcionalmente con "." o ")" al final
    s_clean = s.rstrip(".)")
    if s_clean.isdigit():
        n = int(s_clean)
        return n if n >= 1 else None
    # Caso: prefijo genérico + número (ej. "opción 2", "nro 4") — solo prefijos conocidos
    _PREFIJOS_MENU = {"opcion", "opción", "numero", "número", "nro", "item", "ítem"}
    m = re.match(r'^(\w+)\s+([1-9]\d*)[.)]*$', s.strip().lower())
    if m and m.group(1) in _PREFIJOS_MENU:
        return int(m.group(2))
    return None


def _resolve_message_to_company(mensaje):
    """Devuelve (company_id, company) si el mensaje coincide con una empresa (nombre, id o número de menú)."""
    mensaje_norm = chatbot.normalizar_texto(mensaje)
    if not mensaje_norm:
        return None, None
    lista = _list_companies(include_inactive=False)
    num = _parse_menu_number(mensaje)
    if num is not None and 1 <= num <= len(lista):
        item = lista[num - 1]
        cid = _normalize_company_id(item.get("company_id"))
        return cid, item
    msg_nospace = mensaje_norm.replace(" ", "")
    for item in lista:
        cid = _normalize_company_id(item.get("company_id"))
        raw_name = (item.get("company_name") or cid or "").strip()
        name = chatbot.normalizar_texto(raw_name)
        cid_norm = chatbot.normalizar_texto(cid or "")
        if not name:
            name = cid_norm
        if mensaje_norm == cid_norm or mensaje_norm == name:
            return cid, item
        if name and (mensaje_norm in name or (len(mensaje_norm) >= 2 and name in mensaje_norm)):
            return cid, item
        # Comparación sin espacios: "test3" coincide con "test 3"
        name_nospace = name.replace(" ", "")
        cid_nospace = cid_norm.replace(" ", "")
        if name_nospace and len(name_nospace) >= 2 and (
            msg_nospace == name_nospace or
            name_nospace in msg_nospace or
            cid_nospace == msg_nospace or
            (len(cid_nospace) >= 2 and cid_nospace in msg_nospace)
        ):
            return cid, item
    return None, None


def _resolve_message_to_area(mensaje, company_id, branch=None):
    """Devuelve el nombre del área si el mensaje coincide (nombre o número de menú)."""
    company = _get_company(company_id, include_inactive=False)
    if branch is not None and str(branch or "").strip():
        areas = _get_areas_for_branch(company, str(branch).strip())
    else:
        areas = _get_all_areas_for_company(company)
    mensaje_norm = chatbot.normalizar_texto(mensaje)
    if not mensaje_norm or not areas:
        return None
    num = _parse_menu_number(mensaje)
    if num is not None and 1 <= num <= len(areas):
        return str(areas[num - 1]).strip()
    for a in areas:
        an = str(a).strip()
        if chatbot.normalizar_texto(an) == mensaje_norm or mensaje_norm in chatbot.normalizar_texto(an):
            return an
    return None


def _construir_acciones_empresas(limite=10):
    acciones = []
    for i, item in enumerate(_list_companies(include_inactive=False)[:limite], start=1):
        cid = item.get("company_id")
        name = (item.get("company_name") or cid or "").strip()
        if cid and name:
            acciones.append(_accion(f"{i}. {name}", cid, "topic"))
    return acciones


def _get_branches_for_company(company):
    """Lista de nombres de sucursal para la empresa (para menú chat)."""
    if not company:
        return []
    return [_branch_name(b) for b in (company.get("branches") or []) if _branch_name(b)]


def _resolve_message_to_branch(mensaje, company_id):
    """Devuelve el nombre de la sucursal si el mensaje coincide (nombre o número de menú)."""
    company = _get_company(company_id, include_inactive=False)
    branches = _get_branches_for_company(company)
    mensaje_norm = chatbot.normalizar_texto(mensaje)
    if not mensaje_norm or not branches:
        return None
    num = _parse_menu_number(mensaje)
    if num is not None and 1 <= num <= len(branches):
        return str(branches[num - 1]).strip()
    for b in branches:
        bn = str(b).strip()
        if chatbot.normalizar_texto(bn) == mensaje_norm or mensaje_norm in chatbot.normalizar_texto(bn):
            return bn
    return None


def _construir_acciones_sucursales(company_id, limite=12):
    company = _get_company(company_id, include_inactive=False)
    branches = _get_branches_for_company(company)
    acciones = []
    for i, name in enumerate((branches or [])[:limite], start=1):
        if name:
            acciones.append(_accion(f"{i}. {name}", name, "topic"))
    return acciones


def _get_all_areas_for_company(company):
    """Unión de áreas de todas las sucursales + ámbito empresa (para menú chat)."""
    if not company:
        return []
    areas_by_branch = company.get("areas_by_branch")
    if isinstance(areas_by_branch, dict) and areas_by_branch:
        seen = set()
        out = []
        for branch in [""] + list(company.get("branches") or []):
            branch_key = "" if branch == "" else _branch_name(branch)
            for a in _get_areas_for_branch(company, branch_key):
                k = str(a).strip().lower()
                if k and k not in seen:
                    seen.add(k)
                    out.append(a)
        return out
    return _normalize_areas(company.get("areas"))


def _construir_acciones_areas(company_id, limite=12, branch=None):
    company = _get_company(company_id, include_inactive=False)
    if branch is not None and str(branch or "").strip():
        areas = _get_areas_for_branch(company, str(branch).strip())
    else:
        areas = _get_all_areas_for_company(company)
    acciones = []
    for i, a in enumerate((areas or [])[:limite], start=1):
        an = str(a).strip()
        if an:
            acciones.append(_accion(f"{i}. {an}", an, "topic"))
    return acciones


def _utc_now():
    return datetime.now(timezone.utc)


def _new_conversation_id():
    return f"conv-{uuid.uuid4().hex[:12]}"


def _as_utc_naive(dt):
    if not isinstance(dt, datetime):
        return None
    if dt.tzinfo is None:
        return dt
    return dt.astimezone(timezone.utc).replace(tzinfo=None)


def _as_utc_aware(dt):
    if not isinstance(dt, datetime):
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _iso_utc(dt):
    dt2 = _as_utc_aware(dt)
    if not dt2:
        return ""
    return dt2.isoformat(timespec="seconds").replace("+00:00", "Z")


def _fmt_fecha(dt):
    dt2 = _as_utc_naive(dt)
    if not dt2:
        return "Sin fecha"
    return dt2.strftime("%Y-%m-%d %H:%M")


def _is_true_env(value, default=True):
    raw = str(value if value is not None else "").strip().lower()
    if not raw:
        return default
    return raw in {"1", "true", "yes", "on", "si", "sí"}


def _smtp_settings_from_firestore():
    """Lee la configuración SMTP guardada en Firestore (rrhh_config/smtp)."""
    try:
        if chatbot.db:
            doc = chatbot.db.collection("rrhh_config").document("smtp").get()
            if doc.exists:
                return doc.to_dict() or {}
    except Exception:
        pass
    return {}


def _smtp_settings():
    try:
        port = int(str(os.getenv("SMTP_PORT", "587")).strip() or "587")
    except Exception:
        port = 587
    env_host = str(os.getenv("SMTP_HOST", "")).strip()
    env_user = str(os.getenv("SMTP_USER", "")).strip()
    env_pass = str(os.getenv("SMTP_PASSWORD", "")).strip()
    env_from = str(os.getenv("SMTP_FROM", "")).strip()

    # Si las env vars están configuradas (no son placeholders), usarlas
    _placeholder_hosts = {"", "smtp.ejemplo.com", "smtp.example.com"}
    if env_host and env_host not in _placeholder_hosts:
        return {
            "host": env_host,
            "port": port,
            "username": env_user,
            "password": env_pass,
            "from_email": env_from,
            "use_tls": _is_true_env(os.getenv("SMTP_USE_TLS"), default=True),
        }

    # Fallback: leer de Firestore
    fs = _smtp_settings_from_firestore()
    try:
        fs_port = int(str(fs.get("port", 587) or 587))
    except Exception:
        fs_port = 587
    return {
        "host": str(fs.get("host") or "").strip(),
        "port": fs_port,
        "username": str(fs.get("username") or "").strip(),
        "password": str(fs.get("password") or "").strip(),
        "from_email": str(fs.get("from_email") or "").strip(),
        "use_tls": bool(fs.get("use_tls", True)),
    }


def _send_email(to_email, subject, body_text):
    cfg = _smtp_settings()
    recipient = str(to_email or "").strip()
    if not recipient:
        return False, "Email de destino inválido."
    if not cfg["host"]:
        return False, "SMTP no configurado: falta SMTP_HOST."

    sender = cfg["from_email"] or cfg["username"]
    if not sender:
        return False, "SMTP no configurado: falta SMTP_FROM o SMTP_USER."

    message = EmailMessage()
    message["From"] = sender
    message["To"] = recipient
    message["Subject"] = str(subject or "").strip() or "Notificación"
    message.set_content(str(body_text or "").strip() or "Mensaje automático")

    try:
        with smtplib.SMTP(cfg["host"], cfg["port"], timeout=20) as smtp:
            smtp.ehlo()
            if cfg["use_tls"]:
                smtp.starttls()
                smtp.ehlo()
            if cfg["username"]:
                smtp.login(cfg["username"], cfg["password"])
            smtp.send_message(message)
    except Exception as exc:
        return False, f"No se pudo enviar email: {exc}"
    return True, ""


def _send_password_reset_email(to_email, display_name, reset_url, expires_at_iso):
    name = str(display_name or "usuario").strip()
    subject = "Restablecer contraseña de acceso"
    body = (
        f"Hola {name},\n\n"
        "Recibimos una solicitud para restablecer tu contraseña.\n"
        f"Usá este enlace para crear una nueva clave:\n{reset_url}\n\n"
        f"Este enlace vence el: {expires_at_iso} (UTC).\n\n"
        "Si no solicitaste este cambio, ignorá este mensaje."
    )
    return _send_email(to_email=to_email, subject=subject, body_text=body)


def _default_general_settings():
    return {
        "company_name": BOOTSTRAP_COMPANY_NAME,
        "hr_team_name": BOOTSTRAP_HR_TEAM_NAME,
        "hr_contact": BOOTSTRAP_HR_CONTACT,
        "company_email": "",
        "company_address": "",
        "company_phone": "",
        "company_website": "",
        "handoff_auto_close_enabled": False,
        "handoff_auto_close_minutes": AUTO_CLOSE_DEFAULT_MINUTES,
    }


def _normalize_company_id(value):
    token = chatbot.normalizar_texto(value or "")
    token = token.replace(" ", "-")
    token = "".join(ch for ch in token if ch.isalnum() or ch in {"-", "_", "."})
    token = token.strip("-_.")
    if len(token) < 2:
        return ""
    return token[:64]


def _default_company_id():
    env_value = str(os.getenv("CHATBOT_DEFAULT_COMPANY_ID", "")).strip()
    if env_value:
        normalized = _normalize_company_id(env_value)
        if normalized:
            return normalized
    return _normalize_company_id(_default_general_settings().get("company_name")) or "empresa"


def _normalize_branch_item(item):
    """Normaliza un ítem de sucursal: string -> objeto con name; dict -> objeto con name, address, phone, encargado."""
    if item is None:
        return None
    if isinstance(item, dict):
        name = str(item.get("name") or "").strip()
        if not name:
            return None
        return {
            "name": name,
            "address": _normalize_optional_company_text(item.get("address"), max_len=240),
            "phone": _normalize_optional_company_text(item.get("phone"), max_len=80),
            "encargado": _normalize_optional_company_text(item.get("encargado"), max_len=120),
        }
    branch = str(item or "").strip()
    if not branch:
        return None
    return {
        "name": branch,
        "address": "",
        "phone": "",
        "encargado": "",
    }


def _normalize_branches(raw_branches):
    if raw_branches is None:
        return []
    values = raw_branches if isinstance(raw_branches, list) else [raw_branches]
    cleaned = []
    seen = set()
    for item in values:
        branch = _normalize_branch_item(item)
        if not branch:
            continue
        key = branch["name"].lower()
        if key in seen:
            continue
        cleaned.append(branch)
        seen.add(key)
    return cleaned


def _normalize_areas(raw_areas):
    """Lista de nombres de área (strings) normalizados y sin duplicados."""
    if raw_areas is None:
        return []
    values = raw_areas if isinstance(raw_areas, list) else [raw_areas]
    cleaned = []
    seen = set()
    for item in values:
        name = str(item or "").strip()
        if not name:
            continue
        name = name[:120]
        key = name.lower()
        if key in seen:
            continue
        cleaned.append(name)
        seen.add(key)
    return cleaned


def _normalize_areas_by_branch(raw):
    """Dict sucursal -> lista de áreas. Claves normalizadas (string); '' se guarda como AREA_BRANCH_KEY_EMPRESA por Firestore."""
    if not isinstance(raw, dict):
        return {}
    out = {}
    for branch_key, areas in raw.items():
        branch = str(branch_key or "").strip()[:120]
        key = AREA_BRANCH_KEY_EMPRESA if branch == "" else branch
        out[key] = _normalize_areas(areas)
    return out


def _get_areas_for_branch(company, branch):
    """Devuelve la lista de áreas para una empresa y sucursal. branch '' = ámbito empresa."""
    if not company:
        return []
    branch_key = _branch_name(branch) if isinstance(branch, dict) else (str(branch).strip() if branch is not None else "")
    areas_by_branch = company.get("areas_by_branch")
    if isinstance(areas_by_branch, dict):
        if branch_key in areas_by_branch:
            return list(areas_by_branch[branch_key] or [])
        if "" in areas_by_branch:
            return list(areas_by_branch[""] or [])
        if AREA_BRANCH_KEY_EMPRESA in areas_by_branch:
            return list(areas_by_branch[AREA_BRANCH_KEY_EMPRESA] or [])
    return _normalize_areas(company.get("areas"))


def _areas_by_branch_for_api(areas_by_branch):
    """Convierte areas_by_branch para la API: __e -> '' para que el frontend reciba siempre '' como 'Todas'."""
    if not isinstance(areas_by_branch, dict):
        return areas_by_branch or {}
    out = {}
    for k, v in areas_by_branch.items():
        key = "" if k == AREA_BRANCH_KEY_EMPRESA else k
        out[key] = list(v) if isinstance(v, list) else []
    return out


def _company_for_api(company):
    """Copia la empresa con areas_by_branch con claves aptas para el frontend ('' en lugar de __e)."""
    if not company or not isinstance(company, dict):
        return company
    out = dict(company)
    out["areas_by_branch"] = _areas_by_branch_for_api(company.get("areas_by_branch"))
    return out


def _normalize_optional_company_text(value, max_len=180):
    raw = str(value or "").strip()
    if not raw:
        return ""
    return raw[:max_len]


def _normalize_whatsapp_numbers(raw):
    """Lista de { "phone": str, "label": str } por empresa/área (ej. RRHH, Tesorería)."""
    if not raw:
        return []
    out = []
    seen_phones = set()
    for item in raw if isinstance(raw, list) else []:
        if not isinstance(item, dict):
            continue
        phone = str(item.get("phone") or "").strip()
        if not phone:
            continue
        if not phone.startswith("+"):
            phone = "+" + phone.lstrip("0")
        key = "".join(c for c in phone if c.isdigit())
        if key in seen_phones:
            continue
        seen_phones.add(key)
        label = str(item.get("label") or "").strip() or "Principal"
        notify_email = str(item.get("notify_email") or "").strip()[:200] or None
        out.append({"phone": phone[:30], "label": label[:80], "notify_email": notify_email})
    return out


def _normalize_bool_flag(value, default=False):
    if isinstance(value, bool):
        return value
    raw = str(value if value is not None else "").strip().lower()
    if not raw:
        return bool(default)
    if raw in {"1", "true", "yes", "on", "si", "sí"}:
        return True
    if raw in {"0", "false", "no", "off"}:
        return False
    return bool(default)


def _normalize_auto_close_minutes(value):
    try:
        minutes = int(str(value or 0).strip() or "0")
    except Exception:
        minutes = 0
    if minutes <= 0:
        return 0
    return max(AUTO_CLOSE_MIN_MINUTES, min(minutes, AUTO_CLOSE_MAX_MINUTES))


def _normalize_company_entry(entry):
    if not isinstance(entry, dict):
        return None
    company_name = str(entry.get("company_name") or "").strip()
    if not company_name:
        return None
    company_id = _normalize_company_id(entry.get("company_id") or company_name)
    if not company_id:
        return None
    hr_team_name = str(entry.get("hr_team_name") or "Atención").strip() or "Atención"
    hr_contact = str(entry.get("hr_contact") or "").strip() or "interno 104"
    branches = _normalize_branches(entry.get("branches"))
    company_email = _normalize_optional_company_text(entry.get("company_email"), max_len=200).lower()
    company_address = _normalize_optional_company_text(entry.get("company_address"), max_len=240)
    company_phone = _normalize_optional_company_text(entry.get("company_phone"), max_len=80)
    company_website = _normalize_optional_company_text(entry.get("company_website"), max_len=200)
    auto_close_minutes = _normalize_auto_close_minutes(entry.get("handoff_auto_close_minutes"))
    auto_close_enabled = _normalize_bool_flag(
        entry.get("handoff_auto_close_enabled"),
        default=auto_close_minutes > 0,
    )
    if not auto_close_enabled:
        auto_close_minutes = 0
    permitir_humano = _normalize_bool_flag(entry.get("permitir_hablar_con_humano"), default=True)
    temas_hab_raw = entry.get("temas_habilitados")
    if isinstance(temas_hab_raw, list):
        temas_habilitados = [str(t).strip().lower() for t in temas_hab_raw if str(t).strip()]
    else:
        temas_habilitados = []
    areas = _normalize_areas(entry.get("areas"))
    areas_by_branch = _normalize_areas_by_branch(entry.get("areas_by_branch"))
    if not areas_by_branch and areas:
        areas_by_branch = {AREA_BRANCH_KEY_EMPRESA: areas}
    whatsapp_numbers = _normalize_whatsapp_numbers(entry.get("whatsapp_numbers"))
    return {
        "company_id": company_id,
        "company_name": company_name,
        "hr_team_name": hr_team_name,
        "hr_contact": hr_contact,
        "company_email": company_email,
        "company_address": company_address,
        "company_phone": company_phone,
        "company_website": company_website,
        "handoff_auto_close_enabled": auto_close_enabled,
        "handoff_auto_close_minutes": auto_close_minutes,
        "branches": branches,
        "areas": areas,
        "areas_by_branch": areas_by_branch,
        "active": bool(entry.get("active", True)),
        "permitir_hablar_con_humano": permitir_humano,
        "temas_habilitados": temas_habilitados,
        "whatsapp_numbers": whatsapp_numbers,
        "drive_folder_id": str(entry.get("drive_folder_id") or "").strip()[:128] or None,
        "handoff_notify_email": str(entry.get("handoff_notify_email") or "").strip()[:200] or None,
    }
    if out.get("drive_folder_id") is None:
        out.pop("drive_folder_id", None)


def _default_company_entry():
    defaults = _default_general_settings()
    return {
        "company_id": _default_company_id(),
        "company_name": defaults.get("company_name") or "Empresa",
        "hr_team_name": defaults.get("hr_team_name") or "Atención",
        "hr_contact": defaults.get("hr_contact") or "interno 104",
        "company_email": "",
        "company_address": "",
        "company_phone": "",
        "company_website": "",
        "handoff_auto_close_enabled": False,
        "handoff_auto_close_minutes": AUTO_CLOSE_DEFAULT_MINUTES,
        "branches": [],
        "areas": [],
        "areas_by_branch": {},
        "active": True,
        "permitir_hablar_con_humano": True,
        "temas_habilitados": [],
        "whatsapp_numbers": [],
        "drive_folder_id": None,
    }


def _company_by_whatsapp_phone(phone):
    """Si el número recibido está configurado en alguna empresa, devuelve (company_id, company, label)."""
    if not phone:
        return None, None, None
    norm = _normalize_phone_for_match(phone)
    if not norm:
        return None, None, None
    for company in _list_companies(include_inactive=False):
        for line in (company.get("whatsapp_numbers") or []):
            p = (line.get("phone") or "").strip()
            if _normalize_phone_for_match(p) == norm:
                return company.get("company_id"), company, (line.get("label") or "").strip() or None
    return None, None, None


def _merge_company_entries(current, candidate):
    base = dict(current or {})
    extra = dict(candidate or {})
    for key in ("company_name", "hr_team_name", "hr_contact"):
        value = str(extra.get(key) or "").strip()
        if value:
            base[key] = value
    for key in ("company_email", "company_address", "company_phone", "company_website"):
        value = _normalize_optional_company_text(extra.get(key), max_len=240)
        if value:
            base[key] = value.lower() if key == "company_email" else value
    base_minutes = _normalize_auto_close_minutes(
        extra.get("handoff_auto_close_minutes", base.get("handoff_auto_close_minutes"))
    )
    base_enabled = _normalize_bool_flag(
        extra.get("handoff_auto_close_enabled", base.get("handoff_auto_close_enabled")),
        default=base_minutes > 0,
    )
    if not base_enabled:
        base_minutes = 0
    base["handoff_auto_close_enabled"] = base_enabled
    base["handoff_auto_close_minutes"] = base_minutes
    base["branches"] = _normalize_branches(
        list(base.get("branches") or []) + list(extra.get("branches") or [])
    )
    if "areas" in extra and isinstance(extra.get("areas"), list):
        base["areas"] = _normalize_areas(extra.get("areas"))
    else:
        base.setdefault("areas", [])
    if "areas_by_branch" in extra and isinstance(extra.get("areas_by_branch"), dict):
        base["areas_by_branch"] = _normalize_areas_by_branch(extra.get("areas_by_branch"))
    else:
        base.setdefault("areas_by_branch", base.get("areas_by_branch") or {})
    base["active"] = bool(base.get("active", True) or extra.get("active", True))
    base["company_id"] = str(base.get("company_id") or extra.get("company_id") or "").strip()
    if "permitir_hablar_con_humano" in extra:
        base["permitir_hablar_con_humano"] = _normalize_bool_flag(extra.get("permitir_hablar_con_humano"), default=True)
    else:
        base.setdefault("permitir_hablar_con_humano", True)
    if "temas_habilitados" in extra and isinstance(extra.get("temas_habilitados"), list):
        base["temas_habilitados"] = [str(t).strip().lower() for t in extra["temas_habilitados"] if str(t).strip()]
    else:
        base.setdefault("temas_habilitados", [])
    if "whatsapp_numbers" in extra and isinstance(extra.get("whatsapp_numbers"), list):
        base["whatsapp_numbers"] = _normalize_whatsapp_numbers(extra["whatsapp_numbers"])
    else:
        base.setdefault("whatsapp_numbers", [])
    if "drive_folder_id" in extra:
        v = str(extra.get("drive_folder_id") or "").strip()[:128] or None
        base["drive_folder_id"] = v
    else:
        base.setdefault("drive_folder_id", base.get("drive_folder_id"))
    if "handoff_notify_email" in extra:
        v = str(extra.get("handoff_notify_email") or "").strip()[:200] or None
        base["handoff_notify_email"] = v
    else:
        base.setdefault("handoff_notify_email", base.get("handoff_notify_email"))
    return base


def _list_companies(include_inactive=False):
    cache_key = "companies_all" if include_inactive else "companies_active"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    rows = []
    if chatbot.db:
        try:
            for doc in chatbot.db.collection(COMPANIES_COLLECTION).stream():
                payload = doc.to_dict() or {}
                payload.setdefault("company_id", doc.id)
                normalized = _normalize_company_entry(payload)
                if normalized:
                    rows.append(normalized)
        except Exception:
            rows = []
    else:
        rows = []
        for company_id, payload in IN_MEMORY_COMPANIES.items():
            clone = dict(payload or {})
            clone.setdefault("company_id", company_id)
            normalized = _normalize_company_entry(clone)
            if normalized:
                rows.append(normalized)

    # Deduplica por company_id para evitar selectores con empresas repetidas.
    deduped = {}
    for item in rows:
        key = str(item.get("company_id") or "").strip()
        if not key:
            continue
        if key in deduped:
            deduped[key] = _merge_company_entries(deduped[key], item)
        else:
            deduped[key] = item
    rows = list(deduped.values())

    if not rows:
        rows = [_default_company_entry()]

    default_entry = _default_company_entry()
    if not any(item.get("company_id") == default_entry.get("company_id") for item in rows):
        rows.append(default_entry)

    rows.sort(key=lambda item: (str(item.get("company_name") or "").lower(), item.get("company_id")))

    if include_inactive:
        # Al cargar todo, también pre-cargamos el subconjunto activo para evitar segunda lectura
        _cache_set("companies_all", rows)
        active_rows = [item for item in rows if item.get("active", True)]
        if not active_rows:
            active_rows = [default_entry]
        _cache_set("companies_active", active_rows)
        return rows
    else:
        rows = [item for item in rows if item.get("active", True)]
        if not rows:
            rows = [default_entry]
        _cache_set("companies_active", rows)
        return rows


def _company_map(include_inactive=False):
    return {item["company_id"]: item for item in _list_companies(include_inactive=include_inactive)}


def _get_company(company_id, include_inactive=True):
    key = _normalize_company_id(company_id)
    if not key:
        return None
    return _company_map(include_inactive=include_inactive).get(key)


def _upsert_company(payload):
    normalized = _normalize_company_entry(payload)
    if not normalized:
        return False, None, "Empresa inválida. Completá nombre y datos básicos."

    if chatbot.db:
        try:
            (
                chatbot.db.collection(COMPANIES_COLLECTION)
                .document(normalized["company_id"])
                .set(normalized, merge=False)
            )
            return True, normalized, ""
        except Exception as exc:
            return False, None, f"No pude guardar empresa: {exc}"

    IN_MEMORY_COMPANIES[normalized["company_id"]] = normalized
    return True, normalized, ""


def _delete_company(company_id):
    key = _normalize_company_id(company_id)
    if not key:
        return False, "Empresa inválida."
    if key == _default_company_id():
        return False, "No podés eliminar la empresa por defecto."

    companies = _list_companies(include_inactive=True)
    if len(companies) <= 1:
        return False, "Debe quedar al menos una empresa configurada."

    for user in auth_rrhh.get_users().values():
        assignments = user.get("assignments") if isinstance(user.get("assignments"), list) else []
        if any(str(item.get("company_id") or "").strip().lower() == key for item in assignments if isinstance(item, dict)):
            return False, "No podés eliminar una empresa asignada a usuarios."

    if chatbot.db:
        try:
            chatbot.db.collection(COMPANIES_COLLECTION).document(key).delete()
            return True, ""
        except Exception as exc:
            return False, f"No pude eliminar empresa: {exc}"

    if key not in IN_MEMORY_COMPANIES:
        return False, "Empresa no encontrada."
    IN_MEMORY_COMPANIES.pop(key, None)
    return True, ""


def _set_company_session(company_id):
    company = _get_company(company_id, include_inactive=False)
    if company is None:
        company = _list_companies(include_inactive=False)[0]
    current_user = _current_rrhh_user()
    if current_user and not _user_can_access_company(current_user, company.get("company_id")):
        allowed = _companies_for_user(current_user)
        if allowed:
            company = allowed[0]
    _sess()["company_id"] = company["company_id"]
    _sess()["company_name"] = company["company_name"]
    _sess()["company_hr_team_name"] = company.get("hr_team_name") or "Atención"
    return company


def _current_company():
    selected = _normalize_company_id(_sess().get("company_id"))
    company = _get_company(selected, include_inactive=False) if selected else None
    if company is None:
        if _list_companies(include_inactive=False):
            company = _list_companies(include_inactive=False)[0]
            _sess()["company_id"] = company["company_id"]
    current_user = _current_rrhh_user()
    if current_user and not _user_can_access_company(current_user, company.get("company_id")):
        allowed = _companies_for_user(current_user)
        if allowed:
            company = allowed[0]
            _sess()["company_id"] = company["company_id"]
    return company or _default_company_entry()


def _sanitize_general_settings(payload):
    defaults = _default_general_settings()
    data = dict(defaults)
    if not isinstance(payload, dict):
        return data

    for key in defaults:
        value = str(payload.get(key) or "").strip()
        if value:
            data[key] = value
    return data


def _read_general_settings():
    company = _current_company()
    auto_close_enabled = _normalize_bool_flag(
        company.get("handoff_auto_close_enabled", False),
        default=False,
    )
    auto_close_minutes = _normalize_auto_close_minutes(company.get("handoff_auto_close_minutes"))
    if not auto_close_enabled:
        auto_close_minutes = 0
    return {
        "company_id": company.get("company_id"),
        "company_name": company.get("company_name"),
        "hr_team_name": company.get("hr_team_name"),
        "hr_contact": company.get("hr_contact"),
        "company_email": company.get("company_email") or "",
        "company_address": company.get("company_address") or "",
        "company_phone": company.get("company_phone") or "",
        "company_website": company.get("company_website") or "",
        "handoff_auto_close_enabled": auto_close_enabled,
        "handoff_auto_close_minutes": auto_close_minutes,
        "branches": list(company.get("branches") or []),
        "permitir_hablar_con_humano": _normalize_bool_flag(company.get("permitir_hablar_con_humano"), default=True),
        "temas_habilitados": list(company.get("temas_habilitados") or []),
    }


def _write_general_settings(payload):
    current = _current_company()
    data = payload if isinstance(payload, dict) else {}
    company_id = _normalize_company_id((payload or {}).get("company_id") or current.get("company_id"))
    company_name = str((payload or {}).get("company_name") or current.get("company_name") or "").strip()
    hr_team_name = str((payload or {}).get("hr_team_name") or current.get("hr_team_name") or "").strip()
    hr_contact = str((payload or {}).get("hr_contact") or current.get("hr_contact") or "").strip()
    company_email_raw = (
        data.get("company_email")
        if "company_email" in data
        else current.get("company_email") or ""
    )
    company_email = _normalize_optional_company_text(
        company_email_raw,
        max_len=200,
    ).lower()
    company_address_raw = (
        data.get("company_address")
        if "company_address" in data
        else current.get("company_address") or ""
    )
    company_address = _normalize_optional_company_text(
        company_address_raw,
        max_len=240,
    )
    company_phone_raw = (
        data.get("company_phone")
        if "company_phone" in data
        else current.get("company_phone") or ""
    )
    company_phone = _normalize_optional_company_text(
        company_phone_raw,
        max_len=80,
    )
    company_website_raw = (
        data.get("company_website")
        if "company_website" in data
        else current.get("company_website") or ""
    )
    company_website = _normalize_optional_company_text(
        company_website_raw,
        max_len=200,
    )
    auto_close_enabled_raw = (
        data.get("handoff_auto_close_enabled")
        if "handoff_auto_close_enabled" in data
        else current.get("handoff_auto_close_enabled", False)
    )
    auto_close_enabled = _normalize_bool_flag(
        auto_close_enabled_raw,
        default=bool(current.get("handoff_auto_close_enabled", False)),
    )
    auto_close_minutes_raw = (
        data.get("handoff_auto_close_minutes")
        if "handoff_auto_close_minutes" in data
        else current.get("handoff_auto_close_minutes", AUTO_CLOSE_DEFAULT_MINUTES)
    )
    auto_close_minutes = _normalize_auto_close_minutes(auto_close_minutes_raw)
    if not auto_close_enabled:
        auto_close_minutes = 0
    branches = _normalize_branches((payload or {}).get("branches") or current.get("branches") or [])

    permitir_humano = _normalize_bool_flag(
        data.get("permitir_hablar_con_humano")
        if "permitir_hablar_con_humano" in data
        else current.get("permitir_hablar_con_humano", True),
        default=True,
    )
    temas_hab_raw = data.get("temas_habilitados") if "temas_habilitados" in data else current.get("temas_habilitados")
    temas_habilitados = (
        [str(t).strip().lower() for t in temas_hab_raw if str(t).strip()]
        if isinstance(temas_hab_raw, list)
        else []
    )
    ok, company, error = _upsert_company(
        {
            "company_id": company_id or current.get("company_id"),
            "company_name": company_name or current.get("company_name"),
            "hr_team_name": hr_team_name or current.get("hr_team_name"),
            "hr_contact": hr_contact or current.get("hr_contact"),
            "company_email": company_email,
            "company_address": company_address,
            "company_phone": company_phone,
            "company_website": company_website,
            "handoff_auto_close_enabled": auto_close_enabled,
            "handoff_auto_close_minutes": auto_close_minutes,
            "branches": branches,
            "areas": _normalize_areas(data.get("areas") if "areas" in data else (current.get("areas") or [])),
            "active": True,
            "permitir_hablar_con_humano": permitir_humano,
            "temas_habilitados": temas_habilitados,
        }
    )
    if not ok:
        return False, None, error
    _set_company_session(company.get("company_id"))
    return True, _read_general_settings(), ""


def _apply_company_branding(settings=None):
    cfg = settings or _read_general_settings()
    if hasattr(chatbot, "actualizar_configuracion_empresa"):
        chatbot.actualizar_configuracion_empresa(
            company_name=cfg.get("company_name"),
            hr_team_name=cfg.get("hr_team_name"),
            hr_contact=cfg.get("hr_contact"),
        )
    return cfg


def _farewell_message():
    cfg = _apply_company_branding()
    return (
        f"Gracias por comunicarte con {cfg.get('hr_team_name', 'Atención')} "
        f"de {cfg.get('company_name', 'la empresa')}. ¡Buen día!"
    )


def _manual_agent_id(name):
    normalized = chatbot.normalizar_texto(name)[:48]
    if not normalized:
        normalized = f"agente-{uuid.uuid4().hex[:8]}"
    return f"manual:{normalized}"


def _agent_payload(display_name, agent_id="", role="rrhh", company_id="", area=""):
    shown = str(display_name or "Agente").strip() or "Agente"
    identifier = str(agent_id or "").strip().lower() or _manual_agent_id(shown)
    company_key = _normalize_company_id(company_id)
    area_str = str(area or "").strip()
    return {
        "agent_id": identifier,
        "display_name": shown,
        "role": str(role or "rrhh").strip().lower() or "rrhh",
        "company_id": company_key,
        "area": area_str,
    }


def _active_agent_from_current_user():
    current = _current_rrhh_user()
    if not current:
        return None
    company_id = _selected_company_id_for_rrhh()
    return _agent_payload(
        display_name=current.get("display_name") or current.get("username") or "Agente",
        agent_id=str(current.get("username") or "").strip().lower(),
        role=current.get("role") or "rrhh",
        company_id=company_id,
        area=current.get("area") or "",
    )


def _auth_enabled():
    return auth_rrhh.is_auth_enabled()


def _safe_next_path(next_path, fallback="/rrhh"):
    raw = str(next_path or "").strip()
    if not raw:
        return fallback
    if not raw.startswith("/") or raw.startswith("//"):
        return fallback
    return raw


def _request_path_with_query():
    path = request.path
    if request.query_string:
        path = f"{path}?{request.query_string.decode('utf-8', errors='ignore')}"
    return _safe_next_path(path)


def _current_rrhh_user():
    username = str(session.get("rrhh_user") or "").strip()
    if not username:
        return None
    # Verificar timeout de inactividad
    last_active = session.get("rrhh_last_active")
    if last_active:
        elapsed = datetime.now(timezone.utc).timestamp() - last_active
        if elapsed > RRHH_SESSION_TIMEOUT_SECONDS:
            _clear_rrhh_user()
            return None
    session["rrhh_last_active"] = datetime.now(timezone.utc).timestamp()
    current = {
        "username": username,
        "display_name": str(session.get("rrhh_display_name") or username),
        "role": str(session.get("rrhh_role") or "rrhh"),
        "assignments": session.get("rrhh_assignments") or [],
        "area": str(session.get("rrhh_area") or "").strip(),
    }
    if _auth_enabled():
        # Si el rol/nombre cambian en archivo, sincroniza sesión en caliente.
        entry = auth_rrhh.get_users().get(username.lower())
        if not entry:
            _clear_rrhh_user()
            return None
        current["display_name"] = str(entry.get("display_name") or username)
        current["role"] = str(entry.get("role") or "rrhh")
        current["assignments"] = list(entry.get("assignments") or [])
        current["area"] = str(entry.get("area") or "").strip()
        _set_rrhh_user(current)
    return current


def _set_rrhh_user(user_payload):
    session["rrhh_user"] = str(user_payload.get("username") or "").strip()
    session["rrhh_display_name"] = str(
        user_payload.get("display_name") or session["rrhh_user"]
    ).strip()
    session["rrhh_role"] = str(user_payload.get("role") or "rrhh").strip()
    session["rrhh_assignments"] = list(user_payload.get("assignments") or [])
    session["rrhh_area"] = str(user_payload.get("area") or "").strip()
    session["rrhh_last_active"] = datetime.now(timezone.utc).timestamp()


def _clear_rrhh_user():
    session.pop("rrhh_user", None)
    session.pop("rrhh_display_name", None)
    session.pop("rrhh_role", None)
    session.pop("rrhh_assignments", None)
    session.pop("rrhh_area", None)
    session.pop("rrhh_last_active", None)


def _rrhh_agent_name(default="Agente"):
    current = _current_rrhh_user()
    if not current:
        return default
    return str(current.get("display_name") or current.get("username") or default)


def _resolve_rrhh_agent(payload):
    if _auth_enabled():
        current_agent = _active_agent_from_current_user()
        if current_agent:
            return current_agent
    fallback_name = str((payload or {}).get("agente") or "Agente").strip() or "Agente"
    return _agent_payload(display_name=fallback_name, company_id=_selected_company_id_for_rrhh())


def _upsert_active_agent(agent, source="heartbeat"):
    payload = _agent_payload(
        display_name=agent.get("display_name") or "Agente",
        agent_id=agent.get("agent_id") or "",
        role=agent.get("role") or "rrhh",
        company_id=agent.get("company_id") or _selected_company_id_for_rrhh(),
        area=agent.get("area") or "",
    )
    payload["updated_at"] = _utc_now()
    payload["available"] = True
    payload["source"] = str(source or "heartbeat")

    if chatbot.db:
        try:
            (
                chatbot.db.collection(ACTIVE_AGENTS_COLLECTION)
                .document(payload["agent_id"])
                .set(payload, merge=True)
            )
        except Exception:
            return payload
        return payload

    IN_MEMORY_ACTIVE_AGENTS[payload["agent_id"]] = payload
    return payload


def _list_active_agents(ttl_seconds=ACTIVE_AGENT_TTL_SECONDS, company_id=None, branch=None):
    now = _as_utc_naive(_utc_now()) or datetime.utcnow()
    ttl = max(30, int(ttl_seconds or ACTIVE_AGENT_TTL_SECONDS))
    threshold = now - timedelta(seconds=ttl)

    if chatbot.db:
        rows = []
        try:
            for doc in chatbot.db.collection(ACTIVE_AGENTS_COLLECTION).stream():
                data = doc.to_dict() or {}
                data["agent_id"] = str(data.get("agent_id") or doc.id)
                rows.append(data)
        except Exception:
            rows = []
    else:
        rows = [dict(value) for value in IN_MEMORY_ACTIVE_AGENTS.values()]

    active = []
    target_company = _normalize_company_id(company_id)
    branch_str = str(branch or "").strip() if branch is not None else None
    users_map = auth_rrhh.get_users() if _auth_enabled() else {}

    for item in rows:
        updated_at = _as_utc_naive(item.get("updated_at"))
        if updated_at is None or updated_at < threshold:
            continue
        if item.get("available") is False:
            continue
        item_company = _normalize_company_id(item.get("company_id"))
        if target_company and item_company != target_company:
            continue
        if branch_str is not None and users_map:
            agent_id_raw = str(item.get("agent_id") or "").strip().lower()
            user_entry = users_map.get(agent_id_raw)
            if user_entry and not auth_rrhh.assignment_matches_company_branch(
                user_entry.get("assignments"), target_company, branch_str
            ):
                continue
        role = str(item.get("role") or "").strip().lower()
        if role and not auth_rrhh.role_has_permission(role, auth_rrhh.PERM_CONVERSATIONS_MANAGE):
            continue
        agent_id_raw = str(item.get("agent_id") or "").strip().lower()
        area_stored = str(item.get("area") or "").strip()
        if not area_stored and users_map and agent_id_raw:
            user_entry = users_map.get(agent_id_raw)
            area_stored = str((user_entry or {}).get("area") or "").strip()
        active.append(
            _agent_payload(
                item.get("display_name"),
                item.get("agent_id"),
                role,
                company_id=item_company,
                area=area_stored,
            )
        )
        active[-1]["updated_at"] = item.get("updated_at")

    active.sort(key=lambda x: (str(x.get("display_name") or "").lower(), x.get("agent_id") or ""))
    return active


def _serialize_active_agent(agent):
    return {
        "agent_id": str(agent.get("agent_id") or ""),
        "display_name": str(agent.get("display_name") or ""),
        "role": str(agent.get("role") or ""),
        "company_id": str(agent.get("company_id") or ""),
        "updated_at": _fmt_fecha(agent.get("updated_at")),
    }


def _heartbeat_current_agent(source="rrhh_panel"):
    if _auth_enabled():
        current = _active_agent_from_current_user()
        if not current:
            return None
        if not auth_rrhh.role_has_permission(
            current.get("role"), auth_rrhh.PERM_CONVERSATIONS_MANAGE
        ):
            return None
        agent = _upsert_active_agent(current, source=source)
        _auto_assign_pending_handoffs(agent)
        return agent
    return None


def _auto_assign_pending_handoffs(agent):
    """Al conectarse un agente, asignarle handoffs pendientes sin asignar de su empresa."""
    if not agent:
        return
    agent_company = _normalize_company_id(agent.get("company_id"))
    pending = _list_handoffs(include_closed=False, company_id=agent_company or None)
    now = _utc_now()
    for conv in pending:
        if str(conv.get("estado") or "").strip().lower() != HANDOFF_STATUS_PENDING:
            continue
        if str(conv.get("rrhh_agente_id") or "").strip():
            continue  # ya tiene agente
        conv_id = conv.get("conversation_id")
        if not conv_id:
            continue
        _take_handoff(conv_id, agent, auto_taken=True)
        _upsert_handoff(conv_id, {"updated_at": now}, merge=True)


def _open_handoff_load_by_agent(company_id=None):
    counts = {}
    for conv in _list_handoffs(include_closed=False, limit=1000, company_id=company_id):
        estado = str(conv.get("estado") or "").strip().lower()
        if estado == HANDOFF_STATUS_CLOSED:
            continue
        agent_id = str(conv.get("rrhh_agente_id") or "").strip().lower()
        if not agent_id:
            continue
        counts[agent_id] = counts.get(agent_id, 0) + 1
    return counts


def _agent_areas_set(agent):
    """Conjunto de áreas del agente (normalizadas) para matching."""
    area_str = str(agent.get("area") or "").strip()
    if not area_str:
        return set()
    return {a.strip().lower() for a in area_str.split(",") if a.strip()}


def _select_auto_agent(company_id=None, branch=None, area=None):
    # 1. Búsqueda exacta: empresa + sucursal
    active_agents = _list_active_agents(company_id=company_id, branch=branch or None)
    # 2. Si no hay, relajar sucursal
    if not active_agents and branch:
        active_agents = _list_active_agents(company_id=company_id, branch=None)
    # 3. Si no hay, cualquier agente activo (sin filtro de empresa)
    if not active_agents:
        active_agents = _list_active_agents(company_id=None, branch=None)
    if not active_agents:
        return None
    area_norm = str(area or "").strip().lower() if area else None
    if area_norm:
        by_area = [a for a in active_agents if area_norm in _agent_areas_set(a)]
        if by_area:
            active_agents = by_area
    load = _open_handoff_load_by_agent(company_id=company_id)
    return sorted(
        active_agents,
        key=lambda agent: (
            load.get(agent.get("agent_id"), 0),
            str(agent.get("display_name") or "").lower(),
            str(agent.get("agent_id") or ""),
        ),
    )[0]


def _resolve_target_agent_for_reassignment(payload):
    data = payload or {}
    requested_id = str(data.get("agente_id") or "").strip().lower()
    requested_name = str(data.get("agente") or "").strip()

    if requested_id:
        for agent in _list_active_agents(company_id=_selected_company_id_for_rrhh()):
            if agent.get("agent_id") == requested_id:
                return agent
        return None

    if requested_name:
        return _agent_payload(
            display_name=requested_name,
            company_id=_selected_company_id_for_rrhh(),
        )

    return _resolve_rrhh_agent(data)


def _can_manage_configuration():
    """Acceso a la página Configuración (al menos un módulo habilitado)."""
    if not _auth_enabled():
        return True
    return (
        _has_permission(auth_rrhh.PERM_CONFIG_MANAGE)
        or _has_permission(auth_rrhh.PERM_USERS_MANAGE)
        or _has_permission(auth_rrhh.PERM_ROLES_MANAGE)
        or _has_permission(auth_rrhh.PERM_CONFIG_EMPRESAS)
        or _has_permission(auth_rrhh.PERM_CONFIG_SUCURSALES)
        or _has_permission(auth_rrhh.PERM_CONFIG_AREAS)
        or _has_permission(auth_rrhh.PERM_CONFIG_KNOWLEDGE)
        or _has_permission(auth_rrhh.PERM_CONFIG_CONVENIOS)
        or _has_permission(auth_rrhh.PERM_CONFIG_SMTP)
    )


def _can_manage_general_config():
    """Backward-compat: True si tiene el permiso global de configuración."""
    if not _auth_enabled():
        return True
    return _has_permission(auth_rrhh.PERM_CONFIG_MANAGE)


def _can_config_empresas():
    if not _auth_enabled():
        return True
    return _has_permission(auth_rrhh.PERM_CONFIG_MANAGE) or _has_permission(auth_rrhh.PERM_CONFIG_EMPRESAS)


def _can_config_sucursales():
    if not _auth_enabled():
        return True
    return _has_permission(auth_rrhh.PERM_CONFIG_MANAGE) or _has_permission(auth_rrhh.PERM_CONFIG_SUCURSALES)


def _can_config_areas():
    if not _auth_enabled():
        return True
    return _has_permission(auth_rrhh.PERM_CONFIG_MANAGE) or _has_permission(auth_rrhh.PERM_CONFIG_AREAS)


def _can_config_knowledge():
    if not _auth_enabled():
        return True
    return _has_permission(auth_rrhh.PERM_CONFIG_MANAGE) or _has_permission(auth_rrhh.PERM_CONFIG_KNOWLEDGE)


def _can_config_convenios():
    if not _auth_enabled():
        return True
    return _has_permission(auth_rrhh.PERM_CONFIG_MANAGE) or _has_permission(auth_rrhh.PERM_CONFIG_CONVENIOS)


def _can_config_smtp():
    if not _auth_enabled():
        return True
    return _has_permission(auth_rrhh.PERM_CONFIG_MANAGE) or _has_permission(auth_rrhh.PERM_CONFIG_SMTP)


def _list_companies_for_current_rrhh_user(include_inactive=True):
    """Devuelve las empresas filtradas por las asignaciones del usuario logueado."""
    companies_raw = _list_companies(include_inactive=include_inactive)
    cur_user = _current_rrhh_user()
    if cur_user:
        assignments = [
            item for item in (cur_user.get("assignments") or [])
            if isinstance(item, dict) and item.get("company_id")
        ]
        if assignments:
            allowed_ids = {str(item["company_id"]).strip().lower() for item in assignments}
            companies_raw = [c for c in companies_raw if c.get("company_id") in allowed_ids]
    return companies_raw


def _companies_for_user(user_payload):
    companies = _list_companies(include_inactive=False)
    assignments = list((user_payload or {}).get("assignments") or [])
    if not assignments:
        return companies
    allowed = {
        str(item.get("company_id") or "").strip().lower()
        for item in assignments
        if isinstance(item, dict)
    }
    return [item for item in companies if item.get("company_id") in allowed]


def _user_can_access_company(user_payload, company_id):
    company_key = _normalize_company_id(company_id)
    if not company_key:
        return False
    username = str((user_payload or {}).get("username") or "").strip()
    if username:
        return auth_rrhh.user_has_company_access(username, company_key)
    assignments = list((user_payload or {}).get("assignments") or [])
    if not assignments:
        return True
    return any(
        str(item.get("company_id") or "").strip().lower() == company_key
        for item in assignments
        if isinstance(item, dict)
    )


def _selected_company_id_for_rrhh():
    current = _current_company()
    return current.get("company_id")


def _default_landing_for_user(user_payload):
    role = str((user_payload or {}).get("role") or "")
    if auth_rrhh.role_has_permission(role, auth_rrhh.PERM_CONVERSATIONS_VIEW):
        return "/rrhh"
    if auth_rrhh.role_has_permission(role, auth_rrhh.PERM_HISTORY_VIEW):
        return "/historial"
    if auth_rrhh.role_has_permission(role, auth_rrhh.PERM_LEGAJOS_VIEW):
        return "/legajos"
    if auth_rrhh.role_has_permission(role, auth_rrhh.PERM_USERS_MANAGE) or auth_rrhh.role_has_permission(
        role, auth_rrhh.PERM_ROLES_MANAGE
    ):
        return "/configuracion"
    return "/"


def _has_permission(permission):
    if not _auth_enabled():
        return True
    current = _current_rrhh_user()
    if not current:
        return False
    entry = auth_rrhh.get_users().get((current.get("username") or "").lower()) or current
    selected_company = _selected_company_id_for_rrhh() or ""
    effective_role = auth_rrhh.get_role_for_context(entry, selected_company, branch=None)
    if not effective_role:
        effective_role = current.get("role") or "rrhh"
    return auth_rrhh.role_has_permission(effective_role, permission)


def _is_admin():
    if not _auth_enabled():
        return True
    current = _current_rrhh_user()
    if not current:
        return False
    entry = auth_rrhh.get_users().get((current.get("username") or "").lower()) or current
    selected_company = _selected_company_id_for_rrhh() or ""
    effective_role = auth_rrhh.get_role_for_context(entry, selected_company, branch=None)
    if not effective_role:
        effective_role = current.get("role") or "rrhh"
    return str(effective_role).strip().lower() == "admin"


def _auth_json_error():
    return jsonify({"ok": False, "error": "No autorizado"}), 401


def _forbidden_json_error(message="No tenés permisos para esta acción."):
    return jsonify({"ok": False, "error": message}), 403


def rrhh_auth_required(handler):
    @wraps(handler)
    def wrapped(*args, **kwargs):
        if not _auth_enabled():
            return handler(*args, **kwargs)
        if _current_rrhh_user() is not None:
            return handler(*args, **kwargs)
        if request.path.startswith("/api/"):
            return _auth_json_error()
        return redirect(url_for("login_page", next=_request_path_with_query()))

    return wrapped


def rrhh_permission_required(permission, message="No tenés permisos para esta acción."):
    def decorator(handler):
        @wraps(handler)
        def wrapped(*args, **kwargs):
            if not _auth_enabled():
                return handler(*args, **kwargs)
            if _current_rrhh_user() is None:
                if request.path.startswith("/api/"):
                    return _auth_json_error()
                return redirect(url_for("login_page", next=_request_path_with_query()))
            if not _has_permission(permission):
                if request.path.startswith("/api/"):
                    return _forbidden_json_error(message)
                return (message, 403)
            return handler(*args, **kwargs)

        return wrapped

    return decorator


def _session_chat_id():
    chat_id = _sess().get("chat_session_id")
    if chat_id:
        return chat_id
    chat_id = _new_conversation_id()
    _sess()["chat_session_id"] = chat_id
    return chat_id


def _new_history_id():
    return f"hist-{uuid.uuid4().hex[:14]}"


def _add_chat_history(
    conversation_id,
    remitente,
    texto,
    canal="asistente",
    agente="",
    metadata=None,
):
    payload = {
        "history_id": _new_history_id(),
        "conversation_id": str(conversation_id or "sin_conversacion"),
        "remitente": str(remitente or "desconocido"),
        "texto": str(texto or "").strip(),
        "canal": str(canal or "asistente"),
        "agente": str(agente or "").strip(),
        "fecha": _utc_now(),
        "metadata": metadata or {},
    }
    if not payload["texto"]:
        return

    if chatbot.db:
        chatbot.db.collection("chat_historial").add(payload)
        _cache_del("chat_history_all")
        return

    IN_MEMORY_CHAT_HISTORY.append(payload)


def _list_chat_history(limit=300):
    all_rows = _cache_get("chat_history_all")
    if all_rows is None:
        if chatbot.db:
            all_rows = []
            for doc in chatbot.db.collection("chat_historial").stream():
                data = doc.to_dict() or {}
                data["id"] = doc.id
                all_rows.append(data)
        else:
            all_rows = list(IN_MEMORY_CHAT_HISTORY)
        all_rows.sort(
            key=lambda x: _as_utc_naive(x.get("fecha")) or datetime.min,
            reverse=True,
        )
        _cache_set("chat_history_all", all_rows)
    return all_rows[:limit]


def _serialize_history_item(item):
    fecha = _as_utc_aware(item.get("fecha"))
    return {
        "id": str(item.get("id") or item.get("history_id") or ""),
        "conversation_id": str(item.get("conversation_id") or ""),
        "remitente": str(item.get("remitente") or ""),
        "canal": str(item.get("canal") or ""),
        "agente": str(item.get("agente") or ""),
        "texto": str(item.get("texto") or ""),
        "fecha": _fmt_fecha(fecha),
        "fecha_iso": _iso_utc(fecha),
        "metadata": item.get("metadata") or {},
    }


def _set_handoff_session(conversation_id):
    _sess()["handoff_conversation_id"] = conversation_id
    _sess()["last_rrhh_seen_iso"] = ""


def _clear_handoff_session():
    _sess().pop("handoff_conversation_id", None)
    _sess().pop("last_rrhh_seen_iso", None)


def _get_handoff_session_id():
    return _sess().get("handoff_conversation_id")


def _get_last_seen_rrhh():
    raw = _sess().get("last_rrhh_seen_iso", "")
    if not raw:
        return None
    try:
        parsed = datetime.fromisoformat(raw)
        return _as_utc_naive(parsed)
    except Exception:
        return None


def _set_last_seen_rrhh(dt):
    dt2 = _as_utc_naive(dt)
    session["last_rrhh_seen_iso"] = dt2.isoformat() if dt2 else ""


def _from_firestore_doc(snapshot):
    data = snapshot.to_dict() or {}
    data["id"] = snapshot.id
    return data


def _normalize_phone_for_match(phone):
    """Normaliza número para comparar (whatsapp:+549... o +549... -> dígitos)."""
    p = (phone or "").strip().lower()
    if not p:
        return ""
    if p.startswith("whatsapp:"):
        p = p[9:].strip()
    return "".join(c for c in p if c.isdigit())


def _find_open_handoff_by_whatsapp_phone(phone):
    """Devuelve el conversation_id de un handoff abierto para este número de WhatsApp, o None."""
    if not phone:
        return None
    norm = _normalize_phone_for_match(phone)
    if not norm:
        return None
    for conv in _list_handoffs(include_closed=False, limit=50):
        to_phone = (conv.get("whatsapp_to_phone") or "").strip()
        if not to_phone:
            continue
        if _normalize_phone_for_match(to_phone) == norm:
            return conv.get("id")
    return None


def _fetch_handoff(conversation_id):
    if chatbot.db:
        doc = chatbot.db.collection("rrhh_handoffs").document(conversation_id).get()
        if not doc.exists:
            return None
        data = doc.to_dict() or {}
        data["id"] = conversation_id
        return data

    return IN_MEMORY_HANDOFFS.get(conversation_id)


def _upsert_handoff(conversation_id, payload, merge=True):
    if chatbot.db:
        doc_ref = chatbot.db.collection("rrhh_handoffs").document(conversation_id)
        doc_ref.set(payload, merge=merge)
        _cache_del("handoffs_all")
        return

    existing = IN_MEMORY_HANDOFFS.get(conversation_id, {})
    if merge:
        existing.update(payload)
        IN_MEMORY_HANDOFFS[conversation_id] = existing
    else:
        IN_MEMORY_HANDOFFS[conversation_id] = dict(payload)
    IN_MEMORY_HANDOFFS[conversation_id]["id"] = conversation_id
    IN_MEMORY_HANDOFFS[conversation_id].setdefault("mensajes", [])


def _branch_name(item):
    if isinstance(item, dict) and item.get("name") is not None:
        return str(item.get("name") or "").strip()
    return str(item or "").strip()


def _list_handoffs(include_closed=False, limit=100, company_id=None, branches=None, areas=None):
    docs = _cache_get("handoffs_all")
    if docs is None:
        if chatbot.db:
            docs = [
                _from_firestore_doc(doc)
                for doc in chatbot.db.collection("rrhh_handoffs").stream()
            ]
        else:
            docs = [dict(value) for value in IN_MEMORY_HANDOFFS.values()]
        _cache_set("handoffs_all", docs)

    filtered = []
    target_company = _normalize_company_id(company_id)
    branch_set = None
    if branches is not None and len(branches) > 0:
        branch_set = {_branch_name(b).lower() for b in branches if _branch_name(b)}
    area_set = None
    if areas is not None and len(areas) > 0:
        area_set = {str(a).strip().lower() for a in areas if str(a).strip()}
    for item in docs:
        estado = str(item.get("estado") or "").strip().lower()
        if not include_closed and estado == HANDOFF_STATUS_CLOSED:
            continue
        if target_company:
            item_company = _normalize_company_id(item.get("company_id"))
            if item_company != target_company:
                continue
        if branch_set is not None:
            conv_branch = str(item.get("branch") or "").strip().lower()
            if conv_branch and conv_branch not in branch_set:
                continue
        if area_set is not None:
            conv_area = str(item.get("area") or "").strip().lower()
            if conv_area and conv_area not in area_set:
                continue
        filtered.append(item)

    filtered.sort(key=lambda x: _as_utc_naive(x.get("updated_at")) or datetime.min, reverse=True)
    return filtered[:limit]


def _all_handoff_records_for_stats():
    if chatbot.db:
        docs = _cache_get("handoffs_all")
        if docs is None:
            docs = [_from_firestore_doc(doc) for doc in chatbot.db.collection("rrhh_handoffs").stream()]
            _cache_set("handoffs_all", docs)
        return docs
    return [dict(value) for value in IN_MEMORY_HANDOFFS.values()]


def _companies_for_filter_context():
    """Lista de empresas con branches y areas para filtros (estadísticas, historial)."""
    current = _current_rrhh_user()
    companies = (
        _companies_for_user(current)
        if current
        else _list_companies(include_inactive=False)
    )
    out = []
    for c in companies:
        cid = c.get("company_id")
        if not cid:
            continue
        branches = []
        for b in (c.get("branches") or []):
            name = _branch_name(b)
            if name:
                branches.append({"name": name})
        areas = _get_all_areas_for_company(c)
        areas_list = [str(a).strip() for a in (areas or []) if str(a).strip()]
        out.append({
            "company_id": cid,
            "company_name": (c.get("company_name") or cid).strip(),
            "branches": branches,
            "areas": areas_list,
        })
    return out


def _send_whatsapp_to_collaborator(to_phone, text, from_number=None, media_url=None):
    """Envía texto y/o media al colaborador vía Meta WhatsApp Cloud API.
    Para media: descarga el archivo desde la URL de Storage, lo sube a Meta y lo envía.
    Siempre usa el META_PHONE_NUMBER_ID configurado como phone_number_id,
    ignorando from_number (que puede ser un número Twilio de handoffs viejos).
    """
    if not to_phone:
        return False
    # Usar siempre el Meta phone_number_id configurado en env vars
    pid = _meta_phone_number_id() or None

    media_list = media_url if isinstance(media_url, list) else ([media_url] if media_url else [])
    # "(archivo adjunto)" es un marcador interno, no un mensaje real para el colaborador
    _raw_caption = str(text or "").strip()
    caption_pending = "" if _raw_caption == "(archivo adjunto)" else _raw_caption

    for url in media_list:
        if not url or not isinstance(url, str) or not url.startswith("http"):
            continue
        try:
            import requests as _req
            from urllib.parse import urlparse
            r = _req.get(url, timeout=30)
            if not r.ok:
                logger.warning("_send_whatsapp_to_collaborator: no se pudo descargar %s", url)
                continue
            file_bytes = r.content
            content_type = r.headers.get("content-type", "").split(";")[0].strip()
            parsed = urlparse(url)
            filename = parsed.path.split("/")[-1].split("?")[0] or "archivo"
            ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
            if not content_type or content_type in ("application/octet-stream", "binary/octet-stream"):
                content_type = _EXT_TO_MIME.get(ext, "application/octet-stream")
            # Meta no acepta text/csv — convertir a XLSX antes de subir
            if ext == "csv" or content_type == "text/csv":
                try:
                    import openpyxl as _openpyxl
                    from io import BytesIO as _BytesIO, StringIO as _StringIO
                    import csv as _csv_mod
                    _text = file_bytes.decode("utf-8-sig", errors="replace")
                    try:
                        _dialect = _csv_mod.Sniffer().sniff(_text[:2048])
                    except Exception:
                        _dialect = "excel"
                    _reader = _csv_mod.reader(_StringIO(_text), _dialect)
                    _wb = _openpyxl.Workbook()
                    _ws = _wb.active
                    for _row in _reader:
                        _ws.append(_row)
                    _buf = _BytesIO()
                    _wb.save(_buf)
                    file_bytes = _buf.getvalue()
                    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    filename = filename.rsplit(".", 1)[0] + ".xlsx"
                    ext = "xlsx"
                    logger.info("_send_whatsapp_to_collaborator: CSV convertido a XLSX (%s)", filename)
                except Exception as _conv_err:
                    logger.warning("_send_whatsapp_to_collaborator csv->xlsx: %s", _conv_err)
            media_id = _upload_media_to_meta(file_bytes, content_type, filename, phone_number_id=pid)
            if not media_id:
                logger.warning("_send_whatsapp_to_collaborator: no se pudo subir media a Meta")
                continue
            is_image = content_type.startswith("image/")
            media_type = "image" if is_image else "document"
            _send_meta_whatsapp_media(
                to_phone, media_id, media_type,
                caption=caption_pending or None,
                filename=filename if not is_image else None,
                phone_number_id=pid,
            )
            caption_pending = ""  # ya enviado como caption del archivo
        except Exception as e:
            logger.warning("_send_whatsapp_to_collaborator media: %s", e)

    # Enviar texto restante si no fue incluido como caption
    if caption_pending:
        _send_meta_whatsapp(to_phone, caption_pending, phone_number_id=pid)
    return True


def _add_handoff_message(
    conversation_id,
    remitente,
    texto,
    agente="",
    visible_to_colaborador=True,
    media_url=None,
):
    now = _utc_now()
    texto_str = str(texto).strip()
    media_list = []
    if media_url is not None:
        def _valid_media_url(u):
            s = str(u).strip()
            return s.startswith("http") or s.startswith("meta_media://")
        if isinstance(media_url, (list, tuple)):
            media_list = [str(u).strip() for u in media_url if u and _valid_media_url(u)]
        elif isinstance(media_url, str) and _valid_media_url(media_url):
            media_list = [media_url.strip()]
    payload = {
        "remitente": remitente,
        "texto": texto_str,
        "agente": str(agente or "").strip(),
        "fecha": now,
        "visible_to_colaborador": bool(visible_to_colaborador),
    }
    if media_list:
        payload["media_url"] = media_list

    if chatbot.db:
        (
            chatbot.db.collection("rrhh_handoffs")
            .document(conversation_id)
            .collection("mensajes")
            .add(payload)
        )
    else:
        conv = IN_MEMORY_HANDOFFS.setdefault(conversation_id, {"id": conversation_id})
        conv.setdefault("mensajes", []).append(payload)

    upsert_extra = {
        "updated_at": now,
        "ultimo_mensaje": payload["texto"],
        "ultimo_remitente": str(remitente or "").strip().lower(),
        "ultimo_mensaje_fecha": now,
    }
    _upsert_handoff(conversation_id, upsert_extra, merge=True)
    # Incrementar contador de mensajes del colaborador para notificaciones por mensaje
    if str(remitente or "").strip().lower() == "colaborador" and chatbot.db:
        try:
            from firebase_admin import firestore as _fs_admin
            chatbot.db.collection("rrhh_handoffs").document(conversation_id).update(
                {"colaborador_mensajes_count": _fs_admin.Increment(1)}
            )
        except Exception:
            pass
    try:
        _hist_company_id = _normalize_company_id(_sess().get("company_id"))
    except Exception:
        _hist_company_id = ""
    _add_chat_history(
        conversation_id=conversation_id,
        remitente=remitente,
        texto=payload["texto"],
        canal="rrhh",
        agente=payload["agente"],
        metadata={
            "visible_to_colaborador": payload["visible_to_colaborador"],
            "company_id": _hist_company_id or None,
        },
    )
    # Enviar por WhatsApp al colaborador si la conversación es por WA: mensajes del agente y cierre de conversación
    if payload["visible_to_colaborador"] and remitente in ("rrhh", "sistema"):
        has_text = bool(payload.get("texto"))
        has_media = bool(payload.get("media_url"))
        if has_text or has_media:
            conv = _fetch_handoff(conversation_id)
            to_phone = (conv or {}).get("whatsapp_to_phone", "").strip()
            from_number = (conv or {}).get("whatsapp_from_number", "").strip()
            if to_phone:
                _send_whatsapp_to_collaborator(
                    to_phone,
                    text=payload.get("texto") or None,
                    from_number=from_number or None,
                    media_url=payload.get("media_url") or None,
                )


def _list_handoff_messages(conversation_id, limit=300):
    if chatbot.db:
        docs = (
            chatbot.db.collection("rrhh_handoffs")
            .document(conversation_id)
            .collection("mensajes")
            .stream()
        )
        rows = []
        for doc in docs:
            data = doc.to_dict() or {}
            data["id"] = doc.id
            rows.append(data)
    else:
        conv = IN_MEMORY_HANDOFFS.get(conversation_id, {})
        rows = list(conv.get("mensajes", []))

    rows.sort(key=lambda x: _as_utc_naive(x.get("fecha")) or datetime.min)
    return rows[-limit:]


def _join_messages_for_user(messages):
    if not messages:
        return ""
    lineas = []
    for msg in messages:
        remitente = str(msg.get("remitente") or "").strip().lower()
        texto = str(msg.get("texto") or "").strip()
        agente = str(msg.get("agente") or "").strip()
        fecha = _fmt_fecha(msg.get("fecha"))
        if remitente == "rrhh":
            prefijo = f"👩‍💼 Agente{f' ({agente})' if agente else ''}"
        elif remitente == "sistema":
            prefijo = "ℹ️ Sistema"
        else:
            prefijo = "📩"
        lineas.append(f"{prefijo} [{fecha}]: {texto}")
    return "\n".join(lineas)


def _take_handoff(conversation_id, agente, auto_taken=False):
    conv = _fetch_handoff(conversation_id)
    if not conv:
        return False
    agent = _agent_payload(
        display_name=(agente or {}).get("display_name") if isinstance(agente, dict) else agente,
        agent_id=(agente or {}).get("agent_id") if isinstance(agente, dict) else "",
        role=(agente or {}).get("role") if isinstance(agente, dict) else "rrhh",
        company_id=(
            (agente or {}).get("company_id")
            if isinstance(agente, dict)
            else conv.get("company_id")
        ),
    )

    # Evita mensajes de sistema duplicados cuando la conversación ya estaba tomada
    # por el mismo agente.
    estado_actual = str(conv.get("estado") or "").strip().lower()
    agente_actual_id = str(conv.get("rrhh_agente_id") or "").strip().lower()
    agente_actual_nombre = str(conv.get("rrhh_agente") or "").strip().lower()
    if estado_actual == HANDOFF_STATUS_ACTIVE and (
        (agent.get("agent_id") and agent.get("agent_id") == agente_actual_id)
        or (
            not agent.get("agent_id")
            and agent.get("display_name")
            and agent.get("display_name").strip().lower() == agente_actual_nombre
        )
    ):
        _upsert_handoff(
            conversation_id,
            {"updated_at": _utc_now()},
            merge=True,
        )
        return True

    _upsert_handoff(
        conversation_id,
        {
            "estado": HANDOFF_STATUS_ACTIVE,
            "rrhh_agente": agent["display_name"],
            "rrhh_agente_id": agent["agent_id"],
            "updated_at": _utc_now(),
            "rrhh_asignacion_automatica": bool(auto_taken),
        },
        merge=True,
    )
    if auto_taken:
        _add_handoff_message(
            conversation_id,
            remitente="sistema",
            texto=f"Tu conversación fue asignada automáticamente a un agente ({agent['display_name']}).",
            visible_to_colaborador=False,
        )
    else:
        _add_handoff_message(
            conversation_id,
            remitente="sistema",
            texto=f"Tu conversación fue tomada por un agente ({agent['display_name']}).",
        )
    return True


def _close_handoff(conversation_id, quien):
    conv = _fetch_handoff(conversation_id)
    if not conv:
        return False
    _upsert_handoff(
        conversation_id,
        {
            "estado": HANDOFF_STATUS_CLOSED,
            "updated_at": _utc_now(),
        },
        merge=True,
    )
    _add_handoff_message(
        conversation_id,
        remitente="sistema",
        texto=f"La conversación fue cerrada por {quien}.",
    )
    # Si era conversación de WhatsApp, notificar al colaborador y resetear sesión
    if conv.get("channel") == "whatsapp":
        wa_phone = conv.get("whatsapp_to_phone")
        if wa_phone:
            from_number = conv.get("whatsapp_from_number", "").strip()
            _send_whatsapp_to_collaborator(
                wa_phone,
                "Tu consulta fue cerrada. Si tenés una nueva pregunta, escribinos cuando quieras. 👋",
                from_number=from_number or None,
            )
            _reset_whatsapp_chat_context(wa_phone)
    return True


def _reopen_handoff(conversation_id, quien):
    """Reabre una conversación cerrada (estado -> pendiente)."""
    conv = _fetch_handoff(conversation_id)
    if not conv:
        return False
    estado = str(conv.get("estado") or "").strip().lower()
    if estado != HANDOFF_STATUS_CLOSED:
        return True  # ya está abierta
    _upsert_handoff(
        conversation_id,
        {
            "estado": HANDOFF_STATUS_PENDING,
            "updated_at": _utc_now(),
        },
        merge=True,
    )
    _add_handoff_message(
        conversation_id,
        remitente="sistema",
        texto=f"Conversación reabierta por {quien}.",
    )
    # Notificar al colaborador por WhatsApp
    if conv.get("channel") == "whatsapp":
        wa_phone = conv.get("whatsapp_to_phone", "").strip()
        from_number = conv.get("whatsapp_from_number", "").strip()
        if wa_phone:
            _send_whatsapp_to_collaborator(
                wa_phone,
                "Tu consulta fue reabierta. Un agente se va a comunicar con vos a la brevedad. 🙌",
                from_number=from_number or None,
            )
    return True


def _handoff_auto_close_policy(company=None):
    current = company if isinstance(company, dict) else _current_company()
    enabled = _normalize_bool_flag(current.get("handoff_auto_close_enabled"), default=False)
    minutes = _normalize_auto_close_minutes(
        current.get("handoff_auto_close_minutes", AUTO_CLOSE_DEFAULT_MINUTES)
    )
    if not enabled or minutes <= 0:
        return False, 0
    return True, minutes


def _handoff_last_activity(conv):
    if not isinstance(conv, dict):
        return None
    return (
        _as_utc_naive(conv.get("ultimo_mensaje_fecha"))
        or _as_utc_naive(conv.get("updated_at"))
        or _as_utc_naive(conv.get("created_at"))
    )


def _auto_close_expired_handoffs(company=None, force=False, override_minutes=None):
    target_company = company if isinstance(company, dict) else _current_company()
    company_id = _normalize_company_id(target_company.get("company_id"))
    if not company_id:
        return {
            "company_id": "",
            "auto_close_enabled": False,
            "auto_close_minutes": 0,
            "closed_count": 0,
            "checked_count": 0,
        }

    enabled, minutes = _handoff_auto_close_policy(target_company)
    if override_minutes is not None:
        minutes = _normalize_auto_close_minutes(override_minutes)
        enabled = minutes > 0
    if not enabled:
        return {
            "company_id": company_id,
            "auto_close_enabled": False,
            "auto_close_minutes": 0,
            "closed_count": 0,
            "checked_count": 0,
        }

    now = _as_utc_naive(_utc_now()) or datetime.utcnow()
    conversations = _list_handoffs(include_closed=False, limit=1000, company_id=company_id)
    closed_count = 0
    checked_count = 0
    for conv in conversations:
        estado = str(conv.get("estado") or "").strip().lower()
        if estado not in {HANDOFF_STATUS_PENDING, HANDOFF_STATUS_ACTIVE}:
            continue
        checked_count += 1
        conversation_id = str(conv.get("conversation_id") or conv.get("id") or "").strip()
        if not conversation_id:
            continue
        last_activity = _handoff_last_activity(conv)
        if last_activity is None:
            continue
        idle_minutes = (now - last_activity).total_seconds() / 60.0
        if idle_minutes < minutes:
            continue
        if _close_handoff(
            conversation_id,
            f"sistema por inactividad ({minutes} min)",
        ):
            closed_count += 1

    return {
        "company_id": company_id,
        "auto_close_enabled": True,
        "auto_close_minutes": minutes,
        "closed_count": closed_count,
        "checked_count": checked_count,
        "forced": bool(force),
    }


def _reassign_handoff(conversation_id, agente_destino, reasignado_por=""):
    conv = _fetch_handoff(conversation_id)
    if not conv:
        return False, "Conversación no encontrada"
    target = _agent_payload(
        display_name=(agente_destino or {}).get("display_name")
        if isinstance(agente_destino, dict)
        else agente_destino,
        agent_id=(agente_destino or {}).get("agent_id")
        if isinstance(agente_destino, dict)
        else "",
        role=(agente_destino or {}).get("role")
        if isinstance(agente_destino, dict)
        else "rrhh",
        company_id=(
            (agente_destino or {}).get("company_id")
            if isinstance(agente_destino, dict)
            else conv.get("company_id")
        ),
    )
    if not target.get("display_name"):
        return False, "Agente destino inválido"

    _upsert_handoff(
        conversation_id,
        {
            "estado": HANDOFF_STATUS_ACTIVE,
            "rrhh_agente": target["display_name"],
            "rrhh_agente_id": target["agent_id"],
            "updated_at": _utc_now(),
            "rrhh_asignacion_automatica": False,
        },
        merge=True,
    )
    actor = str(reasignado_por or "").strip() or "sistema"
    _add_handoff_message(
        conversation_id,
        remitente="sistema",
        texto=f"Conversación reasignada a {target['display_name']} por {actor}.",
        visible_to_colaborador=False,
    )
    return True, ""


def _collect_new_messages_for_collaborator(conversation_id):
    last_seen = _get_last_seen_rrhh()
    messages = _list_handoff_messages(conversation_id)
    nuevos = []
    max_seen = last_seen
    for msg in messages:
        remitente = str(msg.get("remitente") or "").strip().lower()
        if remitente not in {"rrhh", "sistema"}:
            continue
        if remitente == "sistema" and msg.get("visible_to_colaborador", True) is False:
            continue
        fecha = _as_utc_naive(msg.get("fecha"))
        if last_seen is not None and fecha is not None and fecha <= last_seen:
            continue
        nuevos.append(msg)
        if fecha is not None and (max_seen is None or fecha > max_seen):
            max_seen = fecha

    if max_seen is not None:
        _set_last_seen_rrhh(max_seen)
    return nuevos


def construir_temas_map(company_id=None, temas_habilitados=None):
    """
    Construye el mapa número -> tema para el menú.
    Si temas_habilitados es una lista no vacía, esos temas son el menú (sin filtrar contra faqs).
    Si no hay temas_habilitados, se usan los temas de la colección faqs o el fallback.
    """
    if company_id is None:
        company_id = (_current_company() or {}).get("company_id")
    # Prioridad: temas_habilitados configurados por la empresa (KB auto-detectados o manuales)
    if temas_habilitados and isinstance(temas_habilitados, list):
        temas = sorted(temas_habilitados, key=chatbot.normalizar_texto)
        return {str(i): tema for i, tema in enumerate(temas, start=1)}
    # Fallback: temas desde colección faqs (legacy) o FAQ_FALLBACK
    temas = chatbot.obtener_temas_desde_firestore(company_id=company_id)
    if not temas:
        temas = sorted(chatbot.FAQ_FALLBACK.keys(), key=chatbot.normalizar_texto)
    return {str(i): tema for i, tema in enumerate(temas, start=1)}


# Límite de temas mostrados en el menú (formato: 1. Tema, 2. Tema, ... H. Hablar con un agente)
MENU_TEMAS_LIMITE = 12


def construir_menu_texto(temas_map, permitir_hablar_con_humano=True):
    """Texto del menú en formato: Menú de temas disponibles: / 1. X / 2. Y / ... / H. Hablar con un agente."""
    lineas = ["Menú de temas disponibles:"]
    for numero, tema in temas_map.items():
        lineas.append(f"{numero}. {tema.capitalize()}")
    if permitir_hablar_con_humano:
        lineas.append("H. Hablar con un agente")
    return "\n".join(lineas)


def construir_acciones_menu(temas_map, limite=None, permitir_hablar_con_humano=True):
    """Botones del menú: numerados (1. Tema, 2. Tema, ...), opcional Ver menú completo, y H. Hablar con un agente."""
    if limite is None:
        limite = MENU_TEMAS_LIMITE
    acciones = []
    items = list(temas_map.items())
    for numero, tema in items[:limite]:
        acciones.append(_accion(f"{numero}. {tema.capitalize()}", numero, "topic"))

    if len(temas_map) > limite:
        acciones.append(_accion("Ver menú completo", "menu", "secondary"))

    if permitir_hablar_con_humano:
        acciones.append(_accion("H. Hablar con un agente", "h", "secondary"))
    return acciones


def construir_acciones_feedback():
    return [
        _accion("Sí", "si", "positive"),
        _accion("No", "no", "negative"),
        _accion("Nueva consulta", "menu", "secondary"),
    ]


def construir_acciones_sugerencias(consulta, temas_map, permitir_hablar_con_humano=True):
    sugerencias = chatbot.sugerir_temas(consulta, temas_map)
    acciones = []
    for tema in sugerencias[:4]:
        acciones.append(_accion(tema.capitalize(), tema, "topic"))
    acciones.append(_accion("Ver menú", "menu", "secondary"))
    if permitir_hablar_con_humano:
        acciones.append(_accion("H. Hablar con un agente", "h", "secondary"))
    return acciones


def construir_acciones_handoff():
    return [
        _accion("Finalizar chat", "__cerrar_rrhh__", "negative"),
    ]


def armar_respuesta_no_entendida(consulta, temas_map):
    if temas_map:
        temas_lista = "\n".join(f"{num}. {tema.capitalize()}" for num, tema in temas_map.items())
        return (
            "No encontré información sobre eso.\n\n"
            "Podés consultar sobre estos temas:\n"
            f"{temas_lista}\n\n"
            "Escribí el número del tema o hablá con un agente si ninguno se adapta."
        )
    return (
        "Todavía no tengo información cargada para esta empresa. "
        "Podés hablar con un agente para que te ayuden."
    )


def limpiar_estado_conversacion():
    _sess().pop("pending_feedback_topic", None)
    _sess().pop("pending_derivacion", None)


def _payload(
    reply,
    await_feedback=False,
    end_session=False,
    quick_actions=None,
    handoff_active=False,
):
    return {
        "reply": reply,
        "await_feedback": await_feedback,
        "end_session": end_session,
        "quick_actions": quick_actions or [],
        "handoff_active": handoff_active,
    }


def _firebase_project_id():
    if chatbot.db is None:
        return "modo_local_sin_firestore"
    return str(getattr(chatbot.db, "project", "desconocido"))


_RESUMEN_IGNORAR = {
    "si", "sí", "no", "ok", "dale", "bueno", "gracias", "menu", "menú",
    "hola", "h", "1", "2", "3", "4", "5", "6", "7", "8", "9", "0",
    "quiero hablar con alguien", "hablar con un agente", "hablar con alguien",
}

def _generar_resumen_conversacion(chat_session_id, mensaje_trigger=None):
    """Genera un resumen interpretativo de los temas consultados para mostrárselo al agente."""
    turnos = []
    try:
        if chatbot.db:
            docs = (
                chatbot.db.collection("chat_historial")
                .where("conversation_id", "==", str(chat_session_id or ""))
                .get()
            )
            all_turnos = [doc.to_dict() or {} for doc in docs]
            all_turnos.sort(key=lambda x: _as_utc_naive(x.get("fecha")) or datetime.min)
            turnos = all_turnos
        else:
            turnos = sorted(
                [h for h in IN_MEMORY_CHAT_HISTORY
                 if str(h.get("conversation_id") or "") == str(chat_session_id or "")],
                key=lambda x: _as_utc_naive(x.get("fecha")) or datetime.min,
            )
    except Exception as _exc:
        logging.warning("_generar_resumen: error consultando historial session=%s: %s", chat_session_id, _exc)
        turnos = []

    logging.info("_generar_resumen: session_id=%r turnos_encontrados=%d", chat_session_id, len(turnos))

    if not turnos:
        return "Sin historial previo disponible."

    # Extraer consultas reales del colaborador (ignorar respuestas cortas/triviales)
    consultas = []
    for t in turnos:
        remitente = str(t.get("remitente") or "").strip().lower()
        texto = str(t.get("texto") or "").strip()
        if not texto:
            continue
        if remitente == "colaborador":
            norm = chatbot.normalizar_texto(texto)
            if norm not in _RESUMEN_IGNORAR and len(texto) > 3:
                consultas.append(texto)

    logging.info("_generar_resumen: session_id=%r consultas_filtradas=%d remitentes=%s",
                 chat_session_id, len(consultas),
                 list({str(t.get("remitente") or "") for t in turnos}))

    if not consultas:
        # Fallback: usar el mensaje que disparó el handoff si aporta información
        if mensaje_trigger:
            _mt = str(mensaje_trigger).strip()
            _mt_norm = chatbot.normalizar_texto(_mt)
            if _mt_norm not in _RESUMEN_IGNORAR and len(_mt) > 3:
                return f"📋 Resumen de la consulta:\nÚltimo mensaje: \"{_mt}\""
        return "El colaborador solicitó hablar con un agente sin realizar consultas previas."

    # Deduplicar manteniendo orden
    vistas = set()
    consultas_unicas = []
    for c in consultas:
        norm = chatbot.normalizar_texto(c)
        if norm not in vistas:
            vistas.add(norm)
            consultas_unicas.append(c)

    # Tomar solo las últimas 5 consultas relevantes
    consultas_unicas = consultas_unicas[-5:]

    partes = ["📋 Resumen de la consulta:"]
    partes.append(f"Temas consultados ({len(consultas_unicas)}): " + " | ".join(f'"{c}"' for c in consultas_unicas))

    return "\n".join(partes)


def _iniciar_handoff_rrhh(mensaje_usuario):
    chat_session_id = _session_chat_id()
    company = _current_company()
    company_id = company.get("company_id")
    active_id = _get_handoff_session_id()
    existing = _fetch_handoff(active_id) if active_id else None
    now = _utc_now()

    if existing is not None:
        estado_actual = str(existing.get("estado") or "").strip().lower()
        existing_company = _normalize_company_id(existing.get("company_id"))
        if estado_actual in {HANDOFF_STATUS_PENDING, HANDOFF_STATUS_ACTIVE} and existing_company == company_id:
            conversation_id = active_id
        else:
            _clear_handoff_session()
            conversation_id = _new_conversation_id()
            existing = None
    else:
        conversation_id = _new_conversation_id()

    if existing is None:
        collaborator_area = _sess().get("chat_context_area") or ""
        collaborator_branch = _sess().get("chat_context_branch") or ""
        assigned_agent = _select_auto_agent(
            company_id=company_id,
            branch=collaborator_branch or None,
            area=collaborator_area or None,
        )
        if not assigned_agent:
            assigned_agent = _select_auto_agent(company_id=company_id, branch=None, area=None)
        estado_inicial = HANDOFF_STATUS_ACTIVE if assigned_agent else HANDOFF_STATUS_PENDING
        handoff_payload = {
            "conversation_id": conversation_id,
            "company_id": company_id,
            "company_name": company.get("company_name"),
            "branch": collaborator_branch,
            "area": collaborator_area,
            "estado": estado_inicial,
            "created_at": now,
            "updated_at": now,
            "rrhh_agente": assigned_agent.get("display_name", "") if assigned_agent else "",
            "rrhh_agente_id": assigned_agent.get("agent_id", "") if assigned_agent else "",
            "rrhh_asignacion_automatica": bool(assigned_agent),
            "ultimo_mensaje": "",
            "ultima_consulta": mensaje_usuario.strip() or "Solicitud de derivación",
            "chat_session_id": chat_session_id,
        }
        if getattr(g, "whatsapp_from", None) and getattr(g, "whatsapp_to", None):
            wa_phone_raw = g.whatsapp_from
            wa_profile = getattr(g, "whatsapp_profile_name", "") or ""
            collab_nombre, collab_tel = _resolve_whatsapp_contact(
                wa_phone_raw, company_id, wa_profile
            )
            handoff_payload["channel"] = "whatsapp"
            handoff_payload["whatsapp_to_phone"] = wa_phone_raw
            handoff_payload["whatsapp_from_number"] = g.whatsapp_to
            handoff_payload["colaborador_nombre"] = collab_nombre
            handoff_payload["colaborador_telefono"] = collab_tel
        _upsert_handoff(conversation_id, handoff_payload, merge=False)
        _send_push_all(
            f"Nueva conversación{' · ' + handoff_payload.get('company_name') if handoff_payload.get('company_name') else ''}",
            f"{handoff_payload.get('colaborador_nombre') or handoff_payload.get('colaborador_telefono') or 'Usuario'}: {handoff_payload.get('ultima_consulta', 'Solicitud de atención')[:150]}",
        )
        # Generar resumen antes de notificar para incluirlo en el email
        resumen = _generar_resumen_conversacion(chat_session_id, mensaje_trigger=mensaje_usuario)
        handoff_payload["resumen_conversacion"] = resumen
        _notify_handoff_via_n8n(handoff_payload, company)
        _add_handoff_message(
            conversation_id,
            remitente="sistema",
            texto=resumen,
            visible_to_colaborador=False,
        )
        if assigned_agent:
            _take_handoff(conversation_id, assigned_agent, auto_taken=True)

    if mensaje_usuario.strip():
        _add_handoff_message(
            conversation_id,
            remitente="colaborador",
            texto=mensaje_usuario,
        )
        update_payload = {
            "ultima_consulta": mensaje_usuario.strip(),
            "updated_at": now,
            "chat_session_id": chat_session_id,
        }
        if getattr(g, "whatsapp_from", None):
            wa_profile = getattr(g, "whatsapp_profile_name", "") or ""
            collab_nombre, collab_tel = _resolve_whatsapp_contact(
                g.whatsapp_from, company_id, wa_profile
            )
            update_payload["colaborador_nombre"] = collab_nombre
            update_payload["colaborador_telefono"] = collab_tel
        _upsert_handoff(conversation_id, update_payload, merge=True)

    _set_handoff_session(conversation_id)
    return conversation_id


def procesar_feedback_pendiente(texto_usuario, tema_pendiente, temas_map, permitir_hablar_con_humano=True, company_id=None):
    tipo, texto_norm = chatbot.clasificar_input_feedback(texto_usuario)

    if tipo == "feedback":
        # Si el tema es del knowledge base y el usuario dijo "no" → mostrar menú y ofrecer derivación
        if texto_norm == "no" and tema_pendiente == "knowledge_answer":
            chatbot.registrar_feedback(tema_pendiente, texto_norm, company_id=company_id)
            limpiar_estado_conversacion()
            hr_name = (_current_company() or {}).get("hr_team_name") or "Atención"
            # Si el mismo mensaje ya pide hablar con alguien → derivar directamente sin preguntar
            texto_usuario_norm = chatbot.normalizar_texto(texto_usuario)
            if permitir_hablar_con_humano and chatbot.solicita_contacto_rrhh(texto_usuario_norm):
                conversation_id = _iniciar_handoff_rrhh(texto_usuario)
                conv = _fetch_handoff(conversation_id) or {}
                assigned = str(conv.get("rrhh_agente") or "").strip()
                if assigned:
                    respuesta_derivacion = (
                        f"Entendido. Te derivé con el equipo de atención.\n"
                        f"Te asigné con {assigned}. Podés seguir escribiendo por este chat."
                    )
                else:
                    respuesta_derivacion = (
                        f"Entendido. Derivé tu consulta al equipo de {hr_name}.\n"
                        "Te van a responder por este mismo chat."
                    )
                return _payload(respuesta_derivacion, handoff_active=True, quick_actions=construir_acciones_handoff())
            menu_texto = construir_menu_texto(temas_map, permitir_hablar_con_humano=False)
            respuesta_no = (
                f"Entendido. Podés consultar sobre estos temas:\n\n{menu_texto}"
            )
            if permitir_hablar_con_humano:
                _sess()["pending_derivacion"] = True
                respuesta_no += f"\n\n¿O preferís que te derive con un agente de {hr_name}?"
            acciones = construir_acciones_menu(temas_map, limite=6, permitir_hablar_con_humano=False)
            if permitir_hablar_con_humano:
                acciones.append({"label": f"Sí, hablar con {hr_name}", "value": "si"})
                acciones.append({"label": "No, gracias", "value": "no"})
            return _payload(respuesta_no, await_feedback=False, quick_actions=acciones)
        guardado = chatbot.registrar_feedback(tema_pendiente, texto_norm, company_id=company_id)
        limpiar_estado_conversacion()
        if guardado:
            texto = (
                "¡Gracias por tu feedback! 🙌\n"
                "Si querés, escribime otra consulta o poné 'menu' para ver temas."
            )
        else:
            texto = (
                "⚠️ Recibí tu feedback, pero no pude guardarlo en la base de datos.\n"
                "Podés seguir usando el chat y revisar tu conexión Firebase."
            )
        return _payload(
            texto,
            quick_actions=construir_acciones_menu(temas_map, limite=6, permitir_hablar_con_humano=permitir_hablar_con_humano),
        )

    if tipo == "menu":
        limpiar_estado_conversacion()
        return _payload(
            construir_menu_texto(temas_map, permitir_hablar_con_humano=permitir_hablar_con_humano),
            quick_actions=construir_acciones_menu(temas_map, permitir_hablar_con_humano=permitir_hablar_con_humano),
        )

    if tipo == "salir":
        limpiar_estado_conversacion()
        return _payload(_farewell_message(), end_session=True)

    if tipo == "consulta":
        limpiar_estado_conversacion()
        return None

    return _payload(
        "Podés responder 'si' o 'no'. Si preferís, también podés escribir una nueva consulta.",
        await_feedback=True,
        quick_actions=construir_acciones_feedback(),
    )


def responder_chat(mensaje_usuario):
    company = _current_company()
    company_id = (company or {}).get("company_id")
    permitir = (company or {}).get("permitir_hablar_con_humano", True)
    temas_habilitados = (company or {}).get("temas_habilitados") or []
    temas_map = construir_temas_map(company_id=company_id, temas_habilitados=temas_habilitados)
    mensaje_norm = chatbot.normalizar_texto(mensaje_usuario)

    def _acciones_menu(limite=8):
        return construir_acciones_menu(temas_map, limite=limite, permitir_hablar_con_humano=permitir)

    if not mensaje_norm:
        return _payload(
            "No llegué a entender tu consulta. ¿Podés reformularla?",
            quick_actions=_acciones_menu(6),
        )

    if mensaje_norm in chatbot.PALABRAS_SALIDA:
        handoff_id = _get_handoff_session_id()
        if handoff_id:
            _close_handoff(handoff_id, "colaborador")
            _clear_handoff_session()
        return _payload(_farewell_message(), end_session=True)

    handoff_id = _get_handoff_session_id()
    if handoff_id:
        _auto_close_expired_handoffs(company=_current_company())
        conv = _fetch_handoff(handoff_id)
        if conv and str(conv.get("estado") or "").strip().lower() == HANDOFF_STATUS_CLOSED:
            _clear_handoff_session()
            handoff_id = None

    if handoff_id:
        if mensaje_norm in HANDOFF_END_COMMANDS:
            _close_handoff(handoff_id, "colaborador")
            _clear_handoff_session()
            return _payload(
                "✅ Cerré la conversación con el agente. Si querés, seguimos con el asistente virtual.",
                quick_actions=_acciones_menu(6),
            )

        # Solo "menu" cierra el handoff y vuelve al bot; saludos y consultas van al agente.
        if mensaje_norm == "menu":
            _close_handoff(handoff_id, "colaborador")
            _clear_handoff_session()
            return _payload(
                construir_menu_texto(temas_map, permitir_hablar_con_humano=permitir),
                quick_actions=construir_acciones_menu(temas_map, permitir_hablar_con_humano=permitir),
            )

        if mensaje_norm in HANDOFF_POLL_COMMANDS:
            nuevos = _collect_new_messages_for_collaborator(handoff_id)
            if nuevos:
                return _payload(
                    _join_messages_for_user(nuevos),
                    handoff_active=True,
                    quick_actions=construir_acciones_handoff(),
                )
            conv = _fetch_handoff(handoff_id) or {}
            estado = str(conv.get("estado") or "").strip().lower()
            if estado == HANDOFF_STATUS_PENDING:
                texto = "⏳ Tu consulta sigue en cola. Un agente te responde en breve."
            else:
                agente = conv.get("rrhh_agente") or "el equipo"
                texto = f"🟢 Chat activo con {agente}. Aún no hay nuevos mensajes."
            return _payload(
                texto,
                handoff_active=True,
                quick_actions=construir_acciones_handoff(),
            )

        # Mientras está activo el handoff, los mensajes van al canal humano.
        media_urls_colab = getattr(g, "whatsapp_media_urls", None) or []
        _add_handoff_message(
            handoff_id,
            remitente="colaborador",
            texto=mensaje_usuario,
            media_url=media_urls_colab if media_urls_colab else None,
        )
        _upsert_handoff(
            handoff_id,
            {"ultima_consulta": mensaje_usuario.strip(), "updated_at": _utc_now()},
            merge=True,
        )
        conv = _fetch_handoff(handoff_id) or {}
        _send_push_all(
            "Nuevo mensaje",
            f"{conv.get('colaborador_nombre') or conv.get('colaborador_telefono') or 'Usuario'}: {mensaje_usuario.strip()[:150]}",
        )
        estado = str(conv.get("estado") or "").strip().lower()
        if estado == HANDOFF_STATUS_PENDING:
            texto = "📨 Mensaje enviado. Un agente todavía no tomó la conversación."
        else:
            agente = conv.get("rrhh_agente") or "un agente"
            texto = f"📨 Mensaje enviado a {agente}. Te responderán por este chat."
        return _payload(
            texto,
            handoff_active=True,
            quick_actions=construir_acciones_handoff(),
        )

    # Frases inequívocas de contacto humano — se procesan siempre, sin importar el estado de sesión.
    # Esto cubre el caso WhatsApp donde pending_derivacion se perdió entre instancias de Cloud Run.
    # "h" es el value del botón "H. Hablar con un agente" en el menú de WhatsApp.
    _FRASES_CONTACTO_DIRECTA = {
        "h",
        "hablar con alguien", "hablar con un agente", "hablar con un asistente",
        "quiero hablar con alguien", "quiero hablar con un agente",
        "necesito hablar con alguien", "necesito un agente",
        "contacto humano", "atencion humana", "derivame con alguien",
    }
    _texto_norm_contacto = chatbot.normalizar_texto(mensaje_usuario)
    _es_contacto_directo = _texto_norm_contacto == "h" or any(
        f in _texto_norm_contacto for f in _FRASES_CONTACTO_DIRECTA if f != "h"
    )
    if permitir and _es_contacto_directo:
        _sess().pop("pending_feedback_topic", None)
        _sess().pop("pending_derivacion", None)
        conversation_id = _iniciar_handoff_rrhh(mensaje_usuario)
        conv = _fetch_handoff(conversation_id) or {}
        assigned = str(conv.get("rrhh_agente") or "").strip()
        if assigned:
            respuesta = f"👩‍💼 Perfecto, te derivé con el equipo de atención.\nTe asigné con {assigned}. Podés seguir escribiendo por este chat."
        else:
            respuesta = "👩‍💼 Perfecto, derivé tu consulta al equipo de atención.\nTe van a responder por este mismo chat."
        return _payload(respuesta, handoff_active=True, quick_actions=construir_acciones_handoff())

    # Estado: el bot ofreció derivación y espera confirmación del usuario
    if _sess().get("pending_derivacion"):
        _sess().pop("pending_derivacion", None)
        msg_norm = chatbot.normalizar_texto(mensaje_usuario)
        _quiere_derivacion = (
            msg_norm in {"si", "sí", "dale", "ok", "bueno", "yes", "quiero", "derivame"}
            or msg_norm.startswith("si ")
            or chatbot.solicita_contacto_rrhh(msg_norm)
        )
        if _quiere_derivacion:
            if not permitir:
                return _payload(
                    "La derivación a un agente está desactivada para esta empresa. "
                    "Podés seguir consultando el menú.",
                    quick_actions=_acciones_menu(6),
                )
            conversation_id = _iniciar_handoff_rrhh(mensaje_usuario)
            conv = _fetch_handoff(conversation_id) or {}
            assigned = str(conv.get("rrhh_agente") or "").strip()
            if assigned:
                respuesta = (
                    "👩‍💼 Perfecto, te derivé con el equipo de atención.\n"
                    f"Te asigné con {assigned}. Podés seguir escribiendo por este chat."
                )
            else:
                respuesta = (
                    "👩‍💼 Perfecto, derivé tu consulta al equipo de atención.\n"
                    "Te van a responder por este mismo chat."
                )
            return _payload(respuesta, handoff_active=True, quick_actions=construir_acciones_handoff())
        elif msg_norm in {"no", "no gracias", "no, gracias"}:
            return _payload(
                "Entendido. Si necesitás algo más, estoy por acá.",
                quick_actions=_acciones_menu(6),
            )
        # Cualquier otra cosa = nueva consulta, cae al flujo normal sin responder "Entendido"

    tema_pendiente = _sess().get("pending_feedback_topic")
    if tema_pendiente:
        payload = procesar_feedback_pendiente(mensaje_usuario, tema_pendiente, temas_map, permitir_hablar_con_humano=permitir, company_id=company_id)
        if payload is not None:
            return payload
        # Si llega una nueva consulta durante feedback, sigue flujo normal.

    # Solo disparar handoff directo si la empresa NO tiene KB cargada
    # Si tiene KB, dejar que obtener_respuesta busque primero (ej: "contacto" puede ser sección de la KB)
    _tiene_kb = bool(company_id and chatbot.obtener_knowledge_empresa(company_id))
    if chatbot.solicita_contacto_rrhh(mensaje_norm) and not _tiene_kb:
        if not permitir:
            return _payload(
                "En este momento la derivación a un agente está desactivada para esta empresa. "
                "Para activarla: entrá a Preferencias en el panel, sección «Reglas del chat», elegí la empresa y marcá «Mostrar botón de derivación y permitir hablar con un agente», luego Guardar. "
                "Mientras tanto podés consultar el menú de temas.",
                quick_actions=_acciones_menu(6),
            )
        conversation_id = _iniciar_handoff_rrhh(mensaje_usuario)
        conv = _fetch_handoff(conversation_id) or {}
        assigned = str(conv.get("rrhh_agente") or "").strip()
        if assigned:
            respuesta = (
                "👩‍💼 Perfecto. Derivé tu consulta al equipo de atención.\n"
                f"Te asigné automáticamente con {assigned}. Podés seguir escribiendo por este chat."
            )
        else:
            respuesta = (
                "👩‍💼 Perfecto. Derivé tu consulta al equipo de atención.\n"
                "Te van a responder por este mismo chat. Podés seguir escribiendo aquí."
            )
        return _payload(
            respuesta,
            handoff_active=True,
            quick_actions=construir_acciones_handoff(),
        )

    if mensaje_norm == "menu":
        return _payload(
            construir_menu_texto(temas_map, permitir_hablar_con_humano=permitir),
            quick_actions=construir_acciones_menu(temas_map, permitir_hablar_con_humano=permitir),
        )

    if chatbot.es_saludo(mensaje_norm):
        company_obj = _current_company()
        c_name = (company_obj or {}).get("company_name") or "tu empresa"
        area_ctx = session.get("chat_context_area") or ""
        area_str = f" del área {area_ctx}" if area_ctx else ""
        return _payload(
            f"👋 ¡Hola! ¿Cómo estás? Soy Debo, tu asistente{area_str} de {c_name}. ¿En qué te puedo ayudar hoy? 😊",
            quick_actions=_acciones_menu(6),
        )

    last_kb_pregunta = _sess().get("last_kb_pregunta")
    _wa_convenio = (_sess().get("wa_convenio") or "").strip() or None
    respuesta, tema_id = chatbot.obtener_respuesta(
        mensaje_usuario, temas_map, company_id=company_id, context_pregunta=last_kb_pregunta, convenio=_wa_convenio
    )
    if respuesta:
        if tema_id == "knowledge_answer":
            _sess()["last_kb_pregunta"] = mensaje_usuario
        elif tema_id in ("saludo", "ayuda", None):
            pass  # mantener contexto anterior
        else:
            _sess().pop("last_kb_pregunta", None)

        if tema_id == "knowledge_derivacion":
            _sess()["pending_derivacion"] = True
            return _payload(
                respuesta,
                await_feedback=False,
                quick_actions=[
                    {"label": "Sí, derivame", "value": "si"},
                    {"label": "No, gracias", "value": "no"},
                    {"label": "Nueva consulta", "value": "nueva consulta"},
                ],
            )
        requiere_feedback = tema_id not in chatbot.TEMAS_SIN_FEEDBACK
        if requiere_feedback:
            _sess()["pending_feedback_topic"] = tema_id
            respuesta = f"{respuesta}\n\n¿Esta información te fue de utilidad? (si/no)"
            return _payload(
                respuesta,
                await_feedback=True,
                quick_actions=construir_acciones_feedback(),
            )
        return _payload(
            respuesta,
            quick_actions=_acciones_menu(4),
        )

    guardado_pendiente = chatbot.registrar_pendiente(mensaje_usuario, company_id=company_id)
    respuesta = armar_respuesta_no_entendida(mensaje_usuario, temas_map)
    if not guardado_pendiente:
        respuesta += "\nℹ️ No pude registrar esta consulta en la base de datos."
    return _payload(
        respuesta,
        quick_actions=construir_acciones_menu(temas_map, limite=6, permitir_hablar_con_humano=permitir),
    )


def _serialize_handoff(conv):
    updated_at = _as_utc_aware(conv.get("updated_at"))
    ultimo_mensaje_fecha = _as_utc_aware(conv.get("ultimo_mensaje_fecha")) or updated_at
    return {
        "conversation_id": conv.get("conversation_id") or conv.get("id"),
        "company_id": conv.get("company_id") or "",
        "company_name": conv.get("company_name") or "",
        "branch": conv.get("branch") or "",
        "area": conv.get("area") or "",
        "estado": conv.get("estado") or HANDOFF_STATUS_PENDING,
        "rrhh_agente": conv.get("rrhh_agente") or "",
        "rrhh_agente_id": conv.get("rrhh_agente_id") or "",
        "rrhh_asignacion_automatica": bool(conv.get("rrhh_asignacion_automatica")),
        "ultima_consulta": conv.get("ultima_consulta") or "",
        "ultimo_remitente": str(conv.get("ultimo_remitente") or "").strip().lower(),
        "ultimo_mensaje_iso": _iso_utc(ultimo_mensaje_fecha),
        "updated_at": _fmt_fecha(updated_at),
        "updated_at_iso": _iso_utc(updated_at),
        "channel": conv.get("channel") or "",
        "colaborador_nombre": conv.get("colaborador_nombre") or "",
        "colaborador_telefono": conv.get("colaborador_telefono") or "",
        "colaborador_mensajes_count": int(conv.get("colaborador_mensajes_count") or 0),
    }


def _serialize_messages(messages):
    payload = []
    for msg in messages:
        fecha = _as_utc_aware(msg.get("fecha"))
        item = {
            "remitente": str(msg.get("remitente") or ""),
            "texto": str(msg.get("texto") or ""),
            "agente": str(msg.get("agente") or ""),
            "fecha": _fmt_fecha(fecha),
            "fecha_iso": _iso_utc(fecha),
        }
        if msg.get("media_url"):
            item["media_url"] = msg["media_url"] if isinstance(msg["media_url"], list) else [msg["media_url"]]
        payload.append(item)
    return payload


def _conversation_matches_selected_company(conv):
    if not isinstance(conv, dict):
        return False
    selected_company = _selected_company_id_for_rrhh()
    if not selected_company:
        return True
    return _normalize_company_id(conv.get("company_id")) == selected_company


def reset_in_memory_handoffs():
    IN_MEMORY_HANDOFFS.clear()
    IN_MEMORY_CHAT_HISTORY.clear()
    IN_MEMORY_ACTIVE_AGENTS.clear()
    IN_MEMORY_GENERAL_SETTINGS.clear()
    IN_MEMORY_COMPANIES.clear()


def _can_view_stats():
    if not _auth_enabled():
        return True
    return _has_permission(auth_rrhh.PERM_STATS_VIEW)


def _can_manage_preferences():
    if not _auth_enabled():
        return True
    return _has_permission(auth_rrhh.PERM_PREFERENCES_MANAGE)


@flask_app.get("/sw.js")
def serve_sw():
    from flask import make_response, send_from_directory
    resp = make_response(send_from_directory("static", "sw.js"))
    resp.headers["Content-Type"] = "application/javascript"
    resp.headers["Service-Worker-Allowed"] = "/"
    return resp


@flask_app.get("/api/push/vapid-key")
@rrhh_auth_required
def push_vapid_key():
    _ensure_vapid_keys()
    if not _VAPID_PUBLIC_KEY:
        return jsonify({"ok": False, "error": "Push no disponible"})
    return jsonify({"ok": True, "public_key": _VAPID_PUBLIC_KEY})


@flask_app.post("/api/push/subscribe")
@rrhh_auth_required
def push_subscribe():
    sub = request.json
    if not sub or not sub.get("endpoint"):
        return jsonify({"ok": False})
    sub_id = _hashlib.sha256(sub["endpoint"].encode()).hexdigest()[:32]
    username = (_current_rrhh_user() or {}).get("username", "anon")
    if chatbot.db:
        chatbot.db.collection("push_subscriptions").document(sub_id).set({
            "subscription": sub,
            "username": username,
            "updated_at": _utc_now(),
        })
    return jsonify({"ok": True})


@flask_app.post("/api/push/unsubscribe")
@rrhh_auth_required
def push_unsubscribe():
    data = request.json or {}
    endpoint = (data.get("endpoint") or "").strip()
    if endpoint and chatbot.db:
        sub_id = _hashlib.sha256(endpoint.encode()).hexdigest()[:32]
        try:
            chatbot.db.collection("push_subscriptions").document(sub_id).delete()
        except Exception:
            pass
    return jsonify({"ok": True})


@flask_app.get("/")
def home():
    """Vista principal: layout con sidebar de módulos y contenido en iframe (por defecto chat)."""
    requested_company = _normalize_company_id(request.args.get("empresa"))
    if requested_company and _get_company(requested_company, include_inactive=False):
        _set_company_session(requested_company)
    else:
        _set_company_session(session.get("company_id") or _default_company_id())
    settings = _apply_company_branding(_read_general_settings())
    company_name = settings.get("company_name") or "Empresa"
    hr_display = (settings.get("hr_team_name") or "Atención").strip()
    if hr_display.upper() == "RRHH":
        hr_display = "Atención"
    user = _current_rrhh_user()
    if _auth_enabled() and user is None:
        return redirect(url_for("login_page", next=_request_path_with_query()))
    show_all = False
    _allowed_modules = {
        "chat",
        "rrhh",
        "configuracion",
        "estadisticas",
        "historial",
        "comunicados",
        "preferencias",
        "legajos",
    }
    initial_module = request.args.get("m", "chat")
    if initial_module not in _allowed_modules:
        initial_module = "chat"
    return render_template(
        "index.html",
        company_name=company_name,
        hr_team_name=hr_display,
        initial_module=initial_module,
        can_view_config=_can_manage_configuration() if not show_all else True,
        can_view_stats=_can_view_stats() if not show_all else True,
        can_manage_preferences=_can_manage_preferences() if not show_all else True,
        can_view_conversations=_has_permission(auth_rrhh.PERM_CONVERSATIONS_VIEW) if not show_all else True,
        can_view_history=_has_permission(auth_rrhh.PERM_HISTORY_VIEW) if not show_all else True,
        can_view_comunicados=_has_permission(auth_rrhh.PERM_COMUNICADOS_SEND) if not show_all else False,
        can_view_legajos=_has_permission(auth_rrhh.PERM_LEGAJOS_VIEW) if not show_all else False,
    )


@flask_app.get("/chat")
def chat_page():
    """Contenido del chatbot para incrustar en el layout principal (iframe o vista por defecto)."""
    _clear_chat_context()
    step = _chat_context_step()
    if step == CHAT_CONTEXT_STEP_READY or step == CHAT_CONTEXT_STEP_AREA or step == CHAT_CONTEXT_STEP_BRANCH:
        requested_company = _normalize_company_id(
            session.get("chat_context_company_id") or session.get("company_id") or _default_company_id()
        )
        _set_company_session(requested_company)
    else:
        _set_company_session(session.get("company_id") or _default_company_id())

    settings = _apply_company_branding(_read_general_settings())
    company = _current_company()
    company_id = (company or {}).get("company_id")
    company_name = settings.get("company_name") or "Empresa"
    hr_display = (settings.get("hr_team_name") or "Atención").strip()
    if hr_display.upper() == "RRHH":
        hr_display = "Atención"

    _bot_name = BOT_NAME
    if step == CHAT_CONTEXT_STEP_COMPANY:
        quick_actions_iniciales = _construir_acciones_empresas(limite=8)
        nombres_empresas = [a.get("label") or a.get("value") for a in quick_actions_iniciales if a.get("label") or a.get("value")]
        if nombres_empresas:
            bienvenida = (
                f"👋 ¡Hola! Soy {_bot_name}, tu asistente de {hr_display}. 😊\n"
                f"Para orientarte mejor, ¿con qué empresa estás relacionado/a?\n\n"
                + "\n".join(nombres_empresas)
                + "\n\nEscribí el número o el nombre."
            )
        else:
            bienvenida = f"👋 ¡Hola! Soy {_bot_name}. 😊 ¿Con qué empresa estás relacionado/a? Escribí el nombre."
    elif step == CHAT_CONTEXT_STEP_BRANCH:
        ctx_company_id = session.get("chat_context_company_id") or company_id
        company_for_branch = _get_company(ctx_company_id, include_inactive=False)
        area_name_company = (company_for_branch or {}).get("company_name") or ctx_company_id or "Empresa"
        quick_actions_iniciales = _construir_acciones_sucursales(ctx_company_id, limite=8)
        nombres_sucursales = [a.get("label") or a.get("value") for a in quick_actions_iniciales if a.get("label") or a.get("value")]
        if nombres_sucursales:
            bienvenida = (
                f"¡Perfecto, {area_name_company}! ¿Y de qué sucursal sos?\n\n"
                + "\n".join(nombres_sucursales)
                + "\n\nEscribí el número o el nombre."
            )
        else:
            bienvenida = f"¡Perfecto, {area_name_company}! ¿Y de qué sucursal sos?"
    elif step == CHAT_CONTEXT_STEP_AREA:
        ctx_company_id = session.get("chat_context_company_id") or company_id
        ctx_branch = session.get("chat_context_branch") or ""
        company_for_area = _get_company(ctx_company_id, include_inactive=False)
        area_name_company = (company_for_area or {}).get("company_name") or ctx_company_id or "Empresa"
        quick_actions_iniciales = _construir_acciones_areas(ctx_company_id, limite=8, branch=ctx_branch or None)
        nombres_areas = [a.get("label") or a.get("value") for a in quick_actions_iniciales if a.get("label") or a.get("value")]
        suf = f" de {ctx_branch}" if ctx_branch else ""
        if nombres_areas:
            bienvenida = (
                f"Entendido{suf}. ¿Y a qué área pertenecés?\n\n"
                + "\n".join(nombres_areas)
                + "\n\nEscribí el número o el nombre."
            )
        else:
            bienvenida = f"Entendido{suf}. ¿A qué área pertenecés?"
    else:
        permitir = (company or {}).get("permitir_hablar_con_humano", True)
        temas_habilitados = (company or {}).get("temas_habilitados") or []
        temas_map = construir_temas_map(company_id=company_id, temas_habilitados=temas_habilitados)
        area_ctx = session.get("chat_context_area") or ""
        branch_ctx = session.get("chat_context_branch") or ""
        partes = []
        if branch_ctx:
            partes.append(branch_ctx)
        if area_ctx:
            partes.append(area_ctx)
        ctx_str = f" ({', '.join(partes)})" if partes else ""
        if hr_display and hr_display.lower() != company_name.lower() and hr_display.lower() != "atención":
            _label = f"de {hr_display} de {company_name}"
        else:
            _label = f"de {company_name}"
        bienvenida = (
            f"👋 ¡Hola! Soy {_bot_name}, tu asistente {_label}{ctx_str}. 😊\n"
            f"Estoy acá para ayudarte. ¿Sobre qué tema querés consultar?"
        )
        quick_actions_iniciales = construir_acciones_menu(temas_map, limite=6, permitir_hablar_con_humano=permitir)

    return render_template(
        "chat.html",
        bienvenida=bienvenida,
        quick_actions_iniciales=quick_actions_iniciales,
        company_name=company_name,
        hr_team_name=hr_display,
    )


@flask_app.after_request
def add_no_cache_headers(response):
    # Evita cache agresivo de Safari en páginas/API dinámicas.
    if not request.path.startswith("/static/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


@flask_app.get("/rrhh")
@rrhh_permission_required(
    auth_rrhh.PERM_CONVERSATIONS_VIEW,
    message="No tenés permisos para ver el panel de conversaciones.",
)
def rrhh_page():
    company = _set_company_session(session.get("company_id") or _default_company_id())
    settings = _apply_company_branding(_read_general_settings())
    current_user = _current_rrhh_user()
    available_companies = (
        _companies_for_user(current_user) if current_user else _list_companies(include_inactive=False)
    )
    return render_template(
        "rrhh.html",
        auth_enabled=_auth_enabled(),
        rrhh_user=current_user,
        can_manage_users=False,
        can_manage_roles=False,
        can_manage_config=_can_manage_configuration(),
        available_companies=available_companies,
        company_name=settings.get("company_name"),
        hr_team_name=settings.get("hr_team_name"),
        selected_company_id=company.get("company_id"),
        selected_company_name=company.get("company_name"),
    )


@flask_app.get("/configuracion")
@rrhh_auth_required
def configuracion_page():
    if not _can_manage_configuration():
        return ("No tenés permisos para acceder a configuración.", 403)
    company = _set_company_session(session.get("company_id") or _default_company_id())
    settings = _read_general_settings()
    companies = [_company_for_api(c) for c in _list_companies_for_current_rrhh_user(include_inactive=True)]
    return render_template(
        "configuracion.html",
        auth_enabled=_auth_enabled(),
        rrhh_user=_current_rrhh_user(),
        can_manage_users=_has_permission(auth_rrhh.PERM_USERS_MANAGE),
        can_manage_roles=_has_permission(auth_rrhh.PERM_ROLES_MANAGE),
        can_manage_general=_can_manage_general_config(),
        can_config_empresas=_can_config_empresas(),
        can_config_sucursales=_can_config_sucursales(),
        can_config_areas=_can_config_areas(),
        can_config_knowledge=_can_config_knowledge(),
        can_config_convenios=_can_config_convenios(),
        can_config_smtp=_can_config_smtp(),
        can_send_comunicados=_has_permission(auth_rrhh.PERM_COMUNICADOS_SEND),
        is_admin=_is_admin(),
        general_settings=settings,
        companies=companies,
        selected_company_id=company.get("company_id"),
        all_tipos_documento=legajos_service.ALL_TIPOS_DOCUMENTO,
    )


@flask_app.get("/historial")
@rrhh_permission_required(
    auth_rrhh.PERM_HISTORY_VIEW,
    message="No tenés permisos para ver el historial.",
)
def historial_page():
    return render_template(
        "historial.html",
        auth_enabled=_auth_enabled(),
        rrhh_user=_current_rrhh_user(),
    )


@flask_app.get("/legajos")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_VIEW,
    message="No tenés permisos para acceder a Legajos digitales.",
)
def legajos_page():
    """Módulo de legajos digitales: listado de colaboradores y documentos en Storage."""
    company = _set_company_session(session.get("company_id") or _default_company_id())
    current_user = _current_rrhh_user()
    available_companies = (
        _companies_for_user(current_user) if current_user else _list_companies(include_inactive=False)
    )
    return render_template(
        "legajos.html",
        auth_enabled=_auth_enabled(),
        rrhh_user=current_user,
        can_manage_legajos=_has_permission(auth_rrhh.PERM_LEGAJOS_MANAGE),
        available_companies=available_companies,
        selected_company_id=company.get("company_id"),
        selected_company_name=company.get("company_name"),
        all_tipos_documento=legajos_service.ALL_TIPOS_DOCUMENTO,
    )


def _legajos_effective_company_id(explicit: str | None):
    """Empresa activa para legajos: parámetro explícito o sesión; valida acceso del usuario."""
    raw = str(explicit or "").strip()
    cid = _normalize_company_id(raw) if raw else ""
    if not cid:
        cid = _normalize_company_id(_selected_company_id_for_rrhh() or _default_company_id() or "")
    if not cid:
        return None, "No hay empresa seleccionada."
    user = _current_rrhh_user()
    if _auth_enabled() and user and not _user_can_access_company(user, cid):
        return None, "Sin acceso a esa empresa."
    return cid, None


def _legajos_empleado_si_acceso(empleado_id: str):
    emp = legajos_service.get_empleado(chatbot.db, empleado_id)
    if not emp:
        return None, "Colaborador no encontrado."
    user = _current_rrhh_user()
    cid = str(emp.get("company_id") or "").strip().lower()
    if _auth_enabled() and user and not _user_can_access_company(user, cid):
        return None, "Sin acceso a este legajo."
    return emp, None


def _legajos_documento_si_acceso(documento_id: str):
    doc = legajos_service.get_documento(chatbot.db, documento_id)
    if not doc:
        return None, "Documento no encontrado."
    user = _current_rrhh_user()
    cid = str(doc.get("company_id") or "").strip().lower()
    if _auth_enabled() and user and not _user_can_access_company(user, cid):
        return None, "Sin acceso a este documento."
    return doc, None


def _legajos_audit(action: str, company_id: str, details: dict | None = None):
    """Registra auditoría de legajos (no interrumpe si falla)."""
    if not chatbot.db:
        return
    try:
        user = _current_rrhh_user()
        uname = str((user or {}).get("username") or "").strip() or "anon"
        legajos_service.append_auditoria(
            chatbot.db,
            str(company_id or "").strip().lower(),
            uname,
            action,
            details,
        )
    except Exception as exc:
        logger.debug("legajos audit omitido: %s", exc)


def _parse_legajos_audit_datetime(param_name: str):
    """Parsea ISO 8601 desde query (ej. toISOString() del navegador) a UTC."""
    raw = (request.args.get(param_name) or "").strip()
    if not raw:
        return None
    try:
        s = raw.replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return None


@flask_app.post("/api/legajos/empresa/seleccionar")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_VIEW,
    message="Sin permiso para usar legajos.",
)
def api_legajos_seleccionar_empresa():
    """Cambia la empresa en sesión (misma lógica que RRHH) para usuarios que solo tienen permiso de legajos."""
    data = request.get_json(silent=True) or {}
    company_id = _normalize_company_id(data.get("company_id"))
    if not company_id:
        return jsonify({"ok": False, "error": "Seleccioná una empresa válida."}), 400
    company = _get_company(company_id, include_inactive=False)
    if not company:
        return jsonify({"ok": False, "error": "Empresa no encontrada o inactiva."}), 404
    current_user = _current_rrhh_user()
    if current_user and not _user_can_access_company(current_user, company_id):
        return _forbidden_json_error("No tenés acceso a la empresa seleccionada.")
    selected = _set_company_session(company.get("company_id"))
    return jsonify(
        {
            "ok": True,
            "company": {
                "company_id": selected.get("company_id"),
                "company_name": selected.get("company_name"),
            },
        }
    )


@flask_app.get("/api/legajos/empleados")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_VIEW,
    message="Sin permiso para ver legajos.",
)
def api_legajos_empleados_list():
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible.", "empleados": []}), 503
    cid, err = _legajos_effective_company_id(request.args.get("company_id"))
    if err:
        return jsonify({"ok": False, "error": err, "empleados": []}), 400
    q = (request.args.get("q") or request.args.get("search") or "").strip()
    activo_param = request.args.get("activo", "true").strip().lower()
    if activo_param == "all":
        activo_filter = None
    elif activo_param == "false":
        activo_filter = False
    else:
        activo_filter = True
    items = legajos_service.list_empleados(chatbot.db, cid, search=q or None, activo=activo_filter)
    return jsonify({"ok": True, "empleados": items, "company_id": cid})


@flask_app.get("/api/legajos/empleados/export")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_VIEW,
    message="Sin permiso para exportar legajos.",
)
def api_legajos_empleados_export():
    """Descarga Excel (.xlsx) con las mismas columnas que el ejemplo de importación (columnas separadas)."""
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible."}), 503
    cid, err = _legajos_effective_company_id(request.args.get("company_id"))
    if err:
        return jsonify({"ok": False, "error": err}), 400
    q = (request.args.get("q") or request.args.get("search") or "").strip()
    items = legajos_service.list_empleados(chatbot.db, cid, search=q or None)
    body, gen_err = legajos_service.build_legajos_export_xlsx_bytes(items)
    if gen_err or not body:
        return jsonify({"ok": False, "error": gen_err or "No se pudo generar el Excel."}), 503
    safe_cid = "".join(c for c in cid if c.isalnum() or c in "._-") or "empresa"
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    filename = f"legajos_colaboradores_{safe_cid}_{stamp}.xlsx"
    return send_file(
        io.BytesIO(body),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@flask_app.get("/api/legajos/empleados/ejemplo-importacion")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_MANAGE,
    message="Sin permiso para descargar el ejemplo de legajos.",
)
def api_legajos_ejemplo_importacion():
    """Descarga un .xlsx con columnas separadas (Excel); al importar, el servidor lee las mismas filas que con CSV."""
    body, err = legajos_service.build_legajos_ejemplo_xlsx_bytes()
    if err or not body:
        return jsonify({"ok": False, "error": err or "No se pudo generar el ejemplo."}), 503
    return send_file(
        io.BytesIO(body),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="legajos_colaboradores_ejemplo.xlsx",
    )


@flask_app.post("/api/legajos/empleados")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_MANAGE,
    message="Sin permiso para crear o editar legajos.",
)
def api_legajos_empleados_create():
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible."}), 503
    data = request.get_json(silent=True) or {}
    cid, err = _legajos_effective_company_id(data.get("company_id"))
    if err:
        return jsonify({"ok": False, "error": err}), 400
    user = _current_rrhh_user()
    uname = str((user or {}).get("username") or "").strip()
    ok, row, msg = legajos_service.create_empleado(
        chatbot.db,
        company_id=cid,
        dni=str(data.get("dni") or "").strip(),
        nombre_completo=str(data.get("nombre_completo") or "").strip(),
        created_by=uname,
        legajo_numero=str(data.get("legajo_numero") or "").strip(),
        sucursal=str(data.get("sucursal") or "").strip(),
        area=str(data.get("area") or "").strip(),
        notas=str(data.get("notas") or "").strip(),
        email=str(data.get("email") or "").strip(),
        convenio=str(data.get("convenio") or "").strip(),
    )
    if not ok:
        return jsonify({"ok": False, "error": msg or "No se pudo crear el colaborador."}), 400
    _legajos_audit(
        legajos_service.LEGAJOS_AUDIT_EMPLEADO_CREAR,
        cid,
        {"empleado_id": (row or {}).get("id"), "dni": (row or {}).get("dni")},
    )
    return jsonify({"ok": True, "empleado": row})


@flask_app.delete("/api/legajos/empleados/<empleado_id>")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_MANAGE,
    message="Sin permiso para eliminar colaboradores.",
)
def api_legajos_empleado_delete(empleado_id):
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible."}), 503
    emp, err = _legajos_empleado_si_acceso(empleado_id)
    if err:
        return jsonify({"ok": False, "error": err}), 404
    cid = str(emp.get("company_id") or "").strip().lower()
    ok, paths, msg = legajos_service.delete_empleado_completo(chatbot.db, empleado_id)
    if not ok:
        return jsonify({"ok": False, "error": msg or "No se pudo eliminar."}), 400
    for p in paths:
        try:
            bucket = _get_storage_bucket()
            if bucket and p:
                bucket.blob(p).delete()
        except Exception:
            pass
    _legajos_audit(
        legajos_service.LEGAJOS_AUDIT_EMPLEADO_ELIMINAR,
        cid,
        {"empleado_id": empleado_id, "dni": emp.get("dni")},
    )
    return jsonify({"ok": True})


@flask_app.patch("/api/legajos/empleados/<empleado_id>/estado")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_MANAGE,
    message="Sin permiso para editar colaboradores.",
)
def api_legajos_empleado_estado(empleado_id):
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible."}), 503
    emp, err = _legajos_empleado_si_acceso(empleado_id)
    if err:
        return jsonify({"ok": False, "error": err}), 404
    body = request.get_json(silent=True) or {}
    activo = bool(body.get("activo", True))
    legajos_service.set_empleado_activo(chatbot.db, empleado_id, activo)
    cid = str(emp.get("company_id") or "").strip().lower()
    _legajos_audit(
        legajos_service.LEGAJOS_AUDIT_EMPLEADO_EDITAR,
        cid,
        {"empleado_id": empleado_id, "dni": emp.get("dni"), "activo": activo},
    )
    return jsonify({"ok": True, "activo": activo})


@flask_app.get("/api/legajos/empleados/<empleado_id>")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_VIEW,
    message="Sin permiso para ver legajos.",
)
def api_legajos_empleado_get(empleado_id):
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible."}), 503
    emp, err = _legajos_empleado_si_acceso(empleado_id)
    if err:
        return jsonify({"ok": False, "error": err}), 404
    return jsonify({"ok": True, "empleado": emp})


@flask_app.patch("/api/legajos/empleados/<empleado_id>")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_MANAGE,
    message="Sin permiso para editar legajos.",
)
def api_legajos_empleado_patch(empleado_id):
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible."}), 503
    emp, err = _legajos_empleado_si_acceso(empleado_id)
    if err:
        return jsonify({"ok": False, "error": err}), 404
    data = request.get_json(silent=True) or {}

    def _patch_str(key, fallback):
        if key in data:
            return str(data.get(key) or "").strip()
        return str(fallback or "").strip()

    user = _current_rrhh_user()
    uname = str((user or {}).get("username") or "").strip()
    ok, row, msg = legajos_service.update_empleado(
        chatbot.db,
        empleado_id=empleado_id,
        legajo_numero=_patch_str("legajo_numero", emp.get("legajo_numero")),
        nombre_completo=_patch_str("nombre_completo", emp.get("nombre_completo")),
        updated_by=uname,
        sucursal=_patch_str("sucursal", emp.get("sucursal")),
        area=_patch_str("area", emp.get("area")),
        notas=_patch_str("notas", emp.get("notas")),
        email=_patch_str("email", emp.get("email")),
        convenio=_patch_str("convenio", emp.get("convenio")),
    )
    if not ok:
        return jsonify({"ok": False, "error": msg or "No se pudo actualizar."}), 400
    cid = str(emp.get("company_id") or "").strip().lower()
    _legajos_audit(
        legajos_service.LEGAJOS_AUDIT_EMPLEADO_EDITAR,
        cid,
        {"empleado_id": empleado_id, "legajo_numero": (row or {}).get("legajo_numero")},
    )
    return jsonify({"ok": True, "empleado": row})


@flask_app.post("/api/legajos/empleados/import")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_MANAGE,
    message="Sin permiso para importar legajos.",
)
def api_legajos_empleados_import():
    """Importa colaboradores desde .xlsx (Excel) o .csv (UTF-8). Mismas columnas; el servidor convierte Excel a filas como CSV."""
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible."}), 503
    cid, err = _legajos_effective_company_id(request.form.get("company_id") or request.args.get("company_id"))
    if err:
        return jsonify({"ok": False, "error": err}), 400
    file_storage = request.files.get("file")
    if not file_storage or not file_storage.filename:
        return jsonify({"ok": False, "error": "Adjuntá un archivo .xlsx o .csv"}), 400
    name = (file_storage.filename or "").lower()
    if not (name.endswith(".csv") or name.endswith(".xlsx") or name.endswith(".xlsm")):
        return jsonify({"ok": False, "error": "El archivo debe ser .xlsx (Excel) o .csv"}), 400
    raw = file_storage.read()
    filas, parse_err = legajos_service.parse_legajos_import_file(file_storage.filename or "", raw)
    if parse_err:
        return jsonify({"ok": False, "error": parse_err}), 400
    if not filas:
        return jsonify({"ok": False, "error": "No hay filas de datos para importar."}), 400
    user = _current_rrhh_user()
    uname = str((user or {}).get("username") or "").strip()
    result = legajos_service.import_empleados_desde_filas(chatbot.db, cid, filas, created_by=uname)
    _legajos_audit(
        legajos_service.LEGAJOS_AUDIT_EMPLEADO_IMPORTAR,
        cid,
        {
            "creados": result.get("created"),
            "omitidos_duplicado": result.get("skipped_duplicate"),
            "errores": len(result.get("errors") or []),
        },
    )
    return jsonify({"ok": True, **result, "company_id": cid})


@flask_app.get("/api/legajos/usuarios-empresa")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_VIEW,
    message="Sin permiso para ver legajos.",
)
def api_legajos_usuarios_empresa():
    """Usuarios RRHH con acceso a la empresa activa (para filtro de auditoría)."""
    cid, err = _legajos_effective_company_id(request.args.get("company_id"))
    if err:
        return jsonify({"ok": False, "error": err, "usuarios": []}), 400
    usuarios = auth_rrhh.list_users_for_company(cid)
    return jsonify({"ok": True, "usuarios": usuarios, "company_id": cid})


@flask_app.get("/api/legajos/auditoria")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_VIEW,
    message="Sin permiso para ver auditoría de legajos.",
)
def api_legajos_auditoria_list():
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible.", "eventos": []}), 503
    cid, err = _legajos_effective_company_id(request.args.get("company_id"))
    if err:
        return jsonify({"ok": False, "error": err, "eventos": []}), 400
    try:
        lim = int(request.args.get("limit") or 80)
    except ValueError:
        lim = 80
    username = (request.args.get("username") or "").strip() or None
    action = (request.args.get("action") or "").strip() or None
    q = (request.args.get("q") or request.args.get("search") or "").strip() or None
    at_from = _parse_legajos_audit_datetime("at_from")
    at_to = _parse_legajos_audit_datetime("at_to")
    eventos = legajos_service.list_auditoria(
        chatbot.db,
        cid,
        limit=lim,
        username=username,
        action=action,
        q=q,
        at_from=at_from,
        at_to=at_to,
    )
    return jsonify({"ok": True, "eventos": eventos, "company_id": cid})


@flask_app.get("/api/legajos/auditoria/export")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_VIEW,
    message="Sin permiso para exportar auditoría de legajos.",
)
def api_legajos_auditoria_export():
    """Descarga Excel con los mismos filtros que el listado (hasta 5000 filas)."""
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible."}), 503
    cid, err = _legajos_effective_company_id(request.args.get("company_id"))
    if err:
        return jsonify({"ok": False, "error": err}), 400
    username = (request.args.get("username") or "").strip() or None
    action = (request.args.get("action") or "").strip() or None
    q = (request.args.get("q") or request.args.get("search") or "").strip() or None
    at_from = _parse_legajos_audit_datetime("at_from")
    at_to = _parse_legajos_audit_datetime("at_to")
    eventos = legajos_service.list_auditoria(
        chatbot.db,
        cid,
        limit=5000,
        username=username,
        action=action,
        q=q,
        at_from=at_from,
        at_to=at_to,
        max_limit=5000,
    )
    body, gen_err = legajos_service.build_auditoria_export_xlsx_bytes(eventos)
    if gen_err or not body:
        return jsonify({"ok": False, "error": gen_err or "No se pudo generar el Excel."}), 503
    safe_cid = "".join(c for c in cid if c.isalnum() or c in "._-") or "empresa"
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    filename = f"legajos_auditoria_{safe_cid}_{stamp}.xlsx"
    return send_file(
        io.BytesIO(body),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@flask_app.get("/api/legajos/documentos/buscar")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_VIEW,
    message="Sin permiso para buscar documentos de legajos.",
)
def api_legajos_documentos_buscar():
    """Busca archivos por nombre y/o carpeta (tipo_documento) en la empresa."""
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible.", "resultados": []}), 503
    cid, err = _legajos_effective_company_id(request.args.get("company_id"))
    if err:
        return jsonify({"ok": False, "error": err, "resultados": []}), 400
    q = (request.args.get("q") or "").strip()
    tipo_documento = (request.args.get("tipo_documento") or "").strip() or None
    empleados_q = (request.args.get("empleados_q") or "").strip() or None
    if not q and not tipo_documento and not empleados_q:
        return jsonify({"ok": True, "resultados": [], "company_id": cid})
    try:
        lim = int(request.args.get("limit") or 40)
    except ValueError:
        lim = 40
    resultados = legajos_service.search_documentos_empresa(
        chatbot.db,
        cid,
        q,
        limit=lim,
        tipo_documento=tipo_documento,
        empleados_q=empleados_q,
    )
    return jsonify({"ok": True, "resultados": resultados, "company_id": cid})


@flask_app.get("/api/legajos/documentos/tipos")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_VIEW,
    message="Sin permiso para ver legajos.",
)
def api_legajos_documentos_tipos():
    """Resumen de cantidad de archivos por tipo (carpeta) en la empresa."""
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible.", "tipos": []}), 503
    cid, err = _legajos_effective_company_id(request.args.get("company_id"))
    if err:
        return jsonify({"ok": False, "error": err, "tipos": []}), 400
    empleados_q = (request.args.get("empleados_q") or "").strip() or None
    tipos = legajos_service.list_documentos_resumen_tipos(
        chatbot.db, cid, empleados_search=empleados_q
    )
    return jsonify({"ok": True, "tipos": tipos, "company_id": cid})


@flask_app.get("/api/legajos/empleados/<empleado_id>/documentos")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_VIEW,
    message="Sin permiso para ver legajos.",
)
def api_legajos_documentos_list(empleado_id):
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible.", "documentos": []}), 503
    emp, err = _legajos_empleado_si_acceso(empleado_id)
    if err:
        return jsonify({"ok": False, "error": err, "documentos": []}), 404
    docs = legajos_service.list_documentos(chatbot.db, emp["id"])
    return jsonify({"ok": True, "documentos": docs, "empleado": emp})


@flask_app.post("/api/legajos/empleados/<empleado_id>/documentos")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_MANAGE,
    message="Sin permiso para subir documentos de legajo.",
)
def api_legajos_documentos_upload(empleado_id):
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible."}), 503
    emp, err = _legajos_empleado_si_acceso(empleado_id)
    if err:
        return jsonify({"ok": False, "error": err}), 404
    bucket = _get_storage_bucket()
    if not bucket:
        return jsonify({"ok": False, "error": "Storage no configurado (Firebase Storage)."}), 501
    files = [f for f in request.files.getlist("file") if f and f.filename]
    if not files:
        return jsonify({"ok": False, "error": "No se envió ningún archivo."}), 400
    tipo = (request.form.get("tipo_documento") or "otro").strip() or "otro"
    user = _current_rrhh_user()
    uname = str((user or {}).get("username") or "").strip()
    documentos: list[dict] = []
    errores: list[dict] = []
    for fs in files:
        meta, up_err = _upload_one_legajo_filestorage(bucket, emp["company_id"], empleado_id, fs)
        orig_name = (fs.filename or "").strip() or (meta or {}).get("filename") or "archivo"
        if up_err or not meta:
            errores.append({"filename": orig_name, "error": up_err or "Error al subir."})
            continue
        ok, row, msg = legajos_service.create_documento(
            chatbot.db,
            empleado_id=empleado_id,
            company_id=emp["company_id"],
            storage_path=meta["storage_path"],
            filename=meta["filename"],
            content_type=meta["content_type"],
            size_bytes=meta["size_bytes"],
            uploaded_by=uname,
            tipo_documento=tipo,
        )
        if not ok:
            errores.append({"filename": orig_name, "error": msg or "No se pudo registrar el documento."})
            continue
        _legajos_audit(
            legajos_service.LEGAJOS_AUDIT_DOCUMENTO_SUBIR,
            emp["company_id"],
            {
                "empleado_id": empleado_id,
                "documento_id": (row or {}).get("id"),
                "filename": (row or {}).get("filename"),
                "tipo": tipo,
            },
        )
        documentos.append(row or {})
    if not documentos and errores:
        first = errores[0].get("error") or "Error al subir."
        return jsonify({"ok": False, "error": first, "errores": errores, "documentos": []}), 400
    body = {
        "ok": True,
        "documentos": documentos,
        "errores": errores,
        "subidos": len(documentos),
    }
    if len(documentos) == 1 and not errores:
        body["documento"] = documentos[0]
    return jsonify(body)


@flask_app.get("/api/legajos/documentos/<documento_id>/link")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_VIEW,
    message="Sin permiso para descargar documentos de legajo.",
)
def api_legajos_documento_link(documento_id):
    doc, err = _legajos_documento_si_acceso(documento_id)
    if err:
        return jsonify({"ok": False, "error": err}), 404
    bucket = _get_storage_bucket()
    if not bucket:
        return jsonify({"ok": False, "error": "Storage no configurado."}), 501
    path = str(doc.get("storage_path") or "").strip()
    if not path:
        return jsonify({"ok": False, "error": "Documento sin ruta de almacenamiento."}), 500
    try:
        blob = bucket.blob(path)
        url = _generate_signed_url(blob, expiration_minutes=15)
        _legajos_audit(
            legajos_service.LEGAJOS_AUDIT_DOCUMENTO_DESCARGAR,
            doc.get("company_id") or "",
            {
                "documento_id": documento_id,
                "empleado_id": doc.get("empleado_id"),
                "filename": doc.get("filename"),
            },
        )
        return jsonify({"ok": True, "url": url, "filename": doc.get("filename") or "archivo"})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@flask_app.delete("/api/legajos/documentos/<documento_id>")
@rrhh_permission_required(
    auth_rrhh.PERM_LEGAJOS_MANAGE,
    message="Sin permiso para eliminar documentos de legajo.",
)
def api_legajos_documento_delete(documento_id):
    doc, err = _legajos_documento_si_acceso(documento_id)
    if err:
        return jsonify({"ok": False, "error": err}), 404
    ok, deleted, msg = legajos_service.delete_documento_record(chatbot.db, documento_id)
    if not ok:
        return jsonify({"ok": False, "error": msg or "No se pudo eliminar."}), 400
    cid_del = str((deleted or {}).get("company_id") or "").strip().lower()
    _legajos_audit(
        legajos_service.LEGAJOS_AUDIT_DOCUMENTO_ELIMINAR,
        cid_del,
        {
            "documento_id": documento_id,
            "empleado_id": (deleted or {}).get("empleado_id"),
            "filename": (deleted or {}).get("filename"),
        },
    )
    path = str((deleted or {}).get("storage_path") or "").strip()
    if path:
        try:
            bucket = _get_storage_bucket()
            if bucket:
                bucket.blob(path).delete()
        except Exception:
            pass
    return jsonify({"ok": True})


@flask_app.get("/api/legajos/carpetas")
def api_legajos_carpetas_list():
    user = _current_rrhh_user()
    if _auth_enabled() and user is None:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    cid = request.args.get("company_id", "").strip().lower()
    if not cid:
        return jsonify({"ok": False, "error": "Falta company_id"}), 400
    carpetas = legajos_service.list_carpetas(chatbot.db, cid)
    return jsonify({"ok": True, "carpetas": carpetas, "is_custom": legajos_service.has_custom_carpetas(chatbot.db, cid)})


@flask_app.post("/api/legajos/carpetas")
def api_legajos_carpetas_create():
    user = _current_rrhh_user()
    if _auth_enabled() and user is None:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    data = request.get_json(silent=True) or {}
    cid = str(data.get("company_id") or "").strip().lower()
    label = str(data.get("label") or "").strip()
    ok, row, msg = legajos_service.create_carpeta(chatbot.db, cid, label)
    if not ok:
        return jsonify({"ok": False, "error": msg}), 400
    return jsonify({"ok": True, "carpeta": row})


@flask_app.post("/api/legajos/carpetas/init-default")
def api_legajos_carpetas_init_default():
    user = _current_rrhh_user()
    if _auth_enabled() and user is None:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    data = request.get_json(silent=True) or {}
    cid = str(data.get("company_id") or "").strip().lower()
    if not cid:
        return jsonify({"ok": False, "error": "Falta company_id"}), 400
    legajos_service.init_carpetas_from_default(chatbot.db, cid)
    carpetas = legajos_service.list_carpetas(chatbot.db, cid)
    return jsonify({"ok": True, "carpetas": carpetas})


@flask_app.delete("/api/legajos/carpetas/<carpeta_id>")
def api_legajos_carpetas_delete(carpeta_id):
    user = _current_rrhh_user()
    if _auth_enabled() and user is None:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    ok, msg = legajos_service.delete_carpeta(chatbot.db, carpeta_id)
    if not ok:
        return jsonify({"ok": False, "error": msg}), 400
    return jsonify({"ok": True})


@flask_app.get("/api/legajos/convenios")
def api_legajos_convenios_list():
    user = _current_rrhh_user()
    if _auth_enabled() and user is None:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    if _auth_enabled() and not _can_config_convenios():
        return jsonify({"ok": False, "error": "Sin permiso para ver convenios."}), 403
    cid = request.args.get("company_id", "").strip().lower()
    if not cid:
        return jsonify({"ok": False, "error": "Falta company_id"}), 400
    convenios = legajos_service.list_convenios(chatbot.db, cid)
    return jsonify({"ok": True, "convenios": convenios})


@flask_app.post("/api/legajos/convenios")
def api_legajos_convenios_create():
    user = _current_rrhh_user()
    if _auth_enabled() and user is None:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    if _auth_enabled() and not _can_config_convenios():
        return jsonify({"ok": False, "error": "Sin permiso para gestionar convenios."}), 403
    data = request.get_json(silent=True) or {}
    cid = str(data.get("company_id") or "").strip().lower()
    nombre = str(data.get("nombre") or "").strip()
    tipos = data.get("tipos_documento") or []
    ok, row, msg = legajos_service.create_convenio(chatbot.db, cid, nombre, tipos)
    if not ok:
        return jsonify({"ok": False, "error": msg}), 400
    return jsonify({"ok": True, "convenio": row})


@flask_app.patch("/api/legajos/convenios/<convenio_id>")
def api_legajos_convenios_update(convenio_id):
    user = _current_rrhh_user()
    if _auth_enabled() and user is None:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    if _auth_enabled() and not _can_config_convenios():
        return jsonify({"ok": False, "error": "Sin permiso para gestionar convenios."}), 403
    data = request.get_json(silent=True) or {}
    nombre = str(data.get("nombre") or "").strip()
    tipos = data.get("tipos_documento") or []
    ok, row, msg = legajos_service.update_convenio(chatbot.db, convenio_id, nombre, tipos)
    if not ok:
        return jsonify({"ok": False, "error": msg}), 400
    return jsonify({"ok": True, "convenio": row})


@flask_app.delete("/api/legajos/convenios/<convenio_id>")
def api_legajos_convenios_delete(convenio_id):
    user = _current_rrhh_user()
    if _auth_enabled() and user is None:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    if _auth_enabled() and not _can_config_convenios():
        return jsonify({"ok": False, "error": "Sin permiso para gestionar convenios."}), 403
    ok, msg = legajos_service.delete_convenio(chatbot.db, convenio_id)
    if not ok:
        return jsonify({"ok": False, "error": msg}), 400
    return jsonify({"ok": True})


def _normalize_phones_for_comunicado(raw_list):
    """Convierte una lista de entradas (números, con espacios/comas) en una lista de números E.164.
    Números argentinos de 10 dígitos (ej. 3515416836) se convierten a +5493515416836."""
    out = []
    seen = set()
    for item in raw_list:
        s = str(item or "").strip()
        if not s:
            continue
        for part in s.replace(",", " ").replace(";", " ").split():
            num = "".join(c for c in part if c.isdigit() or c == "+").strip()
            if not num:
                continue
            if num.startswith("+"):
                num = num.lstrip("+")
            digits = "".join(c for c in num if c.isdigit())
            # Argentina: 10 dígitos sin código país -> agregar 54 y 9 (móvil)
            if len(digits) == 10 and not digits.startswith("54"):
                digits = "54" + "9" + digits
            if not digits:
                continue
            normalized = "+" + digits
            if normalized != "+" and normalized not in seen:
                seen.add(normalized)
                out.append(normalized)
    return out


@flask_app.get("/comunicados")
@rrhh_permission_required(
    auth_rrhh.PERM_COMUNICADOS_SEND,
    message="No tenés permisos para enviar comunicados.",
)
def comunicados_page():
    """Pantalla para enviar comunicados por WhatsApp: elegir empresa, cargar contactos, escribir mensaje."""
    company = _set_company_session(session.get("company_id") or _default_company_id())
    current_user = _current_rrhh_user()
    available_companies = (
        _companies_for_user(current_user) if current_user else _list_companies(include_inactive=False)
    )
    return render_template(
        "comunicados.html",
        auth_enabled=_auth_enabled(),
        rrhh_user=current_user,
        available_companies=available_companies,
        selected_company_id=company.get("company_id"),
        selected_company_name=company.get("company_name"),
    )


COMUNICADOS_CONTACTOS_COLLECTION = "comunicados_contactos"
COMUNICADOS_PROGRAMADOS_COLLECTION = "comunicados_programados"
COMUNICADOS_AUDITORIA_COLLECTION = "comunicados_auditoria"


def _comunicados_audit(action: str, company_id: str, details: dict):
    """Registra una acción de comunicados en Firestore para auditoría."""
    if not chatbot.db:
        return
    try:
        username = (session.get("rrhh_username") or "sistema")
        chatbot.db.collection(COMUNICADOS_AUDITORIA_COLLECTION).add({
            "action": action,
            "company_id": str(company_id or "").strip().lower(),
            "username": username,
            "details": details or {},
            "at": datetime.now(timezone.utc),
        })
    except Exception as e:
        logging.debug("comunicados_audit error: %s", e)


def _get_comunicados_contactos(company_id):
    """Lista de contactos guardados para comunicados (por empresa)."""
    cid = _normalize_company_id(company_id)
    if not cid:
        return []
    if chatbot.db:
        doc = chatbot.db.collection(COMUNICADOS_CONTACTOS_COLLECTION).document(cid).get()
        if doc.exists:
            data = doc.to_dict() or {}
            return list(data.get("contactos") or [])
    return []


def _resolve_whatsapp_contact(phone_raw, company_id, profile_name=""):
    """Devuelve (nombre_display, telefono_limpio) para un número de WhatsApp.

    Prioridad: 1) lista de contactos de la empresa, 2) ProfileName de WA, 3) teléfono limpio.
    """
    phone_clean = re.sub(r"(?i)^whatsapp:", "", phone_raw or "").strip()
    digits_only = re.sub(r"[^\d]", "", phone_clean)

    if company_id and digits_only:
        contacts = _get_comunicados_contactos(company_id)
        for c in contacts:
            c_digits = re.sub(r"[^\d]", "", str(c.get("telefono") or ""))
            if c_digits and (c_digits == digits_only or
                             c_digits[-8:] == digits_only[-8:]):
                return (c.get("nombre") or phone_clean), phone_clean

    if profile_name:
        return profile_name, phone_clean

    return phone_clean, phone_clean


def _add_comunicado_contacto(company_id, nombre, telefono, legajo=None, upsert=False):
    """Agrega o actualiza un contacto de la empresa.

    upsert=False (default): si el teléfono ya existe devuelve "exists".
    upsert=True: si existe actualiza nombre/legajo, devuelve "updated".
    Devuelve "created" al agregar nuevo, False ante error.
    """
    cid = _normalize_company_id(company_id)
    if not cid:
        return False
    nombre = str(nombre or "").strip() or "Sin nombre"
    telefono = str(telefono or "").strip()
    if not telefono:
        return False
    legajo = str(legajo or "").strip() if legajo is not None else ""
    digits_new = re.sub(r"[^\d]", "", telefono)
    if chatbot.db:
        doc_ref = chatbot.db.collection(COMUNICADOS_CONTACTOS_COLLECTION).document(cid)
        doc = doc_ref.get()
        contactos = list((doc.to_dict() or {}).get("contactos") or []) if doc.exists else []
        for i, c in enumerate(contactos):
            c_digits = re.sub(r"[^\d]", "", str(c.get("telefono") or ""))
            if c_digits and c_digits == digits_new:
                if upsert:
                    contactos[i] = {"nombre": nombre, "telefono": c.get("telefono", telefono), "legajo": legajo or c.get("legajo", "")}
                    doc_ref.set({"contactos": contactos}, merge=True)
                    return "updated"
                return "exists"
        contactos.append({"nombre": nombre, "telefono": telefono, "legajo": legajo})
        doc_ref.set({"contactos": contactos}, merge=True)
        return "created"
    return False


@flask_app.get("/api/comunicados/plantilla")
@rrhh_permission_required(
    auth_rrhh.PERM_COMUNICADOS_SEND,
    message="Sin permiso.",
)
def api_comunicados_plantilla():
    """Descarga plantilla CSV (para abrir en Excel: guardar como CSV y subir). Columnas: teléfono, nombre, legajo."""
    from flask import Response
    csv_content = "\ufeffteléfono;nombre;legajo\n=\"+5491112345678\";Ejemplo;12345\n"
    return Response(csv_content, mimetype="text/csv", headers={
        "Content-Disposition": "attachment; filename=plantilla_contactos_comunicados.csv",
    })


@flask_app.get("/api/comunicados/contactos")
@rrhh_permission_required(
    auth_rrhh.PERM_COMUNICADOS_SEND,
    message="Sin permiso.",
)
def api_comunicados_contactos_list():
    """Lista contactos guardados para la empresa (para seleccionar destinatarios)."""
    company_id = _normalize_company_id(request.args.get("company_id") or "")
    contactos = _get_comunicados_contactos(company_id)
    return jsonify({"ok": True, "contactos": contactos})


@flask_app.post("/api/comunicados/contactos")
@rrhh_permission_required(
    auth_rrhh.PERM_COMUNICADOS_SEND,
    message="Sin permiso.",
)
def api_comunicados_contactos_add():
    """Agrega o actualiza un contacto de la empresa.

    Si upsert=true en el body, actualiza nombre/legajo si el teléfono ya existe.
    Si upsert=false (default), devuelve 409 si el número ya existe.
    """
    data = request.get_json(silent=True) or {}
    company_id = _normalize_company_id(data.get("company_id") or "")
    nombre = str(data.get("nombre") or "").strip()
    telefono = str(data.get("telefono") or "").strip()
    legajo = data.get("legajo")
    upsert = bool(data.get("upsert"))
    if not telefono:
        return jsonify({"ok": False, "error": "El teléfono es obligatorio."}), 400
    result = _add_comunicado_contacto(company_id, nombre or "Sin nombre", telefono, legajo=legajo, upsert=upsert)
    if result in ("created", "updated"):
        _comunicados_audit("contacto_" + result, company_id, {"nombre": nombre, "telefono": telefono, "legajo": legajo})
        return jsonify({"ok": True, "result": result, "contactos": _get_comunicados_contactos(company_id)})
    if result == "exists":
        return jsonify({"ok": False, "error": f"Ya existe un contacto con el número {telefono}.", "exists": True}), 409
    return jsonify({"ok": False, "error": "No se pudo guardar."}), 500


def _remove_comunicado_contacto(company_id, telefono):
    """Elimina un contacto de la lista de la empresa por teléfono (misma normalización que envío)."""
    cid = _normalize_company_id(company_id)
    if not cid:
        return False
    telefono = str(telefono or "").strip()
    if not telefono:
        return False
    normalized_input_list = _normalize_phones_for_comunicado([telefono])
    if not normalized_input_list:
        return False
    normalized_input = normalized_input_list[0]
    if chatbot.db:
        doc_ref = chatbot.db.collection(COMUNICADOS_CONTACTOS_COLLECTION).document(cid)
        doc = doc_ref.get()
        contactos = list((doc.to_dict() or {}).get("contactos") or []) if doc.exists else []
        contactos_new = []
        removed = False
        for c in contactos:
            tel_stored = (c.get("telefono") or "").strip()
            norm_stored = _normalize_phones_for_comunicado([tel_stored])
            if norm_stored and norm_stored[0] == normalized_input and not removed:
                removed = True
                continue
            contactos_new.append(c)
        if not removed:
            return False
        doc_ref.set({"contactos": contactos_new}, merge=True)
        return True
    return False


@flask_app.delete("/api/comunicados/contactos")
@rrhh_permission_required(
    auth_rrhh.PERM_COMUNICADOS_SEND,
    message="Sin permiso.",
)
def api_comunicados_contactos_remove():
    """Elimina un contacto de la lista de la empresa (por company_id y telefono)."""
    data = request.get_json(silent=True) or {}
    company_id = _normalize_company_id(data.get("company_id") or "")
    telefono = str(data.get("telefono") or "").strip()
    if not company_id or not telefono:
        return jsonify({"ok": False, "error": "Faltan company_id o teléfono."}), 400
    if _remove_comunicado_contacto(company_id, telefono):
        _comunicados_audit("contacto_eliminado", company_id, {"telefono": telefono})
        return jsonify({"ok": True, "contactos": _get_comunicados_contactos(company_id)})
    return jsonify({"ok": False, "error": "No se encontró el contacto o no se pudo eliminar."}), 404


@flask_app.delete("/api/comunicados/contactos/bulk")
@rrhh_permission_required(
    auth_rrhh.PERM_COMUNICADOS_SEND,
    message="Sin permiso.",
)
def api_comunicados_contactos_bulk_remove():
    """Elimina múltiples contactos de una empresa en una sola operación."""
    data = request.get_json(silent=True) or {}
    company_id = _normalize_company_id(data.get("company_id") or "")
    telefonos = [str(t).strip() for t in (data.get("telefonos") or []) if t]
    if not company_id or not telefonos:
        return jsonify({"ok": False, "error": "Faltan company_id o teléfonos."}), 400
    cid = company_id
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Sin base de datos."}), 500
    normalized_inputs = set()
    for t in telefonos:
        n = _normalize_phones_for_comunicado([t])
        if n:
            normalized_inputs.add(n[0])
    doc_ref = chatbot.db.collection(COMUNICADOS_CONTACTOS_COLLECTION).document(cid)
    doc = doc_ref.get()
    contactos = list((doc.to_dict() or {}).get("contactos") or []) if doc.exists else []
    removed = 0
    contactos_new = []
    for c in contactos:
        tel_stored = (c.get("telefono") or "").strip()
        norm_stored = _normalize_phones_for_comunicado([tel_stored])
        if norm_stored and norm_stored[0] in normalized_inputs:
            removed += 1
        else:
            contactos_new.append(c)
    doc_ref.set({"contactos": contactos_new}, merge=True)
    if removed:
        _comunicados_audit("contacto_bulk_eliminado", cid, {"cantidad": removed, "telefonos": telefonos[:20]})
    return jsonify({"ok": True, "removed": removed, "contactos": _get_comunicados_contactos(cid)})


@flask_app.post("/api/comunicados/programar")
@rrhh_permission_required(
    auth_rrhh.PERM_COMUNICADOS_SEND,
    message="No tenés permisos para enviar comunicados.",
)
def api_comunicados_programar():
    """Guarda un comunicado para enviarse en una fecha/hora futura."""
    data = request.get_json(silent=True) or {}
    company_id = _normalize_company_id(data.get("company_id") or "")
    scheduled_at_str = str(data.get("scheduled_at") or "").strip()
    if not scheduled_at_str:
        return jsonify({"ok": False, "error": "Falta 'scheduled_at'."}), 400
    try:
        # Acepta ISO 8601 con o sin timezone; si no tiene TZ asume UTC
        scheduled_at_str_fixed = scheduled_at_str
        if "T" in scheduled_at_str and "+" not in scheduled_at_str and not scheduled_at_str.endswith("Z"):
            scheduled_at_str_fixed = scheduled_at_str + "Z"
        scheduled_at = datetime.fromisoformat(scheduled_at_str_fixed.replace("Z", "+00:00"))
        if scheduled_at.tzinfo is None:
            scheduled_at = scheduled_at.replace(tzinfo=timezone.utc)
    except ValueError:
        return jsonify({"ok": False, "error": "Formato de fecha inválido. Usá ISO 8601."}), 400
    if scheduled_at <= datetime.now(timezone.utc):
        return jsonify({"ok": False, "error": "La fecha programada debe ser en el futuro."}), 400

    destinatarios_raw = data.get("destinatarios")
    if isinstance(destinatarios_raw, list):
        phones = _normalize_phones_for_comunicado(destinatarios_raw)
    elif isinstance(destinatarios_raw, str):
        phones = _normalize_phones_for_comunicado([destinatarios_raw])
    else:
        phones = []
    if not phones:
        return jsonify({"ok": False, "error": "No hay destinatarios válidos."}), 400

    mensaje = str(data.get("mensaje") or "").strip()
    imagen_url = str(data.get("imagen_url") or "").strip()
    media_urls = [imagen_url] if imagen_url else []
    if not mensaje and not media_urls:
        return jsonify({"ok": False, "error": "Falta el mensaje o imagen."}), 400

    whatsapp_label = str(data.get("whatsapp_label") or "").strip()
    # Resolver número de WA
    whatsapp_phone = str(data.get("whatsapp_phone") or "").strip()
    if not whatsapp_phone and company_id:
        company = _get_company(company_id, include_inactive=False)
        if company:
            nums = company.get("whatsapp_numbers") or []
            if whatsapp_label and nums:
                for line in nums:
                    if (str(line.get("label") or "")).strip().lower() == whatsapp_label.lower():
                        whatsapp_phone = (line.get("phone") or "").strip()
                        break
            if not whatsapp_phone and nums:
                whatsapp_phone = (nums[0].get("phone") or "").strip()
    if not whatsapp_phone:
        whatsapp_phone = os.getenv("TWILIO_WHATSAPP_FROM", "").strip()
    if not whatsapp_phone:
        return jsonify({"ok": False, "error": "WhatsApp no configurado."}), 503

    user = _current_rrhh_user()
    doc = {
        "company_id": company_id,
        "scheduled_at": scheduled_at.isoformat(),
        "mensaje": mensaje,
        "media_urls": media_urls,
        "destinatarios": phones,
        "whatsapp_phone": whatsapp_phone,
        "whatsapp_label": whatsapp_label,
        "estado": "pendiente",
        "created_at": _utc_now().isoformat(),
        "created_by": (user or {}).get("username", ""),
        "result": None,
    }
    if chatbot.db:
        ref = chatbot.db.collection(COMUNICADOS_PROGRAMADOS_COLLECTION).document()
        ref.set(doc)
        doc["id"] = ref.id
    return jsonify({"ok": True, "comunicado": doc})


@flask_app.get("/api/comunicados/programados")
@rrhh_permission_required(
    auth_rrhh.PERM_COMUNICADOS_SEND,
    message="No tenés permisos.",
)
def api_comunicados_programados_list():
    """Lista comunicados programados (pendientes y recientes) de la empresa."""
    company_id = _normalize_company_id(request.args.get("company_id") or "")
    if not company_id or not chatbot.db:
        return jsonify({"ok": True, "comunicados": []})
    try:
        docs = (
            chatbot.db.collection(COMUNICADOS_PROGRAMADOS_COLLECTION)
            .where("company_id", "==", company_id)
            .order_by("scheduled_at", direction="DESCENDING")
            .limit(50)
            .stream()
        )
        result = []
        for d in docs:
            item = d.to_dict() or {}
            item["id"] = d.id
            result.append(item)
    except Exception as exc:
        logging.warning(f"api_comunicados_programados_list: {exc}")
        result = []
    return jsonify({"ok": True, "comunicados": result})


@flask_app.delete("/api/comunicados/programados/<comunicado_id>")
@rrhh_permission_required(
    auth_rrhh.PERM_COMUNICADOS_SEND,
    message="No tenés permisos.",
)
def api_comunicados_programados_cancel(comunicado_id):
    """Cancela un comunicado pendiente."""
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Sin base de datos."}), 503
    ref = chatbot.db.collection(COMUNICADOS_PROGRAMADOS_COLLECTION).document(comunicado_id)
    doc = ref.get()
    if not doc.exists:
        return jsonify({"ok": False, "error": "No encontrado."}), 404
    data = doc.to_dict() or {}
    if data.get("estado") != "pendiente":
        return jsonify({"ok": False, "error": f"No se puede cancelar: estado '{data.get('estado')}'."}), 400
    ref.update({"estado": "cancelado"})
    return jsonify({"ok": True})


@flask_app.post("/api/comunicados/enviar")
@rrhh_permission_required(
    auth_rrhh.PERM_COMUNICADOS_SEND,
    message="No tenés permisos para enviar comunicados.",
)
def api_comunicados_enviar():
    """Envía un comunicado por WhatsApp a la lista de destinatarios (Twilio, en lotes)."""
    data = request.get_json(silent=True) or {}
    company_id = _normalize_company_id(data.get("company_id") or "")
    destinatarios_raw = data.get("destinatarios")
    if isinstance(destinatarios_raw, list):
        phones = _normalize_phones_for_comunicado(destinatarios_raw)
    elif isinstance(destinatarios_raw, str):
        phones = _normalize_phones_for_comunicado([destinatarios_raw])
    else:
        phones = []
    mensaje_raw = str(data.get("mensaje") or "").strip()
    try:
        mensaje = mensaje_raw.encode("utf-8", errors="replace").decode("utf-8")
    except Exception:
        mensaje = "".join(c for c in mensaje_raw if ord(c) < 0xD800 or ord(c) > 0xDFFF)
    imagen_url = (data.get("imagen_url") or data.get("media_url") or "").strip()
    if isinstance(imagen_url, list):
        media_urls = [u.strip() for u in imagen_url if isinstance(u, str) and u.strip()]
    else:
        media_urls = [imagen_url] if imagen_url else []

    if not phones:
        return jsonify({"ok": False, "error": "No hay destinatarios. Pegá números o importá un archivo."}), 400
    if not mensaje and not media_urls:
        return jsonify({"ok": False, "error": "Escribí el texto del comunicado o agregá una imagen (URL)."}), 400

    phone_number_id = (data.get("whatsapp_phone") or "").strip()
    if not phone_number_id and company_id:
        company = _get_company(company_id, include_inactive=False)
        if company:
            nums = company.get("whatsapp_numbers") or []
            label = (data.get("whatsapp_label") or "").strip().lower()
            if label and nums:
                for line in nums:
                    if (str(line.get("label") or "")).strip().lower() == label:
                        phone_number_id = (line.get("phone") or "").strip()
                        break
            if not phone_number_id and nums:
                phone_number_id = (nums[0].get("phone") or "").strip()
    # Detectar si phone_number_id es un ID numérico de Meta
    _pid_raw = str(phone_number_id or "").strip()
    _is_meta_pid_sync = bool(_pid_raw) and _pid_raw.isdigit() and len(_pid_raw) > 10
    # Si no es un ID Meta numérico, preferir META_PHONE_NUMBER_ID si está configurado
    # (el campo "phone" de la empresa puede ser el número de display, no el numeric ID)
    if not _is_meta_pid_sync:
        _meta_pid_fallback = _meta_phone_number_id()
        if _meta_pid_fallback:
            phone_number_id = _meta_pid_fallback
            _is_meta_pid_sync = True
        elif not phone_number_id:
            phone_number_id = os.getenv("TWILIO_WHATSAPP_FROM", "").strip()
    if not phone_number_id:
        return jsonify({"ok": False, "error": "WhatsApp no configurado (falta TWILIO_WHATSAPP_FROM o META_PHONE_NUMBER_ID)."}), 503
    if not _is_meta_pid_sync and not phone_number_id.startswith("whatsapp:"):
        try:
            from twilio_whatsapp import _format_to_whatsapp
            phone_number_id = _format_to_whatsapp(phone_number_id) or phone_number_id
        except Exception:
            pass

    try:
        from whatsapp_broadcast import broadcast_messages, set_send_function as _wb_set_send
    except ImportError:
        return jsonify({"ok": False, "error": "Módulo de envío no disponible."}), 503

    if _is_meta_pid_sync:
        def _meta_send_fn(phone, body=None, media_url=None, phone_number_id=None, **kwargs):
            try:
                urls = media_url if isinstance(media_url, list) else ([media_url] if media_url else [])
                if urls:
                    import requests as _req_bc2
                    img_url = urls[0]
                    _r2 = _req_bc2.get(img_url, timeout=30)
                    if _r2.ok:
                        _mime2 = _r2.headers.get("Content-Type", "image/jpeg").split(";")[0].strip()
                        _fname2 = img_url.split("/")[-1].split("?")[0] or "imagen.jpg"
                        _mid2 = _upload_media_to_meta(_r2.content, _mime2, _fname2, phone_number_id=phone_number_id)
                        if _mid2:
                            _mtype2 = "image" if _mime2.startswith("image/") else "document"
                            _send_meta_whatsapp_media(phone, _mid2, _mtype2, caption=body or None, filename=_fname2 if _mtype2 == "document" else None, phone_number_id=phone_number_id)
                            return True
                if body:
                    return bool(_send_meta_whatsapp(phone, body, phone_number_id=phone_number_id))
                return False
            except Exception as e:
                logger.warning("broadcast meta: error a %s: %s", phone, e)
                return False
        _wb_set_send(_meta_send_fn)

    result = broadcast_messages(
        phone_list=phones,
        body_text=mensaje or None,
        phone_number_id=phone_number_id,
        media_url=media_urls if media_urls else None,
    )
    sent = result.get("sent", 0)
    failed = result.get("failed", 0)
    total = result.get("total", 0)
    resp = {
        "ok": True,
        "sent": sent,
        "failed": failed,
        "total": total,
        "batches_used": result.get("batches_used", 0),
    }
    if failed > 0:
        try:
            from twilio_whatsapp import last_twilio_error as twilio_err
            err_text = (twilio_err or "").strip()
            err_lower = err_text.lower()
            if "50 daily" in err_lower or ("exceeded" in err_lower and "limit" in err_lower) or "63038" in err_lower:
                resp["error_detail"] = (
                    "Tu cuenta Twilio alcanzó el límite de 50 mensajes por día (cuenta trial). "
                    "Mañana se reinicia el límite, o pasate a una cuenta de pago en twilio.com para enviar más."
                )
            else:
                resp["error_detail"] = (
                    "Algunos o todos los envíos fallaron. Revisá TWILIO_WHATSAPP_FROM, "
                    "que los números tengan WhatsApp y que la URL de la imagen sea pública (Twilio tiene que poder descargarla)."
                )
                if err_text:
                    resp["error_detail"] += " Error de Twilio: " + err_text[:400]
        except Exception:
            resp["error_detail"] = (
                "Algunos o todos los envíos fallaron. Revisá en el servidor TWILIO_WHATSAPP_FROM, "
                "TWILIO_ACCOUNT_SID y TWILIO_AUTH_TOKEN. Ver consola del servidor para el error."
            )
    return jsonify(resp)


@flask_app.post("/api/comunicados/enviar-stream")
@rrhh_permission_required(
    auth_rrhh.PERM_COMUNICADOS_SEND,
    message="No tenés permisos para enviar comunicados.",
)
def api_comunicados_enviar_stream():
    """Envía comunicado en lotes con progreso en tiempo real (Server-Sent Events)."""
    import json as _json
    from flask import stream_with_context, Response as FlaskResponse

    data = request.get_json(silent=True) or {}
    company_id = _normalize_company_id(data.get("company_id") or "")
    destinatarios_raw = data.get("destinatarios")
    if isinstance(destinatarios_raw, list):
        phones = _normalize_phones_for_comunicado(destinatarios_raw)
    elif isinstance(destinatarios_raw, str):
        phones = _normalize_phones_for_comunicado([destinatarios_raw])
    else:
        phones = []
    mensaje_raw = str(data.get("mensaje") or "").strip()
    try:
        mensaje = mensaje_raw.encode("utf-8", errors="replace").decode("utf-8")
    except Exception:
        mensaje = "".join(c for c in mensaje_raw if ord(c) < 0xD800 or ord(c) > 0xDFFF)
    imagen_url = (data.get("imagen_url") or data.get("media_url") or "").strip()
    media_urls = [imagen_url] if imagen_url else []

    if not phones:
        def _err():
            yield 'data: ' + _json.dumps({"ok": False, "error": "No hay destinatarios."}) + "\n\n"
        return FlaskResponse(stream_with_context(_err()), mimetype="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})
    if not mensaje and not media_urls:
        def _err2():
            yield 'data: ' + _json.dumps({"ok": False, "error": "Escribí el mensaje o agregá una imagen."}) + "\n\n"
        return FlaskResponse(stream_with_context(_err2()), mimetype="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

    phone_number_id = (data.get("whatsapp_phone") or "").strip()
    if not phone_number_id and company_id:
        company = _get_company(company_id, include_inactive=False)
        if company:
            nums = company.get("whatsapp_numbers") or []
            label = (data.get("whatsapp_label") or "").strip().lower()
            if label and nums:
                for line in nums:
                    if (str(line.get("label") or "")).strip().lower() == label:
                        phone_number_id = (line.get("phone") or "").strip()
                        break
            if not phone_number_id and nums:
                phone_number_id = (nums[0].get("phone") or "").strip()
    # Detectar si phone_number_id es un ID numérico de Meta
    _pid_is_meta = str(phone_number_id or "").strip().isdigit() and len(str(phone_number_id or "").strip()) > 10
    # Si no es un ID Meta numérico, preferir META_PHONE_NUMBER_ID si está configurado
    if not _pid_is_meta:
        _meta_pid_stream = _meta_phone_number_id()
        if _meta_pid_stream:
            phone_number_id = _meta_pid_stream
            _pid_is_meta = True
        elif not phone_number_id:
            phone_number_id = os.getenv("TWILIO_WHATSAPP_FROM", "").strip()
    if not _pid_is_meta and phone_number_id and not phone_number_id.startswith("whatsapp:"):
        try:
            from twilio_whatsapp import _format_to_whatsapp
            phone_number_id = _format_to_whatsapp(phone_number_id) or phone_number_id
        except Exception:
            pass

    def generate():
        import json as _json2
        try:
            import whatsapp_broadcast as wb

            # Detectar si phone_number_id es un ID numérico de Meta (ej. "1078605635336424")
            # vs un número Twilio (ej. "whatsapp:+14155238886")
            _pid = str(phone_number_id or "").strip()
            _is_meta_pid = bool(_pid) and _pid.isdigit() and len(_pid) > 10

            if _is_meta_pid:
                def _send_fn(phone, body=None, media_url=None, phone_number_id=None, **kwargs):
                    try:
                        urls = media_url if isinstance(media_url, list) else ([media_url] if media_url else [])
                        if urls:
                            import requests as _req_bc
                            img_url = urls[0]
                            _r = _req_bc.get(img_url, timeout=30)
                            if _r.ok:
                                import mimetypes as _mt
                                _mime = _r.headers.get("Content-Type", "image/jpeg").split(";")[0].strip()
                                _fname = img_url.split("/")[-1].split("?")[0] or "imagen.jpg"
                                _mid = _upload_media_to_meta(_r.content, _mime, _fname, phone_number_id=phone_number_id)
                                if _mid:
                                    _mtype = "image" if _mime.startswith("image/") else "document"
                                    _send_meta_whatsapp_media(phone, _mid, _mtype, caption=body or None, filename=_fname if _mtype == "document" else None, phone_number_id=phone_number_id)
                                    return True
                        if body:
                            return bool(_send_meta_whatsapp(phone, body, phone_number_id=phone_number_id))
                        return False
                    except Exception as e:
                        logger.warning("broadcast meta: error enviando a %s: %s", phone, e)
                        return False
            else:
                from twilio_whatsapp import send_one as twilio_send_one
                def _send_fn(phone, body=None, media_url=None, phone_number_id=None, **kwargs):
                    try:
                        result = twilio_send_one(phone, body=body, media_url=media_url[0] if isinstance(media_url, list) and media_url else (media_url or None), phone_number_id=phone_number_id)
                        return bool(result)
                    except Exception as e:
                        logger.warning("broadcast: error enviando a %s: %s", phone, e)
                        return False

            wb.set_send_function(_send_fn)

            progress_events = []

            def on_progress(sent, failed, total, batch_num, total_batches, waiting, wait_remaining):
                progress_events.append({
                    "ok": True, "done": False,
                    "sent": sent, "failed": failed, "total": total,
                    "batch": batch_num, "batches": total_batches,
                    "waiting": waiting, "wait_remaining": wait_remaining,
                })

            import threading
            result_holder = {}
            def run():
                result_holder["result"] = wb.broadcast_messages(
                    phone_list=phones,
                    body_text=mensaje or None,
                    media_url=media_urls if media_urls else None,
                    phone_number_id=phone_number_id,
                    on_progress=on_progress,
                )
            t = threading.Thread(target=run, daemon=True)
            t.start()

            yield 'data: ' + _json2.dumps({"ok": True, "done": False, "started": True, "total": len(phones)}) + "\n\n"

            while t.is_alive() or progress_events:
                while progress_events:
                    evt = progress_events.pop(0)
                    yield 'data: ' + _json2.dumps(evt) + "\n\n"
                if t.is_alive():
                    import time as _time
                    _time.sleep(0.1)

            r = result_holder.get("result", {})
            yield 'data: ' + _json2.dumps({
                "ok": True, "done": True,
                "sent": r.get("sent", 0), "failed": r.get("failed", 0),
                "total": r.get("total", len(phones)),
                "batches": r.get("batches_used", 0),
            }) + "\n\n"

        except Exception as e:
            logger.exception("api_comunicados_enviar_stream: error: %s", e)
            yield 'data: ' + _json2.dumps({"ok": False, "error": str(e)}) + "\n\n"

    return FlaskResponse(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@flask_app.get("/estadisticas")
@rrhh_auth_required
def stats_page():
    if not _can_view_stats():
        return ("No tenés permisos para ver estadísticas.", 403)
    return render_template("stats.html")


@flask_app.get("/preferencias")
@rrhh_auth_required
def preferencias_page():
    """Preferencias de uso: modo oscuro/claro, empresa activa, autocierre y reglas del chat."""
    if not _can_manage_preferences():
        return ("No tenés permisos para gestionar preferencias.", 403)
    return render_template("preferencias.html")


def _strip_leading_articles(text):
    """Elimina artículos iniciales comunes para mejorar el matching (ej: 'la central' → 'central')."""
    return re.sub(r'^(el |la |los |las |de |del |un |una )', '', text.strip(), flags=re.IGNORECASE).strip()


def _process_chat_turn(mensaje_trim):
    """Procesa un mensaje del chat y devuelve el dict de respuesta (reply, quick_actions, etc.). Usado por /api/chat y por el webhook de WhatsApp."""
    _apply_company_branding(_read_general_settings())
    mensaje = mensaje_trim

    # Handoff activo: validar que el handoff realmente exista y no esté cerrado
    _handoff_id = _get_handoff_session_id()
    if _handoff_id:
        _handoff_doc = _fetch_handoff(_handoff_id)
        _handoff_estado = str((_handoff_doc or {}).get("estado") or "").strip().lower()
        if not _handoff_doc or _handoff_estado == HANDOFF_STATUS_CLOSED:
            # Handoff cerrado o no existe — limpiar sesión y seguir flujo normal
            _clear_handoff_session()
            limpiar_estado_conversacion()
        else:
            result = responder_chat(mensaje_trim)
            return result if "ok" in result else {"ok": True, **result}

    step = _chat_context_step()

    # ── Paso DNI: solo en WhatsApp, primer contacto ───────────────────────────
    if step == CHAT_CONTEXT_STEP_DNI:
        mensaje_norm_dni = chatbot.normalizar_texto(mensaje_trim)
        # El colaborador puede omitir la identificación
        if mensaje_norm_dni in ("omitir", "saltar", "skip", "no tengo", "sin dni"):
            _sess()["chat_context_step"] = CHAT_CONTEXT_STEP_COMPANY
            opciones = _construir_acciones_empresas(limite=8)
            nombres = [a.get("label") or a.get("value") for a in opciones if a.get("label") or a.get("value")]
            reply = "Sin problema. ¿Con qué empresa estás relacionado/a?"
            if nombres:
                reply += "\n\n" + "\n".join(nombres) + "\n\nEscribí el número o el nombre."
            return {"ok": True, "reply": reply, "await_feedback": False, "end_session": False, "quick_actions": opciones, "handoff_active": False}
        # Intentar leer el DNI del mensaje
        import legajos_service as _ls
        dni_ingresado = re.sub(r"[^\d]", "", mensaje_trim).strip()
        if len(dni_ingresado) >= 6:
            # Buscar en todas las empresas configuradas
            _found_emp = None
            if chatbot.db:
                try:
                    snaps = chatbot.db.collection(_ls.LEGAJOS_EMPLEADOS_COLLECTION)\
                        .where("dni", "==", dni_ingresado).limit(1).get()
                    for snap in snaps:
                        _found_emp = _ls.empleado_from_snap(snap)
                        break
                except Exception as _e:
                    logger.warning("DNI lookup error: %s", _e)
            if _found_emp:
                emp_id = _found_emp.get("id") or ""
                emp_conv = (_found_emp.get("convenio") or "").strip().lower()
                emp_empresa = (_found_emp.get("company_id") or "").strip()
                emp_nombre = (_found_emp.get("nombre_completo") or _found_emp.get("nombre") or "").strip()
                _sess()["wa_empleado_id"] = emp_id
                _sess()["wa_convenio"] = emp_conv
                _sess()["wa_nombre"] = emp_nombre
                # Pre-cargar empresa si está configurada
                if emp_empresa and _get_company(emp_empresa):
                    _set_chat_context_company(emp_empresa)
                    company_obj = _get_company(emp_empresa)
                    branches = _get_branches_for_company(company_obj)
                    areas = _get_all_areas_for_company(company_obj)
                    if not branches and not areas:
                        _set_chat_context_area("")
                        _sess()["chat_context_step"] = CHAT_CONTEXT_STEP_READY
                        saludo = f"¡Hola {emp_nombre or 'colaborador'}! Ya te identifiqué. ¿En qué puedo ayudarte hoy?"
                        company_obj2 = _current_company()
                        temas_map2 = construir_temas_map(
                            company_id=emp_empresa,
                            temas_habilitados=(company_obj2 or {}).get("temas_habilitados") or [],
                        )
                        qa2 = construir_acciones_menu(temas_map2, limite=6, permitir_hablar_con_humano=(company_obj2 or {}).get("permitir_hablar_con_humano", True))
                        # Guardar identidad para futuros mensajes
                        _wa_phone = getattr(g, "whatsapp_phone", None)
                        if _wa_phone:
                            _save_whatsapp_identity(_wa_phone, emp_id, emp_conv, emp_empresa, emp_nombre)
                        return {"ok": True, "reply": saludo, "await_feedback": False, "end_session": False, "quick_actions": qa2, "handoff_active": False}
                    elif branches:
                        _sess()["chat_context_step"] = CHAT_CONTEXT_STEP_BRANCH
                        qa_b = _construir_acciones_sucursales(emp_empresa, limite=8)
                        nombres_b = [a.get("label") or a.get("value") for a in qa_b if a.get("label") or a.get("value")]
                        reply_b = f"¡Hola {emp_nombre or 'colaborador'}! ¿De qué sucursal sos?"
                        if nombres_b:
                            reply_b += "\n\n" + "\n".join(nombres_b) + "\n\nEscribí el número o el nombre."
                        _wa_phone = getattr(g, "whatsapp_phone", None)
                        if _wa_phone:
                            _save_whatsapp_identity(_wa_phone, emp_id, emp_conv, emp_empresa, emp_nombre)
                        return {"ok": True, "reply": reply_b, "await_feedback": False, "end_session": False, "quick_actions": qa_b, "handoff_active": False}
                # Guardar identidad aunque la empresa no matchee
                _wa_phone = getattr(g, "whatsapp_phone", None)
                if _wa_phone:
                    _save_whatsapp_identity(_wa_phone, emp_id, emp_conv, emp_empresa, emp_nombre)
                # Continuar con selección de empresa
                _sess()["chat_context_step"] = CHAT_CONTEXT_STEP_COMPANY
                opciones = _construir_acciones_empresas(limite=8)
                nombres = [a.get("label") or a.get("value") for a in opciones if a.get("label") or a.get("value")]
                reply = f"👋 ¡Hola {emp_nombre or 'colaborador'}! Soy {BOT_NAME}, tu asistente de RRHH. 😊 ¿Con qué empresa estás relacionado/a?"
                if nombres:
                    reply += "\n\n" + "\n".join(nombres) + "\n\nEscribí el número o el nombre."
                return {"ok": True, "reply": reply, "await_feedback": False, "end_session": False, "quick_actions": opciones, "handoff_active": False}
            # DNI no encontrado
            return {"ok": True, "reply": "No encontré ese DNI en el sistema. Verificá el número e intentá de nuevo, o escribí *omitir* para continuar sin identificarte.", "await_feedback": False, "end_session": False, "quick_actions": [], "handoff_active": False}
        # Mensaje que no parece un DNI
        if chatbot.es_saludo(mensaje_norm_dni):
            return {"ok": True, "reply": f"👋 ¡Hola! Soy {BOT_NAME}, tu asistente de RRHH. 😊 Para ayudarte mejor, necesito identificarte. Por favor ingresá tu *DNI* (solo números), o escribí *omitir* para continuar sin identificarte.", "await_feedback": False, "end_session": False, "quick_actions": [], "handoff_active": False}
        return {"ok": True, "reply": f"Soy {BOT_NAME}, tu asistente de RRHH. Por favor ingresá tu *DNI* (solo números), o escribí *omitir* para continuar sin identificarte.", "await_feedback": False, "end_session": False, "quick_actions": [], "handoff_active": False}

    if step == CHAT_CONTEXT_STEP_COMPANY:
        mensaje_norm = chatbot.normalizar_texto(mensaje_trim)
        # Saludo durante el paso de empresa — responder amigablemente y volver a preguntar
        if chatbot.es_saludo(mensaje_norm):
            opciones = _construir_acciones_empresas(limite=8)
            nombres = [a.get("label") or a.get("value") for a in opciones if a.get("label") or a.get("value")]
            reply = f"👋 ¡Hola! Soy {BOT_NAME}, tu asistente de RRHH. 😊 ¿Con qué empresa estás relacionado/a?"
            if nombres:
                reply += "\n\n" + "\n".join(nombres) + "\n\nEscribí el número o el nombre."
            return {"ok": True, "reply": reply, "await_feedback": False, "end_session": False, "quick_actions": opciones, "handoff_active": False}
        if chatbot.solicita_contacto_rrhh(mensaje_norm):
            opciones = _construir_acciones_empresas(limite=8)
            nombres = [a.get("label") or a.get("value") for a in opciones if a.get("label") or a.get("value")]
            reply = "Para hablar con un agente primero elegí tu empresa en el menú."
            if nombres:
                reply += "\n\nMenú:\n" + "\n".join(nombres) + "\n\nEscribí el número o el nombre."
            return {
                "ok": True,
                "reply": reply,
                "await_feedback": False,
                "end_session": False,
                "quick_actions": opciones,
                "handoff_active": False,
            }
        cid, company = _resolve_message_to_company(mensaje_trim)
        if cid and company:
            _set_chat_context_company(cid)
            company_for_next = _get_company(cid, include_inactive=False)
            branches = _get_branches_for_company(company_for_next)
            areas = _get_all_areas_for_company(company_for_next)
            if branches:
                quick_actions = _construir_acciones_sucursales(cid, limite=8)
                nombres_b = [a.get("label") or a.get("value") for a in quick_actions if a.get("label") or a.get("value")]
                reply = f"¡Perfecto, {company.get('company_name') or cid}! ¿Y de qué sucursal sos?"
                if nombres_b:
                    reply += "\n\n" + "\n".join(nombres_b) + "\n\nEscribí el número o el nombre."
            elif areas:
                quick_actions = _construir_acciones_areas(cid, limite=8)
                nombres_a = [a.get("label") or a.get("value") for a in quick_actions if a.get("label") or a.get("value")]
                reply = f"¡Perfecto, {company.get('company_name') or cid}! ¿Y a qué área pertenecés?"
                if nombres_a:
                    reply += "\n\n" + "\n".join(nombres_a) + "\n\nEscribí el número o el nombre."
            else:
                _set_chat_context_area("")
                _sess()["chat_context_step"] = CHAT_CONTEXT_STEP_READY
                reply = "¡Todo listo! ¿En qué puedo ayudarte hoy? Escribí tu consulta o elegí una opción."
                company = _current_company()
                temas_map = construir_temas_map(
                    company_id=cid,
                    temas_habilitados=(company or {}).get("temas_habilitados") or [],
                )
                quick_actions = construir_acciones_menu(
                    temas_map, limite=6,
                    permitir_hablar_con_humano=(company or {}).get("permitir_hablar_con_humano", True),
                )
            return {
                "ok": True,
                "reply": reply,
                "await_feedback": False,
                "end_session": False,
                "quick_actions": quick_actions,
                "handoff_active": False,
            }
        opciones = _construir_acciones_empresas(limite=8)
        nombres = [a.get("label") or a.get("value") for a in opciones if a.get("label") or a.get("value")]
        # Si el mensaje es largo y claramente no es un nombre de empresa, dar mensaje más amigable
        mensaje_norm_fb = chatbot.normalizar_texto(mensaje_trim)
        parece_consulta = len(mensaje_trim.split()) >= 3 and not any(
            chatbot.normalizar_texto(item.get("company_name") or item.get("company_id") or "") in mensaje_norm_fb
            for item in _list_companies(include_inactive=False)
        )
        if parece_consulta:
            intro = "Para ayudarte, primero necesito saber con qué empresa estás relacionado/a."
        else:
            intro = "No encontré esa empresa."
        if nombres:
            reply = intro + "\nMenú:\n" + "\n".join(nombres) + "\n\nEscribí el número o el nombre."
        else:
            reply = intro + " Elegí una de las opciones o escribí el nombre correcto."
        return {
            "ok": True,
            "reply": reply,
            "await_feedback": False,
            "end_session": False,
            "quick_actions": opciones,
            "handoff_active": False,
        }

    if step == CHAT_CONTEXT_STEP_BRANCH:
        ctx_cid = _sess().get("chat_context_company_id") or (_current_company() or {}).get("company_id")
        mensaje_norm_b = chatbot.normalizar_texto(mensaje_trim)
        # Si por alguna razón llega "menu" / "ver menú completo" en el step de sucursal, volver al menú de temas
        _MENU_CMDS = {"menu", "ver menu", "ver menu completo", "menu completo"}
        if mensaje_norm_b in _MENU_CMDS:
            branch_saved = _sess().get("chat_context_branch") or ""
            company_tmp = _set_company_session(ctx_cid)
            permitir_tmp = (company_tmp or {}).get("permitir_hablar_con_humano", True)
            temas_tmp = construir_temas_map(company_id=ctx_cid, temas_habilitados=(company_tmp or {}).get("temas_habilitados") or [])
            if branch_saved:
                _sess()["chat_context_step"] = CHAT_CONTEXT_STEP_READY
            return {"ok": True, "reply": construir_menu_texto(temas_tmp, permitir_hablar_con_humano=permitir_tmp), "await_feedback": False, "end_session": False, "quick_actions": construir_acciones_menu(temas_tmp, permitir_hablar_con_humano=permitir_tmp), "handoff_active": False}
        # Saludo durante el paso de sucursal — repetir la pregunta con opciones
        if chatbot.es_saludo(mensaje_norm_b):
            opciones_b = _construir_acciones_sucursales(ctx_cid, limite=8)
            nombres_b = [a.get("label") or a.get("value") for a in opciones_b if a.get("label") or a.get("value")]
            reply_b = "¡Hola! ¿De qué sucursal sos?"
            if nombres_b:
                reply_b += "\n\n" + "\n".join(nombres_b) + "\n\nEscribí el número o el nombre."
            return {"ok": True, "reply": reply_b, "await_feedback": False, "end_session": False, "quick_actions": opciones_b, "handoff_active": False}
        if chatbot.solicita_contacto_rrhh(mensaje_norm_b):
            opciones_b = _construir_acciones_sucursales(ctx_cid, limite=8)
            nombres_b = [a.get("label") or a.get("value") for a in opciones_b if a.get("label") or a.get("value")]
            reply_b = "Para hablar con un agente primero elegí tu sucursal en el menú."
            if nombres_b:
                reply_b += "\n\nMenú:\n" + "\n".join(nombres_b) + "\n\nEscribí el número o el nombre."
            return {"ok": True, "reply": reply_b, "await_feedback": False, "end_session": False, "quick_actions": opciones_b, "handoff_active": False}
        _stripped1 = _strip_leading_articles(mensaje_trim)
        _stripped2 = _strip_leading_articles(_stripped1)
        branch = (_resolve_message_to_branch(mensaje_trim, ctx_cid)
                  or _resolve_message_to_branch(_stripped1, ctx_cid)
                  or _resolve_message_to_branch(_stripped2, ctx_cid))
        if branch:
            _set_chat_context_branch(branch)
            company = _set_company_session(ctx_cid)
            company_for_area = _get_company(ctx_cid, include_inactive=False)
            areas = _get_areas_for_branch(company_for_area, branch)
            if areas:
                quick_actions = _construir_acciones_areas(ctx_cid, limite=8, branch=branch)
                nombres_a2 = [a.get("label") or a.get("value") for a in quick_actions if a.get("label") or a.get("value")]
                reply = f"¡Genial, {branch}! ¿Y a qué área pertenecés?"
                if nombres_a2:
                    reply += "\n\n" + "\n".join(nombres_a2) + "\n\nEscribí el número o el nombre."
            else:
                _set_chat_context_area("")
                _sess()["chat_context_step"] = CHAT_CONTEXT_STEP_READY
                settings = _apply_company_branding(_read_general_settings())
                company_name = (settings.get("company_name") or "Empresa").strip() or "Empresa"
                hr_display = (settings.get("hr_team_name") or "Atención").strip() or "Atención"
                if hr_display.upper() == "RRHH":
                    hr_display = "Atención"
                if hr_display and hr_display.lower() != company_name.lower() and hr_display.lower() != "atención":
                    asistente_label = f"de {hr_display} de {company_name}"
                else:
                    asistente_label = f"de {company_name}"
                _bot = "Debo"
                reply = (
                    f"👋 ¡Hola! Soy {_bot}, tu asistente {asistente_label} (sucursal: {branch}). 😊\n"
                    "Estoy acá para ayudarte. ¿Sobre qué tema querés consultar?"
                )
                permitir = (company or {}).get("permitir_hablar_con_humano", True)
                temas_map = construir_temas_map(
                    company_id=ctx_cid,
                    temas_habilitados=(company or {}).get("temas_habilitados") or [],
                )
                quick_actions = construir_acciones_menu(temas_map, limite=6, permitir_hablar_con_humano=permitir)
            return {
                "ok": True,
                "reply": reply,
                "await_feedback": False,
                "end_session": False,
                "quick_actions": quick_actions,
                "handoff_active": False,
            }
        opciones_branch = _construir_acciones_sucursales(ctx_cid, limite=8)
        nombres_branch = [a.get("label") or a.get("value") for a in opciones_branch if a.get("label") or a.get("value")]
        if nombres_branch:
            reply = f"No encontré esa sucursal.\nMenú:\n" + "\n".join(nombres_branch) + "\n\nEscribí el número o el nombre."
        else:
            reply = "No encontré esa sucursal. Elegí una de las opciones."
        return {
            "ok": True,
            "reply": reply,
            "await_feedback": False,
            "end_session": False,
            "quick_actions": opciones_branch,
            "handoff_active": False,
        }

    if step == CHAT_CONTEXT_STEP_AREA:
        ctx_cid = _sess().get("chat_context_company_id") or (_current_company() or {}).get("company_id")
        ctx_branch = _sess().get("chat_context_branch") or None
        mensaje_norm_a = chatbot.normalizar_texto(mensaje_trim)
        # Si llega "menu" / "ver menú completo" en el step de área, ir directo al menú de temas
        _MENU_CMDS_A = {"menu", "ver menu", "ver menu completo", "menu completo"}
        if mensaje_norm_a in _MENU_CMDS_A:
            company_tmp = _set_company_session(ctx_cid)
            permitir_tmp = (company_tmp or {}).get("permitir_hablar_con_humano", True)
            temas_tmp = construir_temas_map(company_id=ctx_cid, temas_habilitados=(company_tmp or {}).get("temas_habilitados") or [])
            _sess()["chat_context_step"] = CHAT_CONTEXT_STEP_READY
            return {"ok": True, "reply": construir_menu_texto(temas_tmp, permitir_hablar_con_humano=permitir_tmp), "await_feedback": False, "end_session": False, "quick_actions": construir_acciones_menu(temas_tmp, permitir_hablar_con_humano=permitir_tmp), "handoff_active": False}
        # Saludo durante el paso de área — repetir la pregunta con opciones
        if chatbot.es_saludo(mensaje_norm_a):
            opciones_a = _construir_acciones_areas(ctx_cid, limite=8, branch=ctx_branch)
            nombres_a = [a.get("label") or a.get("value") for a in opciones_a if a.get("label") or a.get("value")]
            reply_a = "¡Hola! ¿A qué área pertenecés?"
            if nombres_a:
                reply_a += "\n\n" + "\n".join(nombres_a) + "\n\nEscribí el número o el nombre."
            return {"ok": True, "reply": reply_a, "await_feedback": False, "end_session": False, "quick_actions": opciones_a, "handoff_active": False}
        # Intentar resolver el área primero (puede ser "rrhh", "ventas", etc.)
        area = _resolve_message_to_area(mensaje_trim, ctx_cid, branch=ctx_branch) or _resolve_message_to_area(_strip_leading_articles(mensaje_trim), ctx_cid, branch=ctx_branch)
        # Solo interceptar "hablar con agente" si el mensaje NO resuelve a un área válida
        if not area and chatbot.solicita_contacto_rrhh(mensaje_norm_a):
            opciones_a = _construir_acciones_areas(ctx_cid, limite=8, branch=ctx_branch)
            nombres_a = [a.get("label") or a.get("value") for a in opciones_a if a.get("label") or a.get("value")]
            reply_a = "Para hablar con un agente primero elegí tu área en el menú."
            if nombres_a:
                reply_a += "\n\nMenú:\n" + "\n".join(nombres_a) + "\n\nEscribí el número o el nombre."
            return {"ok": True, "reply": reply_a, "await_feedback": False, "end_session": False, "quick_actions": opciones_a, "handoff_active": False}
        if area:
            _set_chat_context_area(area)
            company = _set_company_session(ctx_cid)
            permitir = (company or {}).get("permitir_hablar_con_humano", True)
            temas_habilitados = (company or {}).get("temas_habilitados") or []
            temas_map = construir_temas_map(company_id=ctx_cid, temas_habilitados=temas_habilitados)
            settings = _apply_company_branding(_read_general_settings())
            company_name = (settings.get("company_name") or "Empresa").strip() or "Empresa"
            hr_display = (settings.get("hr_team_name") or "Atención").strip() or "Atención"
            if hr_display.upper() == "RRHH":
                hr_display = "Atención"
            if hr_display and hr_display.lower() != company_name.lower() and hr_display.lower() != "atención":
                asistente_label = f"de {hr_display} de {company_name}"
            else:
                asistente_label = f"de {company_name}"
            _bot = "Debo"
            reply = (
                f"👋 ¡Hola! Soy {_bot}, tu asistente {asistente_label} (área: {area}). 😊\n"
                "Estoy acá para ayudarte. ¿Sobre qué tema querés consultar?"
            )
            return {
                "ok": True,
                "reply": reply,
                "await_feedback": False,
                "end_session": False,
                "quick_actions": construir_acciones_menu(temas_map, limite=6, permitir_hablar_con_humano=permitir),
                "handoff_active": False,
            }
        opciones_area = _construir_acciones_areas(ctx_cid, limite=8, branch=ctx_branch)
        nombres_area = [a.get("label") or a.get("value") for a in opciones_area if a.get("label") or a.get("value")]
        if nombres_area:
            reply = f"No encontré ese área.\nMenú:\n" + "\n".join(nombres_area) + "\n\nEscribí el número o el nombre."
        else:
            reply = "No encontré ese área. Elegí una de las opciones."
        return {
            "ok": True,
            "reply": reply,
            "await_feedback": False,
            "end_session": False,
            "quick_actions": opciones_area,
            "handoff_active": False,
        }

    if step == CHAT_CONTEXT_STEP_READY:
        mensaje_norm_ready = chatbot.normalizar_texto(mensaje_trim)
        quiere_cambiar = getattr(chatbot, "solicita_cambiar_empresa", None) and chatbot.solicita_cambiar_empresa(mensaje_norm_ready)
        if not quiere_cambiar:
            cid_actual = _normalize_company_id(_sess().get("company_id") or _default_company_id())
            # Solo detectar cambio de empresa si el mensaje NO es un número puro (que puede ser selección de menú de temas)
            if _parse_menu_number(mensaje_trim) is None:
                cid_otro, company_otra = _resolve_message_to_company(mensaje_trim)
                if cid_otro and company_otra and cid_otro != cid_actual:
                    quiere_cambiar = True
        if quiere_cambiar:
            _sess()["chat_context_step"] = CHAT_CONTEXT_STEP_COMPANY
            _sess().pop("chat_context_company_id", None)
            _sess().pop("chat_context_branch", None)
            _sess().pop("chat_context_area", None)
            _sess().pop("company_id", None)
            _sess().pop("company_name", None)
            opciones = _construir_acciones_empresas(limite=8)
            nombres = [a.get("label") or a.get("value") for a in opciones if a.get("label") or a.get("value")]
            reply = "Sin problema. Elegí la empresa con la que querés hablar."
            if nombres:
                reply += "\n\nMenú:\n" + "\n".join(nombres) + "\n\nEscribí el número o el nombre."
            return {
                "ok": True,
                "reply": reply,
                "await_feedback": False,
                "end_session": False,
                "quick_actions": opciones,
                "handoff_active": False,
            }

    company = _set_company_session(_sess().get("company_id") or _default_company_id())
    mensaje_norm = chatbot.normalizar_texto(mensaje)
    handoff_before = bool(_get_handoff_session_id())
    log_asistente_input = not (
        handoff_before
        and mensaje_norm not in HANDOFF_POLL_COMMANDS
        and mensaje_norm not in HANDOFF_END_COMMANDS
    )
    if log_asistente_input:
        input_conversation_id = _get_handoff_session_id() or _session_chat_id()
        _add_chat_history(
            conversation_id=input_conversation_id,
            remitente="colaborador",
            texto=mensaje,
            canal="asistente",
            metadata={"source": "api_chat", "company_id": company.get("company_id")},
        )

    payload = responder_chat(mensaje)
    output_conversation_id = _get_handoff_session_id() or _session_chat_id()
    _add_chat_history(
        conversation_id=output_conversation_id,
        remitente="bot",
        texto=payload.get("reply", ""),
        canal="asistente",
        metadata={
            "await_feedback": bool(payload.get("await_feedback")),
            "handoff_active": bool(payload.get("handoff_active")),
            "end_session": bool(payload.get("end_session")),
            "company_id": company.get("company_id"),
        },
    )
    return {"ok": True, **payload}


@flask_app.post("/api/chat")
def chat_api():
    data = request.get_json(silent=True) or {}
    mensaje = data.get("message", "")
    if not isinstance(mensaje, str):
        return jsonify({"ok": False, "error": "Formato de mensaje inválido"}), 400
    mensaje_trim = (mensaje or "").strip()
    result = _process_chat_turn(mensaje_trim)
    return jsonify(result)


@flask_app.post("/webhook/n8n/sync-knowledge")
def webhook_n8n_sync_knowledge():
    """
    Webhook para N8N: recibe notificación de cambio en Drive, sincroniza la KB
    de la empresa correspondiente y regenera los temas del menú automáticamente.

    Auth: header X-Webhook-Secret con el valor de N8N_WEBHOOK_SECRET.
    Body JSON: { "folder_id": "...", "company_id": "..." }
      - folder_id: ID de la subcarpeta de la empresa en Drive (recomendado)
      - company_id: ID de la empresa en el sistema (alternativo)
      Se puede enviar uno o ambos. Si solo se envía folder_id, el sistema busca
      qué empresa tiene esa carpeta configurada.
    """
    expected_secret = os.getenv("N8N_WEBHOOK_SECRET", "").strip()
    if not expected_secret:
        return jsonify({"ok": False, "error": "N8N_WEBHOOK_SECRET no configurado en el servidor."}), 503
    incoming_secret = (request.headers.get("X-Webhook-Secret") or "").strip()
    _valid_secrets = {s.strip() for s in expected_secret.split(",") if s.strip()}
    if incoming_secret not in _valid_secrets:
        logging.warning("webhook_n8n_sync_knowledge: secret inválido desde %s", request.remote_addr)
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    data = request.get_json(silent=True) or {}
    folder_id = (data.get("folder_id") or "").strip()
    company_id = _normalize_company_id(data.get("company_id") or "")

    # Resolver empresa por folder_id si no se pasó company_id
    if not company_id and folder_id:
        for company in _list_companies(include_inactive=False):
            if (company.get("drive_folder_id") or "").strip() == folder_id:
                company_id = _normalize_company_id(company.get("company_id"))
                break

    if not company_id:
        return jsonify({"ok": False, "error": "No se pudo determinar la empresa. Pasá company_id o folder_id configurado en la empresa."}), 400

    # Si no hay folder_id, usar el de la empresa
    if not folder_id:
        company_data = _get_company(company_id, include_inactive=False)
        folder_id = ((company_data or {}).get("drive_folder_id") or "").strip()

    if not folder_id:
        return jsonify({"ok": False, "error": f"La empresa '{company_id}' no tiene carpeta Drive configurada."}), 400

    logging.info("N8N sync triggered: company=%s folder=%s", company_id, folder_id)
    count, err = _sync_knowledge_from_drive(company_id, folder_id)
    if err:
        return jsonify({"ok": False, "company_id": company_id, "error": err}), 400

    return jsonify({
        "ok": True,
        "company_id": company_id,
        "folder_id": folder_id,
        "entries_count": count,
        "message": f"Sync OK: {count} preguntas/respuestas actualizadas para '{company_id}'.",
    })


@flask_app.post("/webhook/n8n/procesar-comunicados")
def webhook_n8n_procesar_comunicados():
    """
    Webhook para N8N: procesa y envía comunicados programados pendientes cuya fecha ya llegó.
    Llamar cada 5 minutos desde un Schedule Trigger de N8N.

    Auth: header X-Webhook-Secret con el valor de N8N_WEBHOOK_SECRET.
    """
    expected_secret = os.getenv("N8N_WEBHOOK_SECRET", "").strip()
    if not expected_secret:
        return jsonify({"ok": False, "error": "N8N_WEBHOOK_SECRET no configurado."}), 503
    _valid_secrets_pc = {s.strip() for s in expected_secret.split(",") if s.strip()}
    if (request.headers.get("X-Webhook-Secret") or "").strip() not in _valid_secrets_pc:
        return jsonify({"ok": False, "error": "Unauthorized"}), 401
    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible."}), 503

    now_iso = _utc_now().isoformat()
    try:
        docs = list(
            chatbot.db.collection(COMUNICADOS_PROGRAMADOS_COLLECTION)
            .where("estado", "==", "pendiente")
            .where("scheduled_at", "<=", now_iso)
            .stream()
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500

    try:
        from whatsapp_broadcast import broadcast_messages
        broadcast_available = True
    except ImportError:
        broadcast_available = False

    processed = []
    for doc in docs:
        data = doc.to_dict() or {}
        doc_id = doc.id
        phones = data.get("destinatarios") or []
        mensaje = data.get("mensaje") or ""
        media_urls = data.get("media_urls") or []
        whatsapp_phone = data.get("whatsapp_phone") or os.getenv("TWILIO_WHATSAPP_FROM", "")
        # Aplicar fallback a Meta si el phone guardado no es un ID numérico de Meta
        _sched_pid = str(whatsapp_phone or "").strip()
        _sched_is_meta = _sched_pid.isdigit() and len(_sched_pid) > 10
        if not _sched_is_meta:
            _meta_fb = _meta_phone_number_id()
            if _meta_fb:
                whatsapp_phone = _meta_fb
                _sched_is_meta = True

        if not phones or (not mensaje and not media_urls):
            chatbot.db.collection(COMUNICADOS_PROGRAMADOS_COLLECTION).document(doc_id).update({
                "estado": "error",
                "result": {"error": "Datos incompletos"}
            })
            processed.append({"id": doc_id, "estado": "error"})
            continue

        # Marcar como enviando para evitar doble procesamiento
        chatbot.db.collection(COMUNICADOS_PROGRAMADOS_COLLECTION).document(doc_id).update({"estado": "enviando"})

        if not broadcast_available:
            chatbot.db.collection(COMUNICADOS_PROGRAMADOS_COLLECTION).document(doc_id).update({
                "estado": "error",
                "result": {"error": "Módulo de envío no disponible"}
            })
            processed.append({"id": doc_id, "estado": "error"})
            continue

        try:
            if _sched_is_meta:
                from whatsapp_broadcast import set_send_function as _wb_set_sched
                def _sched_meta_fn(phone, body=None, media_url=None, phone_number_id=None, **kw):
                    try:
                        urls = media_url if isinstance(media_url, list) else ([media_url] if media_url else [])
                        if urls:
                            import requests as _rsched
                            _r = _rsched.get(urls[0], timeout=30)
                            if _r.ok:
                                _mime = _r.headers.get("Content-Type", "image/jpeg").split(";")[0].strip()
                                _fname = urls[0].split("/")[-1].split("?")[0] or "imagen.jpg"
                                _mid = _upload_media_to_meta(_r.content, _mime, _fname, phone_number_id=phone_number_id)
                                if _mid:
                                    _mt = "image" if _mime.startswith("image/") else "document"
                                    _send_meta_whatsapp_media(phone, _mid, _mt, caption=body or None, filename=_fname if _mt == "document" else None, phone_number_id=phone_number_id)
                                    return True
                        if body:
                            return bool(_send_meta_whatsapp(phone, body, phone_number_id=phone_number_id))
                        return False
                    except Exception as _e:
                        logger.warning("scheduler meta send error %s: %s", phone, _e)
                        return False
                _wb_set_sched(_sched_meta_fn)
            result = broadcast_messages(
                phone_list=phones,
                body_text=mensaje or None,
                phone_number_id=whatsapp_phone,
                media_url=media_urls if media_urls else None,
            )
            chatbot.db.collection(COMUNICADOS_PROGRAMADOS_COLLECTION).document(doc_id).update({
                "estado": "enviado",
                "sent_at": _utc_now().isoformat(),
                "result": {
                    "sent": result.get("sent", 0),
                    "failed": result.get("failed", 0),
                    "total": result.get("total", 0),
                }
            })
            processed.append({"id": doc_id, "estado": "enviado", **result})
        except Exception as exc:
            chatbot.db.collection(COMUNICADOS_PROGRAMADOS_COLLECTION).document(doc_id).update({
                "estado": "error",
                "result": {"error": str(exc)}
            })
            processed.append({"id": doc_id, "estado": "error", "error": str(exc)})

    return jsonify({"ok": True, "procesados": len(processed), "detalle": processed})


@flask_app.post("/webhook/n8n/reporte-semanal")
def webhook_n8n_reporte_semanal():
    """
    Webhook para N8N: genera y envía el reporte semanal de handoffs por empresa.
    Envía un email separado a cada empresa que tenga 'Email empresa' configurado.

    Auth: header X-Webhook-Secret con el valor de N8N_WEBHOOK_SECRET.
    Body JSON: { "days": 7 }
    """
    expected_secret = os.getenv("N8N_WEBHOOK_SECRET", "").strip()
    if not expected_secret:
        return jsonify({"ok": False, "error": "N8N_WEBHOOK_SECRET no configurado."}), 503
    _valid_secrets_rs = {s.strip() for s in expected_secret.split(",") if s.strip()}
    if (request.headers.get("X-Webhook-Secret") or "").strip() not in _valid_secrets_rs:
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    if not chatbot.db:
        return jsonify({"ok": False, "error": "Firestore no disponible."}), 503

    data = request.get_json(silent=True) or {}
    days = max(1, min(int(data.get("days") or 7), 90))

    desde = _utc_now() - timedelta(days=days)
    fecha_desde = desde.strftime("%d/%m/%Y")
    fecha_hasta = _utc_now().strftime("%d/%m/%Y")

    # Consultar todos los handoffs del período
    try:
        docs = chatbot.db.collection("rrhh_handoffs").where("created_at", ">=", desde).stream()
        handoffs = [d.to_dict() for d in docs]
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500

    # Agrupar handoffs por empresa
    by_company: dict = {}
    for h in handoffs:
        cid = str(h.get("company_id") or "default").strip()
        by_company.setdefault(cid, []).append(h)

    # Cargar empresas con email configurado
    companies = _list_companies(include_inactive=True)
    company_map = {str(c.get("company_id") or "").strip(): c for c in companies if c.get("company_id")}

    sent = []
    errors = []

    for cid, items in by_company.items():
        company = company_map.get(cid) or {}
        dest_email = str(company.get("company_email") or "").strip()
        if not dest_email:
            continue

        company_name = company.get("company_name") or cid
        total = len(items)
        cerradas = sum(1 for h in items if h.get("estado") == HANDOFF_STATUS_CLOSED)
        pendientes = sum(1 for h in items if h.get("estado") == HANDOFF_STATUS_PENDING)
        en_atencion = sum(1 for h in items if h.get("estado") == HANDOFF_STATUS_ACTIVE)

        por_area: dict = {}
        por_dia: dict = {}
        for h in items:
            area = (h.get("area") or "Sin área").strip()
            por_area[area] = por_area.get(area, 0) + 1
            created = h.get("created_at")
            if created:
                try:
                    dia = created.strftime("%d/%m/%Y") if hasattr(created, "strftime") else str(created)[:10]
                    por_dia[dia] = por_dia.get(dia, 0) + 1
                except Exception:
                    pass

        def fmt_tabla(filas):
            if not filas:
                return "  (sin datos)"
            ancho = max(len(str(f[0])) for f in filas)
            return "\n".join(f"  {str(f[0]).ljust(ancho)}  {f[1]}" for f in filas)

        lineas = [
            f"Reporte semanal — {fecha_desde} al {fecha_hasta}",
            f"Empresa: {company_name}",
            "=" * 50,
            "",
            f"Total de consultas: {total}",
            f"  Cerradas:    {cerradas}",
            f"  Pendientes:  {pendientes}",
            f"  En atención: {en_atencion}",
            "",
            "Por área:",
            fmt_tabla(sorted(por_area.items(), key=lambda x: -x[1])),
        ]
        if por_dia:
            lineas += [
                "",
                "Por día:",
                fmt_tabla(sorted(por_dia.items())),
            ]
        lineas += ["", "Ver detalles: https://debo-chat.web.app/?m=rrhh"]
        body = "\n".join(lineas)
        subject = f"[{company_name}] Reporte semanal de consultas RRHH — {fecha_desde} al {fecha_hasta}"

        ok, err = _send_email(dest_email, subject, body)
        if ok:
            sent.append(cid)
        else:
            errors.append({"company_id": cid, "error": err})
            logging.warning(f"reporte-semanal: error enviando a {dest_email}: {err}")

    return jsonify({
        "ok": True,
        "days": days,
        "total_handoffs": len(handoffs),
        "sent_to": sent,
        "errors": errors,
    })


@flask_app.get("/webhook/twilio/whatsapp")
def webhook_twilio_whatsapp_get():
    """Twilio a veces valida la URL con GET. Respondemos TwiML vacío."""
    from twilio.twiml.messaging_response import MessagingResponse
    resp = MessagingResponse()
    return str(resp), 200, {"Content-Type": "text/xml"}


@flask_app.post("/webhook/twilio/whatsapp")
def webhook_twilio_whatsapp():
    """Recibe mensajes entrantes de Twilio (colaborador escribe por WhatsApp al número de Bacar)."""
    from_phone = (request.form.get("From") or "").strip()
    to_phone = (request.form.get("To") or "").strip()
    body = (request.form.get("Body") or "").strip()
    # Twilio envía fotos/audios con NumMedia y MediaUrl0, MediaUrl1, ...
    media_urls = []
    try:
        num_media = int(request.form.get("NumMedia") or 0)
        for i in range(num_media):
            url = (request.form.get(f"MediaUrl{i}") or "").strip()
            if url:
                media_urls.append(url)
    except (ValueError, TypeError):
        pass
    logger.info(
        "Webhook Twilio WhatsApp: From=%r To=%r Body_len=%s NumMedia=%s",
        from_phone, to_phone, len(body), len(media_urls)
    )
    if not from_phone:
        logger.warning("Webhook Twilio: sin From, devolviendo 400")
        return ("", 400)
    if not body and not media_urls:
        logger.info("Webhook Twilio: sin Body ni medios, ignorando (200)")
        return ("", 200)
    if media_urls:
        if body:
            body = body + "\n📎 Archivos: " + " ".join(media_urls)
        else:
            body = "📎 Envió " + str(len(media_urls)) + " archivo(s)"
    g.whatsapp_from = from_phone
    g.whatsapp_to = to_phone
    g.whatsapp_phone = from_phone
    g.whatsapp_profile_name = (request.form.get("ProfileName") or "").strip()
    g.whatsapp_session = WHATSAPP_SESSIONS.setdefault(from_phone, {})
    # Siempre recargar desde Firestore para evitar inconsistencias entre instancias Cloud Run
    _load_whatsapp_chat_context(from_phone)
    # Si el colaborador ya fue identificado por DNI, cargar sus datos en sesión
    if not g.whatsapp_session.get("wa_empleado_id"):
        identity = _get_whatsapp_identity(from_phone)
        if identity:
            g.whatsapp_session["wa_empleado_id"] = identity.get("empleado_id") or ""
            g.whatsapp_session["wa_convenio"] = identity.get("convenio") or ""
            g.whatsapp_session["wa_nombre"] = identity.get("nombre") or ""
    if not g.whatsapp_session.get("chat_context_company_id"):
        cid, company, line_label = _company_by_whatsapp_phone(to_phone)
        if cid and company:
            g.whatsapp_session["chat_context_company_id"] = cid
            g.whatsapp_session["company_id"] = cid
            g.whatsapp_session["company_name"] = company.get("company_name") or cid
            branches = _get_branches_for_company(company)
            areas = _get_all_areas_for_company(company)
            label_norm = (line_label or "").strip().lower()
            if label_norm and areas and any(str(a).strip().lower() == label_norm for a in areas):
                g.whatsapp_session["chat_context_area"] = line_label.strip()
                g.whatsapp_session["chat_context_step"] = CHAT_CONTEXT_STEP_READY
            elif branches:
                g.whatsapp_session["chat_context_step"] = CHAT_CONTEXT_STEP_BRANCH
            elif areas:
                g.whatsapp_session["chat_context_step"] = CHAT_CONTEXT_STEP_AREA
            else:
                g.whatsapp_session["chat_context_step"] = CHAT_CONTEXT_STEP_READY
        elif not g.whatsapp_session.get("chat_context_step") and not g.whatsapp_session.get("wa_empleado_id"):
            # Sesión nueva sin empresa asignada y sin identidad → pedir DNI
            g.whatsapp_session["chat_context_step"] = CHAT_CONTEXT_STEP_DNI
    if not g.whatsapp_session.get("handoff_conversation_id"):
        open_handoff_id = _find_open_handoff_by_whatsapp_phone(from_phone)
        if open_handoff_id:
            g.whatsapp_session["handoff_conversation_id"] = open_handoff_id
            logger.info("Webhook Twilio: sesión recuperada, handoff=%s", open_handoff_id)
    g.whatsapp_media_urls = media_urls
    try:
        result = _process_chat_turn(body)
    except Exception as e:
        logger.exception("Webhook Twilio: error en _process_chat_turn: %s", e)
        from twilio.twiml.messaging_response import MessagingResponse
        resp = MessagingResponse()
        return str(resp), 200, {"Content-Type": "text/xml"}
    _save_whatsapp_chat_context(from_phone)
    reply = (result.get("reply") or "").strip()
    if reply:
        try:
            from twilio_whatsapp import send_one
            send_one(from_phone, body=reply, phone_number_id=to_phone)
        except Exception as e:
            logger.warning("Webhook Twilio: no se pudo enviar respuesta por WhatsApp: %s", e)
    from twilio.twiml.messaging_response import MessagingResponse
    resp = MessagingResponse()
    return str(resp), 200, {"Content-Type": "text/xml"}


# ══════════════════════════════════════════════════════════════════════════════
#  META WHATSAPP CLOUD API
# ══════════════════════════════════════════════════════════════════════════════

def _meta_access_token():
    return os.getenv("META_ACCESS_TOKEN", "").strip()

def _meta_phone_number_id():
    return os.getenv("META_PHONE_NUMBER_ID", "").strip()


def _send_meta_whatsapp(to_phone, text, phone_number_id=None):
    """Envía un mensaje de texto por WhatsApp vía Meta Cloud API."""
    import requests as _req
    pid = phone_number_id or _meta_phone_number_id()
    token = _meta_access_token()
    if not pid or not token:
        logger.warning("_send_meta_whatsapp: META_PHONE_NUMBER_ID o META_ACCESS_TOKEN no configurado")
        return False
    # Normalizar número: Meta espera solo dígitos sin '+'
    to_norm = re.sub(r"[^\d]", "", to_phone)

    def _try_send(number):
        url = f"https://graph.facebook.com/v18.0/{pid}/messages"
        payload = {
            "messaging_product": "whatsapp",
            "to": number,
            "type": "text",
            "text": {"body": str(text)[:4096]},
        }
        return _req.post(url, json=payload, headers={"Authorization": f"Bearer {token}"}, timeout=15)

    try:
        r = _try_send(to_norm)
        if not r.ok:
            # Error 131030: número no está en la lista permitida (sandbox).
            # Los números argentinos tienen dos formatos posibles:
            #   Moderno (wa_id): 5493515416836  = 54 + 9 + área(351) + local(5416836)
            #   Viejo (lista):   54351155416836 = 54 + área(351) + 15 + local(5416836)
            # Reintentamos con el formato alternativo.
            try:
                err_code = (r.json() or {}).get("error", {}).get("code")
            except Exception:
                err_code = None
            if err_code == 131030 and to_norm.startswith("549") and len(to_norm) == 13:
                # Formato moderno → viejo: 549 + área(3) + local(7) → 54 + área(3) + 15 + local(7)
                rest = to_norm[3:]  # quitar "549" → ej. "3515416836"
                alt = "54" + rest[:3] + "15" + rest[3:]
                r2 = _try_send(alt)
                if r2.ok:
                    return True
                logger.warning("_send_meta_whatsapp: error con formato alt %s: %s %s", alt, r2.status_code, r2.text[:300])
            elif err_code == 131030 and len(to_norm) == 14 and to_norm.startswith("54") and to_norm[5:7] == "15":
                # Formato viejo → moderno: 54 + área(3) + 15 + local(7) → 549 + área(3) + local(7)
                area = to_norm[2:5]
                local = to_norm[7:]
                alt = "549" + area + local
                r2 = _try_send(alt)
                if r2.ok:
                    return True
                logger.warning("_send_meta_whatsapp: error con formato alt %s: %s %s", alt, r2.status_code, r2.text[:300])
            logger.warning("_send_meta_whatsapp: error %s: %s", r.status_code, r.text[:500])
        return r.ok
    except Exception as e:
        logger.warning("_send_meta_whatsapp: excepción: %s", e)
        return False


def _send_meta_interactive(to_phone, body_text, options, phone_number_id=None):
    """Envía un mensaje interactivo de WhatsApp vía Meta Cloud API.
    - ≤3 opciones → botones inline (sin "Ver opciones")
    - 4-10 opciones → lista desplegable con botón "Ver opciones"
    - >10 o si falla → fallback texto plano
    Reintenta con formato argentino alternativo ante error 131030.
    """
    import requests as _req
    pid = phone_number_id or _meta_phone_number_id()
    token = _meta_access_token()
    if not pid or not token:
        return False
    to_norm = re.sub(r"[^\d]", "", to_phone)
    valid_opts = [o for o in (options or []) if str(o.get("label") or "").strip()][:10]
    if not valid_opts:
        return _send_meta_whatsapp(to_phone, body_text, phone_number_id=pid)

    body_trunc = str(body_text or "")[:1024].strip()

    if len(valid_opts) <= 3:
        buttons = []
        for opt in valid_opts:
            title = str(opt["label"]).strip()[:20]
            btn_id = str(opt.get("value") or opt["label"]).strip()[:256]
            buttons.append({"type": "reply", "reply": {"id": btn_id, "title": title}})
        interactive_payload = {
            "type": "button",
            "body": {"text": body_trunc},
            "action": {"buttons": buttons},
        }
    else:
        rows = []
        for opt in valid_opts:
            title = str(opt["label"]).strip()[:24]
            row_id = str(opt.get("value") or opt["label"]).strip()[:200]
            rows.append({"id": row_id, "title": title})
        interactive_payload = {
            "type": "list",
            "body": {"text": body_trunc},
            "action": {
                "button": "Ver opciones",
                "sections": [{"title": "Opciones", "rows": rows}],
            },
        }

    def _try_send(number):
        payload = {
            "messaging_product": "whatsapp",
            "to": number,
            "type": "interactive",
            "interactive": interactive_payload,
        }
        url = f"https://graph.facebook.com/v18.0/{pid}/messages"
        return _req.post(url, json=payload, headers={"Authorization": f"Bearer {token}"}, timeout=15)

    try:
        r = _try_send(to_norm)
        if not r.ok:
            try:
                err_code = (r.json() or {}).get("error", {}).get("code")
            except Exception:
                err_code = None
            # Retry con formato argentino alternativo
            if err_code == 131030 and to_norm.startswith("549") and len(to_norm) == 13:
                rest = to_norm[3:]
                alt = "54" + rest[:3] + "15" + rest[3:]
                r2 = _try_send(alt)
                if r2.ok:
                    return True
            elif err_code == 131030 and len(to_norm) == 14 and to_norm.startswith("54") and to_norm[5:7] == "15":
                area = to_norm[2:5]
                local = to_norm[7:]
                alt = "549" + area + local
                r2 = _try_send(alt)
                if r2.ok:
                    return True
            # Fallback: texto plano
            logger.warning("_send_meta_interactive: fallo %s, fallback texto", r.status_code)
            return _send_meta_whatsapp(to_phone, body_text, phone_number_id=pid)
        return r.ok
    except Exception as e:
        logger.warning("_send_meta_interactive: %s", e)
        return _send_meta_whatsapp(to_phone, body_text, phone_number_id=pid)


def _upload_media_to_meta(file_bytes, mime_type, filename, phone_number_id=None):
    """Sube un archivo a Meta y devuelve el media_id."""
    import requests as _req
    pid = phone_number_id or _meta_phone_number_id()
    token = _meta_access_token()
    if not pid or not token:
        return None
    url = f"https://graph.facebook.com/v18.0/{pid}/media"
    try:
        r = _req.post(
            url,
            files={"file": (filename, file_bytes, mime_type)},
            data={"messaging_product": "whatsapp"},
            headers={"Authorization": f"Bearer {token}"},
            timeout=60,
        )
        if r.ok:
            return (r.json() or {}).get("id")
        logger.warning("_upload_media_to_meta: %s %s", r.status_code, r.text[:200])
        return None
    except Exception as e:
        logger.warning("_upload_media_to_meta: %s", e)
        return None


def _send_meta_whatsapp_media(to_phone, media_id, media_type, caption=None, filename=None, phone_number_id=None):
    """Envía un archivo (imagen, documento, etc.) vía Meta API usando un media_id ya subido.
    Reintenta con formato de número argentino alternativo si recibe error 131030.
    """
    import requests as _req
    pid = phone_number_id or _meta_phone_number_id()
    token = _meta_access_token()
    if not pid or not token or not media_id:
        return False
    to_norm = re.sub(r"[^\d]", "", to_phone)

    def _try_send(number):
        media_obj = {"id": media_id}
        if caption:
            media_obj["caption"] = str(caption)[:1024]
        if filename and media_type == "document":
            media_obj["filename"] = filename
        payload = {
            "messaging_product": "whatsapp",
            "to": number,
            "type": media_type,
            media_type: media_obj,
        }
        url = f"https://graph.facebook.com/v18.0/{pid}/messages"
        return _req.post(url, json=payload, headers={"Authorization": f"Bearer {token}"}, timeout=30)

    try:
        r = _try_send(to_norm)
        if not r.ok:
            try:
                err_code = (r.json() or {}).get("error", {}).get("code")
            except Exception:
                err_code = None
            if err_code == 131030 and to_norm.startswith("549") and len(to_norm) == 13:
                rest = to_norm[3:]
                alt = "54" + rest[:3] + "15" + rest[3:]
                r2 = _try_send(alt)
                if r2.ok:
                    return True
                logger.warning("_send_meta_whatsapp_media: error alt %s: %s %s", alt, r2.status_code, r2.text[:200])
            elif err_code == 131030 and len(to_norm) == 14 and to_norm.startswith("54") and to_norm[5:7] == "15":
                area = to_norm[2:5]
                local = to_norm[7:]
                alt = "549" + area + local
                r2 = _try_send(alt)
                if r2.ok:
                    return True
                logger.warning("_send_meta_whatsapp_media: error alt %s: %s %s", alt, r2.status_code, r2.text[:200])
            logger.warning("_send_meta_whatsapp_media: %s %s", r.status_code, r.text[:200])
        return r.ok
    except Exception as e:
        logger.warning("_send_meta_whatsapp_media: %s", e)
        return False


def _download_meta_media(media_id):
    """Descarga un archivo de Meta por su media_id. Devuelve (bytes, mime_type) o (None, None)."""
    import requests as _req
    token = _meta_access_token()
    if not token or not media_id:
        logger.warning("_download_meta_media: token o media_id vacío")
        return None, None
    try:
        # 1) Obtener URL de descarga
        r = _req.get(
            f"https://graph.facebook.com/v18.0/{media_id}",
            headers={"Authorization": f"Bearer {token}"},
            timeout=15,
        )
        if not r.ok:
            logger.warning("_download_meta_media: step1 error %s — %s", r.status_code, r.text[:300])
            return None, None
        data = r.json() or {}
        dl_url = data.get("url")
        mime = data.get("mime_type", "application/octet-stream")
        if not dl_url:
            logger.warning("_download_meta_media: sin url en respuesta: %s", data)
            return None, None
        # 2) Descargar el archivo
        r2 = _req.get(dl_url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
        if not r2.ok:
            logger.warning("_download_meta_media: step2 error %s", r2.status_code)
            return None, None
        return r2.content, mime
    except Exception as e:
        logger.warning("_download_meta_media: %s", e)
        return None, None


@flask_app.get("/webhook/meta/whatsapp")
def webhook_meta_whatsapp_get():
    """Verificación del webhook por Meta (GET)."""
    verify_token = os.getenv("META_VERIFY_TOKEN", "chatbot_rrhh_verify").strip()
    mode = request.args.get("hub.mode", "")
    token = request.args.get("hub.verify_token", "")
    challenge = request.args.get("hub.challenge", "")
    if mode == "subscribe" and token == verify_token:
        return challenge, 200
    return "Forbidden", 403


@flask_app.post("/webhook/meta/whatsapp")
def webhook_meta_whatsapp():
    """Recibe mensajes entrantes de Meta WhatsApp Cloud API."""
    data = request.get_json(silent=True) or {}

    for entry in (data.get("entry") or []):
        for change in (entry.get("changes") or []):
            value = change.get("value") or {}
            phone_number_id = (value.get("metadata") or {}).get("phone_number_id") or _meta_phone_number_id()

            # Ignorar eventos de estado (delivered, read, etc.)
            for status in (value.get("statuses") or []):
                pass  # podríamos loguear, pero no procesamos

            for msg in (value.get("messages") or []):
                from_phone = (msg.get("from") or "").strip()
                if not from_phone:
                    continue

                msg_type = (msg.get("type") or "").strip()
                body = ""
                media_info = None  # (media_id, mime_type, filename)

                if msg_type == "text":
                    body = ((msg.get("text") or {}).get("body") or "").strip()

                elif msg_type == "location":
                    loc = msg.get("location") or {}
                    lat = loc.get("latitude")
                    lng = loc.get("longitude")
                    if lat is not None and lng is not None:
                        body = f"__location__ {lat},{lng}"

                elif msg_type in ("image", "video", "audio", "sticker"):
                    media_obj = msg.get(msg_type) or {}
                    mid = media_obj.get("id")
                    mime = media_obj.get("mime_type", "")
                    caption = (media_obj.get("caption") or "").strip()
                    if mid:
                        body = caption or f"📎 Envió {msg_type}"
                        media_info = (mid, mime, msg_type)

                elif msg_type == "document":
                    doc = msg.get("document") or {}
                    mid = doc.get("id")
                    mime = doc.get("mime_type", "application/octet-stream")
                    fname = (doc.get("filename") or "archivo").strip()
                    caption = (doc.get("caption") or "").strip()
                    if mid:
                        body = caption or f"📎 {fname}"
                        media_info = (mid, mime, fname)

                elif msg_type == "interactive":
                    # Respuestas de botones o listas interactivas
                    # Usamos "id" (= value original) para matching correcto; fallback a "title"
                    interactive = msg.get("interactive") or {}
                    itype = interactive.get("type")
                    if itype == "button_reply":
                        btn = interactive.get("button_reply") or {}
                        body = btn.get("id") or btn.get("title") or ""
                    elif itype == "list_reply":
                        row = interactive.get("list_reply") or {}
                        body = row.get("id") or row.get("title") or ""

                if not body and not media_info:
                    continue

                profile_name = ((value.get("contacts") or [{}])[0].get("profile") or {}).get("name") or ""

                logger.info("Webhook Meta: from=%s type=%s body_len=%s pid=%s", from_phone, msg_type, len(body), phone_number_id)

                # ── Inicializar sesión (igual que Twilio) ─────────────────────
                g.whatsapp_from = from_phone
                g.whatsapp_to = phone_number_id
                g.whatsapp_phone = from_phone
                g.whatsapp_phone_number_id = phone_number_id
                g.whatsapp_profile_name = profile_name
                g.whatsapp_session = WHATSAPP_SESSIONS.setdefault(from_phone, {})
                _load_whatsapp_chat_context(from_phone)

                # Cargar identidad persistida (DNI ya vinculado)
                if not g.whatsapp_session.get("wa_empleado_id"):
                    identity = _get_whatsapp_identity(from_phone)
                    if identity:
                        g.whatsapp_session["wa_empleado_id"] = identity.get("empleado_id") or ""
                        g.whatsapp_session["wa_convenio"] = identity.get("convenio") or ""
                        g.whatsapp_session["wa_nombre"] = identity.get("nombre") or ""

                if not g.whatsapp_session.get("chat_context_company_id"):
                    cid, company, line_label = _company_by_whatsapp_phone(phone_number_id)
                    if cid and company:
                        g.whatsapp_session["chat_context_company_id"] = cid
                        g.whatsapp_session["company_id"] = cid
                        g.whatsapp_session["company_name"] = company.get("company_name") or cid
                        # Solo setear el step si no hay uno ya cargado desde Firestore.
                        # Si hay step (ej. "ready"), no pisarlo aunque no tengamos company_id.
                        if not g.whatsapp_session.get("chat_context_step"):
                            branches = _get_branches_for_company(company)
                            areas = _get_all_areas_for_company(company)
                            label_norm = (line_label or "").strip().lower()
                            if label_norm and areas and any(str(a).strip().lower() == label_norm for a in areas):
                                g.whatsapp_session["chat_context_area"] = line_label.strip()
                                g.whatsapp_session["chat_context_step"] = CHAT_CONTEXT_STEP_READY
                            elif branches:
                                g.whatsapp_session["chat_context_step"] = CHAT_CONTEXT_STEP_BRANCH
                            elif areas:
                                g.whatsapp_session["chat_context_step"] = CHAT_CONTEXT_STEP_AREA
                            else:
                                g.whatsapp_session["chat_context_step"] = CHAT_CONTEXT_STEP_READY
                    elif not g.whatsapp_session.get("chat_context_step") and not g.whatsapp_session.get("wa_empleado_id"):
                        g.whatsapp_session["chat_context_step"] = CHAT_CONTEXT_STEP_DNI

                if not g.whatsapp_session.get("handoff_conversation_id"):
                    open_handoff_id = _find_open_handoff_by_whatsapp_phone(from_phone)
                    if open_handoff_id:
                        g.whatsapp_session["handoff_conversation_id"] = open_handoff_id

                # Guardar phone_number_id en sesión para que el panel lo use al responder
                g.whatsapp_session["meta_phone_number_id"] = phone_number_id

                # Adjuntos: poner URL de descarga en g para que el panel los registre
                if media_info:
                    from urllib.parse import quote as _uq
                    g.whatsapp_media_urls = [f"meta_media://{media_info[0]}?mime={_uq(media_info[1] or '')}&fname={_uq(str(media_info[2] or ''))}"]
                else:
                    g.whatsapp_media_urls = []

                try:
                    result = _process_chat_turn(body)
                except Exception as e:
                    logger.exception("Webhook Meta: error en _process_chat_turn: %s", e)
                    result = {"reply": ""}

                _save_whatsapp_chat_context(from_phone)

                reply = (result.get("reply") or "").strip()
                quick_actions = result.get("quick_actions") or []
                if reply:
                    if quick_actions:
                        # Extraer header (texto antes del menú numerado) para el body interactivo
                        header_match = re.search(r'\n+\d+[\.\)]\s', reply)
                        header = reply[:header_match.start()].strip() if header_match else reply
                        _send_meta_interactive(from_phone, header, quick_actions, phone_number_id=phone_number_id)
                    else:
                        _send_meta_whatsapp(from_phone, reply, phone_number_id=phone_number_id)

    return jsonify({"status": "ok"}), 200


@flask_app.get("/login")
def login_page():
    if not _auth_enabled():
        return redirect(url_for("rrhh_page"))
    if _current_rrhh_user() is not None:
        return redirect(_safe_next_path(request.args.get("next"), fallback="/rrhh"))
    companies = _list_companies(include_inactive=False)
    selected_company = _normalize_company_id(
        request.args.get("empresa") or session.get("company_id") or _default_company_id()
    )
    if not any(item.get("company_id") == selected_company for item in companies):
        selected_company = companies[0]["company_id"] if companies else _default_company_id()
    return render_template(
        "login.html",
        error="",
        next_path=_safe_next_path(request.args.get("next"), fallback="/rrhh"),
        companies=companies,
        selected_company=selected_company,
    )


@flask_app.post("/login")
def login_submit():
    if not _auth_enabled():
        return redirect(url_for("rrhh_page"))

    data = request.get_json(silent=True) if request.is_json else request.form
    username = str((data or {}).get("username") or "").strip()
    password = str((data or {}).get("password") or "")
    company_id = _normalize_company_id((data or {}).get("company_id") or request.args.get("empresa"))
    next_path = _safe_next_path((data or {}).get("next") or request.args.get("next"), "/rrhh")

    ok, user_payload, error = auth_rrhh.authenticate(username, password)
    if not ok:
        if request.is_json:
            return jsonify({"ok": False, "error": error}), 401
        companies = _list_companies(include_inactive=False)
        selected_company = company_id or (companies[0]["company_id"] if companies else _default_company_id())
        return (
            render_template(
                "login.html",
                error=error,
                next_path=next_path,
                companies=companies,
                selected_company=selected_company,
            ),
            401,
        )

    user_companies = _companies_for_user(user_payload)
    if not user_companies:
        message = "Tu usuario no tiene empresas asignadas o activas."
        if request.is_json:
            return jsonify({"ok": False, "error": message}), 403
        return (
            render_template(
                "login.html",
                error=message,
                next_path=next_path,
                companies=_list_companies(include_inactive=False),
                selected_company=company_id or _default_company_id(),
            ),
            403,
        )
    if company_id and not any(item.get("company_id") == company_id for item in user_companies):
        message = "No tenés acceso a la empresa seleccionada."
        if request.is_json:
            return jsonify({"ok": False, "error": message}), 403
        return (
            render_template(
                "login.html",
                error=message,
                next_path=next_path,
                companies=user_companies,
                selected_company=company_id,
            ),
            403,
        )

    chosen_company = company_id or (
        user_companies[0]["company_id"] if user_companies else _default_company_id()
    )
    if not _user_can_access_company(user_payload, chosen_company):
        message = "No tenés acceso a la empresa seleccionada."
        if request.is_json:
            return jsonify({"ok": False, "error": message}), 403
        return (
            render_template(
                "login.html",
                error=message,
                next_path=next_path,
                companies=user_companies,
                selected_company=chosen_company,
            ),
            403,
        )

    company = _set_company_session(chosen_company)
    _set_rrhh_user(user_payload)
    if request.is_json:
        return jsonify(
            {
                "ok": True,
                "redirect_to": next_path,
                "user": user_payload,
                "company": {
                    "company_id": company.get("company_id"),
                    "company_name": company.get("company_name"),
                },
            }
        )
    import json as _json
    _module_map = {
        "/rrhh": "rrhh", "/historial": "historial", "/configuracion": "configuracion",
        "/estadisticas": "estadisticas", "/comunicados": "comunicados",
        "/legajos": "legajos", "/preferencias": "preferencias",
    }
    _module = _module_map.get(next_path)
    _top_url = f"/?m={_module}" if _module else next_path
    return f'<script>window.top.location.href = {_json.dumps(_top_url)};</script>', 200, {"Content-Type": "text/html"}


@flask_app.route("/logout", methods=["GET", "POST"])
def logout_page():
    session.clear()
    session_cookie_name = flask_app.config.get("SESSION_COOKIE_NAME", "session")
    target_url = url_for("login_page") if _auth_enabled() else url_for("rrhh_page")
    if request.method == "POST" and not request.is_json:
        resp = make_response(
            f'<script>window.top.location.href = {__import__("json").dumps(target_url)};</script>',
            200,
        )
        resp.content_type = "text/html"
        resp.delete_cookie(session_cookie_name)
        return resp
    response = redirect(target_url)
    response.delete_cookie(session_cookie_name)
    return response


@flask_app.get("/recuperar-clave")
def password_recovery_page():
    if not _auth_enabled():
        return redirect(url_for("login_page"))
    if _current_rrhh_user() is not None:
        return redirect(_safe_next_path(request.args.get("next"), fallback="/rrhh"))
    return render_template(
        "recover_password.html",
        error="",
        success=False,
        message="",
        reset_link="",
    )


@flask_app.post("/recuperar-clave")
def password_recovery_submit():
    if not _auth_enabled():
        return redirect(url_for("login_page"))

    data = request.get_json(silent=True) if request.is_json else request.form
    username = str((data or {}).get("username") or "").strip()
    email = str((data or {}).get("email") or "").strip()
    generic_success_message = (
        "Si los datos coinciden, enviamos un enlace de restablecimiento al email indicado."
    )

    if not username or not email:
        if request.is_json:
            return jsonify({"ok": False, "error": "Ingresá usuario y email."}), 400
        return (
            render_template(
                "recover_password.html",
                error="Ingresá usuario y email.",
                success=False,
                message="",
                reset_link="",
            ),
            400,
        )

    ok, payload, _error = auth_rrhh.create_password_reset_token_for_identity(
        username=username,
        email=email,
        ttl_minutes=60,
        requested_by=f"self:{username}",
    )
    if not ok:
        if request.is_json:
            return jsonify({"ok": True, "message": generic_success_message})
        return render_template(
            "recover_password.html",
            error="",
            success=True,
            message=generic_success_message,
            reset_link="",
        )

    reset_url = url_for("password_reset_page", token=payload.get("token"), _external=True)
    try:
        mail_ok, mail_error = _send_password_reset_email(
            to_email=payload.get("email"),
            display_name=payload.get("display_name") or payload.get("username"),
            reset_url=reset_url,
            expires_at_iso=payload.get("expires_at"),
        )
    except Exception as exc:
        mail_ok, mail_error = False, str(exc)
    smtp_not_configured = mail_error and "SMTP no configurado" in (mail_error or "")
    if request.is_json:
        if mail_ok:
            return jsonify({"ok": True, "mail_sent": True, "message": generic_success_message})
        if smtp_not_configured:
            return jsonify({
                "ok": True,
                "mail_sent": False,
                "message": "El envío por email no está configurado. Usá este enlace para restablecer tu contraseña.",
                "reset_link": reset_url,
            })
        return jsonify({"ok": False, "error": mail_error or "No se pudo enviar el email."}), 502

    if mail_ok:
        return render_template(
            "recover_password.html",
            error="",
            success=True,
            message=generic_success_message,
            reset_link="",
        )
    if smtp_not_configured:
        return render_template(
            "recover_password.html",
            error="",
            success=True,
            message="El envío por email no está configurado. Copiá el enlace de abajo para restablecer tu contraseña (válido por 1 hora).",
            reset_link=reset_url,
        )
    return (
        render_template(
            "recover_password.html",
            error=mail_error or "No se pudo enviar el email de restablecimiento.",
            success=False,
            message="",
            reset_link="",
        ),
        502,
    )


@flask_app.get("/restablecer-clave/<token>")
def password_reset_page(token):
    return render_template(
        "reset_password.html",
        token=str(token or "").strip(),
        error="",
        success=False,
        message="",
    )


@flask_app.post("/restablecer-clave/<token>")
def password_reset_submit(token):
    data = request.get_json(silent=True) if request.is_json else request.form
    password = str((data or {}).get("password") or "")
    confirm = str((data or {}).get("confirm_password") or "")
    if not password:
        return (
            render_template(
                "reset_password.html",
                token=str(token or "").strip(),
                error="Ingresá una contraseña nueva.",
                success=False,
                message="",
            ),
            400,
        )
    if password != confirm:
        return (
            render_template(
                "reset_password.html",
                token=str(token or "").strip(),
                error="Las contraseñas no coinciden.",
                success=False,
                message="",
            ),
            400,
        )

    ok, user, error = auth_rrhh.reset_password_with_token(str(token or "").strip(), password)
    if not ok:
        return (
            render_template(
                "reset_password.html",
                token=str(token or "").strip(),
                error=error or "No se pudo restablecer la contraseña.",
                success=False,
                message="",
            ),
            400,
        )

    return render_template(
        "reset_password.html",
        token="",
        error="",
        success=True,
        message=f"Contraseña actualizada para {user.get('username')}. Ya podés iniciar sesión.",
    )


@flask_app.get("/api/historial")
@rrhh_permission_required(auth_rrhh.PERM_HISTORY_VIEW, message="Sin permiso para ver historial.")
def historial_api():
    try:
        limit = int(request.args.get("limit", "200"))
    except Exception:
        limit = 200
    limit = max(1, min(limit, 1000))

    remitente = str(request.args.get("remitente", "")).strip().lower()
    canal = str(request.args.get("canal", "")).strip().lower()
    conversation_id = str(request.args.get("conversation_id", "")).strip()
    q = str(request.args.get("q", "")).strip().lower()
    company_id_raw = request.args.get("company_id", "").strip() or None
    company_id = _normalize_company_id(company_id_raw) if company_id_raw else None
    branches_param = request.args.get("branches")
    branches = [b.strip() for b in (branches_param or "").split(",") if b.strip()] if branches_param else None
    areas_param = request.args.get("areas")
    areas = [a.strip() for a in (areas_param or "").split(",") if a.strip()] if areas_param else None

    raw_items = _list_chat_history(limit=1500)
    handoff_ids = set()
    handoff_names = {}  # conv_id → colaborador_nombre
    if company_id:
        handoffs = _list_handoffs(
            include_closed=True,
            limit=10000,
            company_id=company_id,
            branches=branches if branches else None,
            areas=areas if areas else None,
        )
        for h in handoffs or []:
            hid = str(h.get("id") or h.get("conversation_id") or "").strip()
            if hid:
                handoff_ids.add(hid)
                nombre = str(h.get("colaborador_nombre") or "").strip()
                if nombre:
                    handoff_names[hid] = nombre
    else:
        # Sin filtro de empresa: cargar todos los handoffs para obtener nombres
        all_handoffs = _all_handoff_records_for_stats()
        for h in all_handoffs:
            hid = str(h.get("id") or h.get("conversation_id") or "").strip()
            nombre = str(h.get("colaborador_nombre") or "").strip()
            if hid and nombre:
                handoff_names[hid] = nombre

    items = []
    for item in raw_items:
        if company_id:
            meta = item.get("metadata") or {}
            item_company = _normalize_company_id(meta.get("company_id"))
            conv_id = str(item.get("conversation_id") or "").strip()
            if item_company != company_id and conv_id not in handoff_ids:
                continue
        serialized = _serialize_history_item(item)
        conv_id_s = serialized["conversation_id"]
        if conv_id_s in handoff_names:
            serialized["colaborador_nombre"] = handoff_names[conv_id_s]
        if remitente and serialized["remitente"].lower() != remitente:
            continue
        if canal and serialized["canal"].lower() != canal:
            continue
        if conversation_id:
            conv_match = serialized["conversation_id"] == conversation_id
            nombre_match = conversation_id.lower() in (serialized.get("colaborador_nombre") or "").lower()
            if not conv_match and not nombre_match:
                continue
        if q and q not in serialized["texto"].lower():
            continue
        items.append(serialized)

    filter_applied = {
        "company_id": company_id,
        "branches": branches or [],
        "areas": areas or [],
    }
    return jsonify(
        {
            "ok": True,
            "total": len(items),
            "limit": limit,
            "items": items[:limit],
            "filter_applied": filter_applied,
        }
    )


@flask_app.get("/api/chat/poll")
def chat_poll_api():
    _auto_close_expired_handoffs(company=_current_company())
    handoff_id = _get_handoff_session_id()
    if not handoff_id:
        return jsonify({"ok": True, "handoff_active": False, "messages": []})

    conv = _fetch_handoff(handoff_id)
    if not conv:
        _clear_handoff_session()
        return jsonify({"ok": True, "handoff_active": False, "messages": []})

    estado = str(conv.get("estado") or "").strip().lower()
    if estado == HANDOFF_STATUS_CLOSED:
        _clear_handoff_session()

    nuevos = _collect_new_messages_for_collaborator(handoff_id)
    return jsonify(
        {
            "ok": True,
            "handoff_active": estado in {HANDOFF_STATUS_PENDING, HANDOFF_STATUS_ACTIVE},
            "messages": _serialize_messages(nuevos),
            "quick_actions": construir_acciones_handoff()
            if estado in {HANDOFF_STATUS_PENDING, HANDOFF_STATUS_ACTIVE}
            else [],
        }
    )


@flask_app.get("/api/configuracion/general")
@rrhh_auth_required
def configuracion_general_api():
    if not _can_manage_configuration() and not _can_manage_preferences():
        return _forbidden_json_error("Sin permiso para ver configuración.")
    settings = _read_general_settings()
    return jsonify({"ok": True, "settings": settings, "selected_company_id": _current_company().get("company_id")})


@flask_app.post("/api/configuracion/general")
@rrhh_auth_required
def configuracion_general_update_api():
    if not _can_manage_preferences() and not _can_manage_general_config():
        return _forbidden_json_error("Sin permiso para editar configuración general.")
    data = request.get_json(silent=True) or {}
    ok, settings, error = _write_general_settings(data)
    if not ok:
        return jsonify({"ok": False, "error": error}), 400
    _apply_company_branding(settings)
    return jsonify({"ok": True, "settings": settings})


@flask_app.get("/api/configuracion/empresas")
@rrhh_auth_required
def configuracion_empresas_api():
    if not _can_config_empresas() and not _can_config_sucursales() and not _can_config_areas() and not _can_manage_configuration():
        return _forbidden_json_error("Sin permiso para ver empresas.")
    companies = [_company_for_api(c) for c in _list_companies_for_current_rrhh_user(include_inactive=True)]
    return jsonify(
        {
            "ok": True,
            "companies": companies,
            "selected_company_id": _current_company().get("company_id"),
        }
    )


@flask_app.post("/api/configuracion/empresas")
@rrhh_auth_required
def configuracion_crear_empresa_api():
    if not _is_admin():
        return _forbidden_json_error("Solo el admin puede crear empresas.")
    data = request.get_json(silent=True) or {}
    ok, company, error = _upsert_company(data)
    if not ok:
        return jsonify({"ok": False, "error": error}), 400
    _cache_del("companies_active", "companies_all")
    return jsonify({"ok": True, "company": _company_for_api(company)})


@flask_app.post("/api/configuracion/empresas/<company_id>")
@rrhh_auth_required
def configuracion_editar_empresa_api(company_id):
    if not _can_config_empresas() and not _can_config_sucursales() and not _can_config_areas():
        return _forbidden_json_error("Sin permiso para editar empresas.")
    current = _get_company(company_id, include_inactive=True)
    if not current:
        return jsonify({"ok": False, "error": "Empresa no encontrada."}), 404
    data = request.get_json(silent=True) or {}
    payload = {
        "company_id": current.get("company_id"),
        "company_name": data.get("company_name", current.get("company_name")),
        "hr_team_name": data.get("hr_team_name", current.get("hr_team_name")),
        "hr_contact": data.get("hr_contact", current.get("hr_contact")),
        "company_email": data.get("company_email", current.get("company_email")),
        "company_address": data.get("company_address", current.get("company_address")),
        "company_phone": data.get("company_phone", current.get("company_phone")),
        "company_website": data.get("company_website", current.get("company_website")),
        "handoff_auto_close_enabled": data.get(
            "handoff_auto_close_enabled",
            current.get("handoff_auto_close_enabled", False),
        ),
        "handoff_auto_close_minutes": data.get(
            "handoff_auto_close_minutes",
            current.get("handoff_auto_close_minutes", AUTO_CLOSE_DEFAULT_MINUTES),
        ),
        "branches": data.get("branches", current.get("branches") or []),
        "areas": _normalize_areas(data.get("areas", current.get("areas") or [])),
        "areas_by_branch": _normalize_areas_by_branch(
            data.get("areas_by_branch") if "areas_by_branch" in data else (current.get("areas_by_branch") or {})
        ),
        "active": bool(data.get("active", current.get("active", True))),
        "permitir_hablar_con_humano": _normalize_bool_flag(
            data.get("permitir_hablar_con_humano", current.get("permitir_hablar_con_humano", True)),
            default=True,
        ),
        "temas_habilitados": current.get("temas_habilitados") or [],
        "whatsapp_numbers": _normalize_whatsapp_numbers(
            data.get("whatsapp_numbers") if "whatsapp_numbers" in data else (current.get("whatsapp_numbers") or [])
        ),
        "drive_folder_id": str(data.get("drive_folder_id", current.get("drive_folder_id") or "") or "").strip()[:128] or None,
        "handoff_notify_email": str(data.get("handoff_notify_email", current.get("handoff_notify_email") or "") or "").strip()[:200] or None,
    }
    th = data.get("temas_habilitados")
    if isinstance(th, list):
        payload["temas_habilitados"] = [str(t).strip().lower() for t in th if str(t).strip()]
    ok, company, error = _upsert_company(payload)
    if not ok:
        return jsonify({"ok": False, "error": error}), 400
    _cache_del("companies_active", "companies_all")
    if _normalize_company_id(session.get("company_id")) == company.get("company_id"):
        _set_company_session(company.get("company_id"))
    return jsonify({"ok": True, "company": _company_for_api(company)})


@flask_app.get("/api/configuracion/smtp")
@rrhh_auth_required
def configuracion_smtp_get():
    if not _can_config_smtp():
        return _forbidden_json_error("Sin permiso.")
    cfg = _smtp_settings()
    return jsonify({
        "ok": True,
        "host": cfg["host"],
        "port": cfg["port"],
        "username": cfg["username"],
        "from_email": cfg["from_email"],
        "use_tls": cfg["use_tls"],
        "has_password": bool(cfg["password"]),
        "source": "env" if (cfg["host"] and cfg["host"] not in {"smtp.ejemplo.com", "smtp.example.com"} and os.getenv("SMTP_HOST", "").strip() == cfg["host"]) else "firestore",
    })


@flask_app.post("/api/configuracion/smtp")
@rrhh_auth_required
def configuracion_smtp_save():
    if not _can_config_smtp():
        return _forbidden_json_error("Sin permiso.")
    data = request.get_json(silent=True) or {}
    host = str(data.get("host") or "").strip()[:200]
    try:
        port = int(str(data.get("port") or 587))
    except Exception:
        port = 587
    username = str(data.get("username") or "").strip()[:200]
    password = str(data.get("password") or "").strip()[:500]
    from_email = str(data.get("from_email") or "").strip()[:200]
    use_tls = bool(data.get("use_tls", True))
    if not host:
        return jsonify({"ok": False, "error": "Ingresá el host SMTP."}), 400
    doc = {
        "host": host,
        "port": port,
        "username": username,
        "from_email": from_email,
        "use_tls": use_tls,
        "updated_at": _utc_now().isoformat(),
    }
    if password:
        doc["password"] = password
    elif not data.get("clear_password"):
        # Mantener contraseña existente si no se envió una nueva
        existing = _smtp_settings_from_firestore()
        if existing.get("password"):
            doc["password"] = existing["password"]
    try:
        if chatbot.db:
            chatbot.db.collection("rrhh_config").document("smtp").set(doc, merge=False)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"No se pudo guardar: {exc}"}), 500
    return jsonify({"ok": True, "message": "Configuración SMTP guardada."})


@flask_app.post("/api/configuracion/smtp/test")
@rrhh_auth_required
def configuracion_smtp_test():
    if not _can_config_smtp():
        return _forbidden_json_error("Sin permiso.")
    data = request.get_json(silent=True) or {}
    to_email = str(data.get("to") or "").strip()
    if not to_email:
        return jsonify({"ok": False, "error": "Ingresá un email de prueba."}), 400
    ok, err = _send_email(to_email, "Test email - ChatBot RRHH", "Si llegó este mensaje, el SMTP está correctamente configurado.")
    if ok:
        return jsonify({"ok": True, "message": f"Email de prueba enviado a {to_email}."})
    return jsonify({"ok": False, "error": err})


@flask_app.delete("/api/configuracion/empresas/<company_id>")
@rrhh_auth_required
def configuracion_eliminar_empresa_api(company_id):
    if not _is_admin():
        return _forbidden_json_error("Solo el admin puede eliminar empresas.")
    ok, error = _delete_company(company_id)
    if not ok:
        status_code = 404 if "no encontrada" in error.lower() else 409
        return jsonify({"ok": False, "error": error}), status_code
    _cache_del("companies_active", "companies_all")
    if _normalize_company_id(session.get("company_id")) == _normalize_company_id(company_id):
        _set_company_session(_default_company_id())
    return jsonify({"ok": True})


@flask_app.post("/api/configuracion/empresa/seleccionar")
@rrhh_auth_required
def configuracion_seleccionar_empresa_api():
    if not _can_manage_configuration():
        return _forbidden_json_error("Sin permiso para cambiar empresa activa.")
    data = request.get_json(silent=True) or {}
    company_id = _normalize_company_id(data.get("company_id"))
    company = _get_company(company_id, include_inactive=True)
    if not company:
        available = _list_companies(include_inactive=True)
        if not available:
            return jsonify({"ok": False, "error": "Empresa no encontrada."}), 404
        company = available[0]
    selected = _set_company_session(company.get("company_id"))
    settings = _read_general_settings()
    _apply_company_branding(settings)
    return jsonify({"ok": True, "company": selected, "settings": settings})


@flask_app.get("/api/configuracion/knowledge")
@rrhh_auth_required
def configuracion_knowledge_get():
    """Devuelve la base de conocimiento (cantidad y opcionalmente entradas) de una empresa."""
    if not _can_config_knowledge():
        return _forbidden_json_error("Sin permiso.")
    company_id = _normalize_company_id(request.args.get("company_id") or "")
    if not company_id:
        return jsonify({"ok": False, "error": "Falta company_id."}), 400
    entries = chatbot.obtener_knowledge_empresa(company_id)
    return jsonify({"ok": True, "count": len(entries), "entries": entries})


def _extract_text_document_ai(pdf_bytes):
    """Extrae texto de un PDF con Document AI. Devuelve (texto, None) o (None, mensaje_error)."""
    project_id = os.getenv("DOCUMENT_AI_PROJECT_ID") or os.getenv("GOOGLE_CLOUD_PROJECT") or ""
    location = os.getenv("DOCUMENT_AI_LOCATION") or "us"
    processor_id = os.getenv("DOCUMENT_AI_PROCESSOR_ID") or ""
    if not project_id or not processor_id:
        return None, "Faltan DOCUMENT_AI_PROJECT_ID y DOCUMENT_AI_PROCESSOR_ID en .env (ver docs/CONECTAR_DOCUMENT_AI.md)."
    try:
        from google.cloud import documentai_v1 as documentai
    except ImportError:
        return None, "Instalá el cliente: pip install google-cloud-documentai"
    try:
        client = documentai.DocumentProcessorServiceClient()
        name = f"projects/{project_id}/locations/{location}/processors/{processor_id}"
        raw_doc = documentai.RawDocument(content=pdf_bytes, mime_type="application/pdf")
        request = documentai.ProcessRequest(name=name, raw_document=raw_doc)
        result = client.process_document(request=request)
        text = (result.document.text or "").strip()
        return text if text else None, None if text else "Document AI no devolvió texto."
    except Exception as e:
        return None, f"Document AI: {str(e)[:200]}"


def _ai_generate_faqs_from_text(text):
    """Usa Gemini (google-genai) para extraer pares pregunta/respuesta de cualquier documento.
    Requiere GEMINI_API_KEY en .env. Si no está configurada, cae al parser rule-based."""
    if not (text or "").strip():
        return []
    api_key = os.getenv("GEMINI_API_KEY", "").strip()
    if not api_key:
        logging.info("GEMINI_API_KEY no configurada; usando parser de FAQs por reglas.")
        return _parse_pdf_text_to_entries(text)
    try:
        import json as _json
        from google import genai as _genai
        _client = _genai.Client(api_key=api_key)
        prompt = (
            "Analizá el siguiente documento de recursos humanos y extraé TODOS los pares pregunta/respuesta "
            "útiles para un empleado. El documento puede ser un manual con secciones, una lista de Q&A, o texto libre.\n\n"
            "Reglas OBLIGATORIAS:\n"
            "- Por cada sección o tema del documento, generá AL MENOS una pregunta natural que haría un empleado.\n"
            "- Las respuestas deben ser completas y basadas ÚNICAMENTE en el documento.\n"
            "- Si hay 10 secciones, debe haber al menos 10 pares. Si hay 5, al menos 5.\n"
            "- Respondé ÚNICAMENTE con el array JSON, sin texto antes ni después:\n"
            '[{"pregunta": "¿...?", "respuesta": "..."}, {"pregunta": "¿...?", "respuesta": "..."}, ...]\n'
            "- No incluyas explicaciones, solo el JSON.\n\n"
            f"Documento:\n{text[:15000]}"
        )
        from google.genai import types as _gtypes
        response = _client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config=_gtypes.GenerateContentConfig(
                thinking_config=_gtypes.ThinkingConfig(thinking_budget=0)
            ),
        )
        content = (response.text or "").strip()
        # Extraer JSON del response: buscar último [{ ... }] para ignorar texto previo
        start = content.rfind("[{")
        end = content.rfind("}]")
        if start != -1 and end != -1 and end > start:
            content = content[start:end + 2]
        elif content.startswith("```"):
            content = re.sub(r"^```(?:json)?\n?", "", content)
            content = re.sub(r"\n?```$", "", content)
        entries_raw = _json.loads(content)
        if not isinstance(entries_raw, list):
            return []
        result = []
        for e in entries_raw:
            if isinstance(e, dict):
                p = str(e.get("pregunta") or "").strip()[:500]
                r = str(e.get("respuesta") or "").strip()[:8000]
                if p or r:
                    result.append({"pregunta": p, "respuesta": r})
        logging.info(f"Gemini FAQ generation: {len(result)} pares extraídos. Raw inicio: {(response.text or '')[:300]!r}")
        # Comparar con el parser por reglas: si el parser detecta más secciones, usarlo
        rule_result = _parse_pdf_text_to_entries(text)
        logging.info(f"Parser por reglas: {len(rule_result)} secciones.")
        if len(rule_result) > len(result):
            return rule_result
        return result
    except Exception as exc:
        logging.warning(f"Gemini FAQ generation falló ({exc}); usando parser por reglas.")
        return _parse_pdf_text_to_entries(text)


def _parse_pdf_text_to_entries(text):
    """Convierte texto extraído en lista de {pregunta, respuesta}.
    Acepta: 'Pregunta: ... Respuesta: ...' o manuales con secciones en MAYÚSCULAS."""
    if not (text or "").strip():
        return []
    entries = []
    text = text.strip()
    # Patrón "Pregunta:" / "Respuesta:" (o P: / R:)
    blocks = re.split(r"\s*(?:Pregunta|P):\s*", text, flags=re.I)
    for block in blocks:
        if not block.strip():
            continue
        parts = re.split(r"\s*(?:Respuesta|R):\s*", block, maxsplit=1, flags=re.I)
        if len(parts) >= 2:
            p, r = parts[0].strip(), parts[1].strip()
            if p or r:
                entries.append({"pregunta": p[:500], "respuesta": r[:8000]})
    if entries:
        return entries
    # Fallback: detectar encabezados MAYÚSCULAS línea por línea
    import unicodedata

    def _es_encabezado(ln):
        s = unicodedata.normalize("NFC", ln.strip())
        if not s or len(s) > 80 or len(s) < 3:
            return False
        if s.endswith(".") or s.endswith(","):
            return False
        # Contar letras y verificar que todas las letras sean mayúsculas
        letras = [c for c in s if c.isalpha()]
        if not letras:
            return False
        uppercase_ratio = sum(1 for c in letras if c.isupper()) / len(letras)
        return uppercase_ratio >= 0.85  # al menos 85% mayúsculas

    lines = [ln.rstrip() for ln in text.splitlines()]
    current_heading = ""
    current_body_lines = []

    def _flush():
        if current_body_lines:
            body = " ".join(ln for ln in current_body_lines if ln)
            if body.strip():
                heading = current_heading or (body[:80] + "..." if len(body) > 80 else body)
                entries.append({"pregunta": heading[:500], "respuesta": body.strip()[:8000]})

    for ln in lines:
        stripped = ln.strip()
        if not stripped:
            continue
        if _es_encabezado(stripped):
            _flush()
            current_heading = stripped
            current_body_lines = []
        else:
            current_body_lines.append(stripped)
    _flush()
    logging.info(f"Parser por reglas: {len(entries)} secciones detectadas.")
    return entries


def _parse_knowledge_file(file_storage):
    """Parsea CSV, Excel o PDF (pregunta, respuesta) y devuelve lista de dicts."""
    if not file_storage or not file_storage.filename:
        return [], "No se envió ningún archivo."
    raw = file_storage.read()
    if not raw:
        return [], "El archivo está vacío."
    filename = (file_storage.filename or "").lower()
    entries = []
    if filename.endswith(".csv"):
        import csv
        import io
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            return [], "No se pudo decodificar el archivo (usá UTF-8 o Latin-1)."
        reader = csv.DictReader(io.StringIO(text), delimiter=",")
        for row in reader:
            p = (row.get("pregunta") or row.get("Pregunta") or "").strip()
            r = (row.get("respuesta") or row.get("Respuesta") or "").strip()
            if p or r:
                entries.append({"pregunta": p, "respuesta": r})
        if not entries and reader.fieldnames:
            return [], "El CSV no tiene columnas 'pregunta' y 'respuesta' (o Pregunta/Respuesta)."
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            import openpyxl
        except ImportError:
            return [], "Para subir Excel instalá: pip install openpyxl"
        try:
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(raw), read_only=True)
            sh = wb.active
            headers = [str(c.value or "").strip().lower() for c in sh[1]]
            ip = ipr = -1
            for i, h in enumerate(headers):
                if "pregunta" in (h or ""):
                    ip = i
                if "respuesta" in (h or ""):
                    ipr = i
            if ip < 0 or ipr < 0:
                wb.close()
                return [], "La hoja debe tener columnas 'Pregunta' y 'Respuesta'."
            for row in sh.iter_rows(min_row=2):
                cells = [row[i].value if i < len(row) else None for i in range(max(ip, ipr) + 1)]
                p = str(cells[ip] or "").strip()
                r = str(cells[ipr] or "").strip()
                if p or r:
                    entries.append({"pregunta": p, "respuesta": r})
            wb.close()
        except Exception as e:
            return [], f"Error al leer Excel: {e}"
    elif filename.endswith(".pdf"):
        text, err = _extract_text_document_ai(raw)
        if err:
            return [], err
        entries = _ai_generate_faqs_from_text(text or "")
        if not entries:
            return [], "No se encontraron preguntas/respuestas en el PDF. Verificá que el documento tenga contenido de RRHH relevante."
    elif filename.endswith(".txt"):
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            return [], "No se pudo decodificar el archivo de texto."
        entries = _ai_generate_faqs_from_text(text)
        if not entries:
            return [], "No se encontraron preguntas/respuestas en el archivo. Verificá el formato."
    else:
        return [], "Formato no soportado. Usá CSV (.csv), Excel (.xlsx), PDF (.pdf) o texto plano (.txt)."
    return entries, None


@flask_app.post("/api/configuracion/knowledge/upload")
@rrhh_auth_required
def configuracion_knowledge_upload():
    """Sube un archivo CSV, Excel o PDF (Pregunta/Respuesta) y guarda la base de conocimiento. En PDF se usa Document AI para extraer texto."""
    if not _can_config_knowledge():
        return _forbidden_json_error("Sin permiso para gestionar base de conocimiento.")
    company_id = _normalize_company_id(request.form.get("company_id") or "")
    if not company_id:
        return jsonify({"ok": False, "error": "Falta company_id."}), 400
    if not _get_company(company_id, include_inactive=True):
        return jsonify({"ok": False, "error": "Empresa no encontrada."}), 404
    entries, err = _parse_knowledge_file(request.files.get("file"))
    if err:
        return jsonify({"ok": False, "error": err}), 400
    if not entries:
        return jsonify({"ok": False, "error": "No se encontraron filas con pregunta o respuesta."}), 400
    if not chatbot.guardar_company_knowledge(company_id, entries):
        return jsonify({"ok": False, "error": "No se pudo guardar (revisá Firestore)."}), 500
    _auto_update_temas_from_knowledge(company_id, entries)
    return jsonify({"ok": True, "count": len(entries), "message": f"Se cargaron {len(entries)} preguntas/respuestas."})


@flask_app.get("/api/configuracion/knowledge/plantilla")
@rrhh_auth_required
def configuracion_knowledge_plantilla():
    """Descarga una plantilla CSV de ejemplo para base de conocimiento."""
    from flask import Response
    csv_content = "\ufeffpregunta,respuesta\n"
    csv_content += "¿Cuántos días de vacaciones tengo?,Según antigüedad: hasta 5 años 14 días, hasta 10 años 21 días.\n"
    csv_content += "¿Dónde obtengo mi recibo?,En la intranet, sección Mi Legajo, el cuarto día hábil de cada mes.\n"
    return Response(csv_content, mimetype="text/csv", headers={
        "Content-Disposition": "attachment; filename=base_conocimiento_empresa.csv",
    })


def _sync_knowledge_from_drive(company_id, folder_id):
    """
    Lista archivos en la carpeta de Drive, extrae texto (Docs/Sheets export, PDF con Document AI,
    archivos de texto), usa Gemini AI para generar pares Pregunta/Respuesta y guarda en
    company_knowledge. Devuelve (total_entries, error).
    """
    if not folder_id or not str(folder_id).strip():
        return 0, "Falta el ID de la carpeta de Drive. Configuralo en la empresa o pasalo en la solicitud."
    folder_id = str(folder_id).strip()
    # Aceptar URL completa de Drive: extraer el ID de /folders/<ID>
    _m = re.search(r"/folders/([a-zA-Z0-9_-]+)", folder_id)
    if _m:
        folder_id = _m.group(1)
    # Limpiar parámetros pegados al ID (ej: "ID?usp=drive_link")
    if "?" in folder_id:
        folder_id = folder_id.split("?")[0].strip()
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseDownload
    except ImportError:
        return 0, "Instalá: pip install google-api-python-client google-auth"
    _DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
    creds = None
    # 1) JSON inline en variable de entorno GOOGLE_DRIVE_CREDENTIALS_JSON
    _creds_json = os.getenv("GOOGLE_DRIVE_CREDENTIALS_JSON", "").strip()
    if _creds_json:
        try:
            import json as _json
            _info = _json.loads(_creds_json)
            creds = service_account.Credentials.from_service_account_info(_info, scopes=_DRIVE_SCOPES)
        except Exception as _e:
            return 0, f"GOOGLE_DRIVE_CREDENTIALS_JSON inválido: {_e}"
    # 2) Ruta a archivo en GOOGLE_APPLICATION_CREDENTIALS
    if creds is None:
        creds_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
        if creds_path and os.path.isfile(creds_path):
            try:
                creds = service_account.Credentials.from_service_account_file(creds_path, scopes=_DRIVE_SCOPES)
            except Exception as _e:
                return 0, f"Error leyendo credenciales desde archivo: {_e}"
    if creds is None:
        return 0, "Configurá GOOGLE_DRIVE_CREDENTIALS_JSON (JSON de la cuenta de servicio) en las variables de entorno."
    try:
        drive = build("drive", "v3", credentials=creds, cache_discovery=False)
    except Exception as e:
        return 0, f"Error al conectar con Drive: {str(e)[:200]}"
    # ── Helper: extraer texto de un archivo de Drive ─────────────────────────
    def _extraer_texto_drive_file(f):
        from io import BytesIO
        fid = f.get("id")
        fname = (f.get("name") or "").lower()
        fmime = (f.get("mimeType") or "").lower()
        if "folder" in fmime:
            return None
        try:
            if "document" in fmime or "presentation" in fmime or fname.endswith((".gdoc", ".gslides")):
                content = drive.files().export(fileId=fid, mimeType="text/plain").execute()
                return content.decode("utf-8", errors="replace") if isinstance(content, bytes) else (content or "")
            elif "spreadsheet" in fmime or fname.endswith(".gsheet"):
                content = drive.files().export(fileId=fid, mimeType="text/csv").execute()
                return content.decode("utf-8", errors="replace") if isinstance(content, bytes) else (content or "")
            elif "pdf" in fmime or fname.endswith(".pdf"):
                req = drive.files().get_media(fileId=fid)
                buf = BytesIO()
                dl = MediaIoBaseDownload(buf, req)
                done = False
                while not done:
                    _, done = dl.next_chunk()
                raw = buf.getvalue()
                text, err = _extract_text_document_ai(raw)
                return text if not err and text else None
            elif fmime.startswith("text/") or fname.endswith(".txt"):
                req = drive.files().get_media(fileId=fid)
                buf = BytesIO()
                dl = MediaIoBaseDownload(buf, req)
                done = False
                while not done:
                    _, done = dl.next_chunk()
                return buf.getvalue().decode("utf-8", errors="replace")
        except Exception as _ex:
            logging.warning("_extraer_texto_drive_file: error exportando '%s' (%s): %s", f.get("name"), fmime, _ex)
            return None
        return None

    # ── Helper: procesar lista de archivos con un tag de convenio ─────────────
    def _procesar_archivos_drive(archivos, convenio_tag):
        resultado = []
        for f in archivos:
            texto = _extraer_texto_drive_file(f)
            if texto:
                nuevas = _ai_generate_faqs_from_text(texto)
                for e in nuevas:
                    if convenio_tag:
                        e["convenio"] = convenio_tag
                resultado.extend(nuevas)
            else:
                logging.warning("_procesar_archivos_drive: sin texto en '%s'", f.get("name"))
        return resultado

    # ── Helper: mapear nombre de subcarpeta → convenio ────────────────────────
    company_cfg = _get_company(company_id, include_inactive=True)
    convenios_empresa = [
        str(c.get("nombre") or "").strip().lower()
        for c in ((company_cfg or {}).get("convenios") or [])
        if (c.get("nombre") or "").strip()
    ]

    def _carpeta_a_convenio(folder_name):
        """Mapea el nombre de una subcarpeta al convenio correspondiente o None si es genérica."""
        n = folder_name.strip().lower()
        # Carpetas genéricas (sin convenio)
        if any(g in n for g in ("generica", "general", "todos", "global")):
            return None
        # Quitar prefijo "faqs " si existe
        nombre_limpio = n[5:].strip() if n.startswith("faqs ") else n
        if not nombre_limpio or nombre_limpio in ("generica", "general", "todos"):
            return None
        # Buscar coincidencia con convenios configurados
        for conv in convenios_empresa:
            if conv and (conv in nombre_limpio or nombre_limpio in conv):
                return conv
        # Si no matchea ningún convenio configurado, usar el nombre como convenio
        return nombre_limpio

    # ── Listar contenido de la carpeta raíz (archivos + subcarpetas) ──────────
    try:
        q = f"'{folder_id}' in parents and trashed = false"
        items_list = drive.files().list(
            q=q,
            pageSize=100,
            fields="files(id, name, mimeType)",
            orderBy="name",
        ).execute()
        items = items_list.get("files") or []
    except Exception as e:
        return 0, f"Error al listar carpeta: {str(e)[:200]}. ¿Compartiste la carpeta con el email de la cuenta de servicio?"

    archivos_raiz = [f for f in items if "folder" not in (f.get("mimeType") or "").lower()]
    subcarpetas = [f for f in items if "folder" in (f.get("mimeType") or "").lower()]

    logging.info("Drive sync: carpeta raíz tiene %d archivos y %d subcarpetas", len(archivos_raiz), len(subcarpetas))

    all_entries = []
    diag_lines = []  # diagnóstico para mostrar en el error si falla

    # Archivos directamente en la raíz → genéricos (sin convenio)
    if archivos_raiz:
        all_entries.extend(_procesar_archivos_drive(archivos_raiz, None))

    # Subcarpetas → cada una con su convenio (o genérica si el nombre lo indica)
    for carpeta in subcarpetas:
        nombre_carpeta = carpeta.get("name") or ""
        convenio_tag = _carpeta_a_convenio(nombre_carpeta)
        try:
            sub_q = f"'{carpeta['id']}' in parents and trashed = false"
            sub_result = drive.files().list(
                q=sub_q,
                pageSize=50,
                fields="files(id, name, mimeType)",
                orderBy="name",
            ).execute()
            sub_files = [f for f in (sub_result.get("files") or []) if "folder" not in (f.get("mimeType") or "").lower()]
            logging.info("Drive sync: subcarpeta '%s' → %d archivos (convenio=%s)", nombre_carpeta, len(sub_files), convenio_tag)
            diag_lines.append(f"'{nombre_carpeta}': {len(sub_files)} archivo(s)")
        except Exception as _se:
            logging.warning("Drive sync: error listando subcarpeta '%s': %s", nombre_carpeta, _se)
            diag_lines.append(f"'{nombre_carpeta}': error al listar ({str(_se)[:80]})")
            continue
        all_entries.extend(_procesar_archivos_drive(sub_files, convenio_tag))

    if not all_entries:
        diag = ("; ".join(diag_lines)) if diag_lines else "carpeta vacía o sin archivos soportados"
        return 0, f"No se extrajeron preguntas/respuestas. Diagnóstico: {diag}. Verificá que los documentos tengan contenido y que la carpeta esté compartida con la cuenta de servicio."
    if not chatbot.guardar_company_knowledge(company_id, all_entries):
        return 0, "No se pudo guardar en Firestore."
    # Auto-generar temas del menú a partir de las preguntas extraídas
    _auto_update_temas_from_knowledge(company_id, all_entries)
    return len(all_entries), None


_KB_TOPIC_KEYWORDS = [
    ("vacaciones",      ["vacacion", "dias vacaciones", "licencia anual", "vacaciones anuales",
                         "licencias anuales", "dias de vacacion"]),
    ("recibo",          ["recibo de sueldo", "recibo", "boleta de sueldo", "firma digital"]),
    ("adelanto",        ["adelanto de sueldo", "adelanto", "anticipo de sueldo", "adelanto de salario",
                         "adelanto sueldo"]),
    ("aguinaldo",       ["aguinaldo", " sac ", "sueldo anual complementario", "anual complementario"]),
    ("licencias",       ["licencia por", "licencia especial", "licencias especiales", "licencia matrimonio",
                         "licencia nacimiento", "licencia fallecimiento", "licencia examen",
                         "licencias por", "licencias especial", "dias corridos por", "permiso especial",
                         "dias de licencia", "licencias especiales"]),
    ("art",             ["art ", " art,", "accidente de trabajo", "accidente laboral", "aseguradora",
                         "riesgos del trabajo", "galeno art", "formulario de denuncia", "art\n", "\nart"]),
    ("home office",     ["home office", "trabajo remoto", "teletrabajo", "trabajar desde", "dias remotos",
                         "trabajo desde casa", "politica de home", "dias por semana", "modalidad remota",
                         "trabajo hibrido", "modalidad hibrida"]),
    ("uniforme",        ["uniforme", "ropa de trabajo", "elementos de trabajo", "deposito de suministros",
                         "equipamiento", "indumentaria"]),
    ("capacitacion",    ["capacitacion", "capacitar", "cursos", "formacion", "instituto xyz",
                         "convenio de capacitacion", "entrenamiento", "training"]),
    ("obra social",     ["obra social", "cobertura medica", "cobertura de salud", "osde", "prepaga",
                         "cambio de plan", "agregar familiares", "medicina prepaga"]),
    ("contacto",        ["contacto rrhh", "comunicarse con", "area de recursos humanos", "interno 200",
                         "correo de rrhh", "horario de atencion", "oficina de rrhh", "contacto de rrhh",
                         "datos de contacto", "como contactar"]),
    ("nacimiento",      ["nacimiento", "maternidad", "paternidad", "licencia por hijo", "adopcion"]),
    ("casamiento",      ["casamiento", "matrimonio", "boda"]),
    ("fraccionamiento", ["fraccionamiento", "fraccionar"]),
]


_HEADING_TO_TEMA = {
    "licencias": "licencias", "licencia": "licencias",
    "licencias especiales": "licencias",
    "home office": "home office", "teletrabajo": "home office",
    "adelanto": "adelanto", "adelanto de sueldo": "adelanto",
    "aguinaldo": "aguinaldo", "sac": "aguinaldo",
    "vacaciones": "vacaciones",
    "recibo": "recibo", "recibo de sueldo": "recibo",
    "uniforme": "uniforme",
    "art": "art",
    "capacitacion": "capacitacion",
    "obra social": "obra social",
    "contacto": "contacto", "contacto rrhh": "contacto",
    "nacimiento": "nacimiento", "maternidad": "nacimiento",
    "casamiento": "casamiento", "matrimonio": "casamiento",
    "fraccionamiento": "fraccionamiento",
}


def _extraer_temas_de_knowledge(entries):
    """Extrae temas del menú buscando keywords en pregunta Y respuesta de cada entrada."""
    temas_encontrados = []
    for entry in entries:
        pregunta_norm = chatbot.normalizar_texto(entry.get("pregunta") or "")
        respuesta_norm = chatbot.normalizar_texto(entry.get("respuesta") or "")
        texto_completo = pregunta_norm + " " + respuesta_norm
        if not texto_completo.strip():
            continue
        matched = False
        # Primero: ¿el heading es exactamente un nombre de tema conocido?
        tema_directo = _HEADING_TO_TEMA.get(pregunta_norm.strip())
        if tema_directo:
            if tema_directo not in temas_encontrados:
                temas_encontrados.append(tema_directo)
            matched = True
        else:
            for tema, keywords in _KB_TOPIC_KEYWORDS:
                if any(kw in texto_completo for kw in keywords):
                    if tema not in temas_encontrados:
                        temas_encontrados.append(tema)
                    matched = True
                    break
        if not matched:
            # Extraer primera palabra significativa (>=4 letras) de pregunta o respuesta
            _STOPWORDS = {"cuant", "donde", "como", "cuales", "puedo", "sobre", "cuanta",
                          "tengo", "hacer", "pedir", "quien", "queda", "cual", "para",
                          "esta", "este", "sido", "sera", "tipo", "toda", "todo", "otra", "otro",
                          "debo", "debe", "deben", "hay", "que"}
            for texto_fb in (pregunta_norm, respuesta_norm):
                found_fb = False
                for w in texto_fb.split():
                    if len(w) >= 4 and w not in _STOPWORDS:
                        if w not in temas_encontrados:
                            temas_encontrados.append(w)
                        found_fb = True
                        break
                if found_fb:
                    break
    return temas_encontrados[:12]


def _auto_update_temas_from_knowledge(company_id, entries):
    """Actualiza temas_habilitados de la empresa a partir de las preguntas de la KB. Retorna la lista de temas."""
    if not entries or not company_id:
        return []
    temas_clean = _extraer_temas_de_knowledge(entries)
    if temas_clean and chatbot.db:
        try:
            cid_key = _normalize_company_id(company_id)
            chatbot.db.collection(COMPANIES_COLLECTION).document(cid_key).set(
                {"temas_habilitados": temas_clean}, merge=True
            )
            for ck in ("companies_active", "companies_all"):
                _cache_set(ck, None)
            logging.info(f"Auto-temas para {company_id}: {temas_clean}")
        except Exception as exc:
            logging.warning(f"_auto_update_temas_from_knowledge: error al guardar en Firestore: {exc}")
    return temas_clean


@flask_app.post("/api/configuracion/knowledge/sync-from-drive")
@rrhh_auth_required
def configuracion_knowledge_sync_from_drive():
    """Sincroniza la base de conocimiento desde una carpeta de Google Drive usando IA (Gemini) para extraer FAQs de cualquier documento."""
    if not _can_config_knowledge():
        return _forbidden_json_error("Sin permiso para gestionar base de conocimiento.")
    data = request.get_json(silent=True) or {}
    company_id = _normalize_company_id(data.get("company_id") or "")
    if not company_id:
        return jsonify({"ok": False, "error": "Falta company_id."}), 400
    if not _get_company(company_id, include_inactive=True):
        return jsonify({"ok": False, "error": "Empresa no encontrada."}), 404
    folder_id = str(data.get("folder_id") or "").strip()
    company = _get_company(company_id, include_inactive=True)
    saved_folder_id = (company.get("drive_folder_id") or "").strip()
    if not folder_id:
        folder_id = saved_folder_id
    if not folder_id:
        return jsonify({"ok": False, "error": "Indicá folder_id en el cuerpo o configurá 'Carpeta Drive' en la empresa."}), 400
    # Si el folder_id fue ingresado manualmente y difiere del guardado, guardarlo en la empresa
    if folder_id != saved_folder_id and chatbot.db:
        try:
            chatbot.db.collection(COMPANIES_COLLECTION).document(company_id).set(
                {"drive_folder_id": folder_id}, merge=True
            )
            _cache_del("companies_active", "companies_all")
        except Exception as exc:
            logging.warning(f"No se pudo guardar drive_folder_id en empresa {company_id}: {exc}")
    count, err = _sync_knowledge_from_drive(company_id, folder_id)
    if err:
        return jsonify({"ok": False, "error": err}), 400
    return jsonify({"ok": True, "count": count, "folder_id": folder_id, "message": f"Se sincronizaron {count} preguntas/respuestas desde Drive."})


@flask_app.post("/api/configuracion/knowledge/regenerar-temas")
@rrhh_auth_required
def configuracion_knowledge_regenerar_temas():
    """Regenera los temas del menú a partir de la base de conocimiento existente (sin re-sincronizar Drive)."""
    if not _can_config_knowledge():
        return _forbidden_json_error("Sin permiso para regenerar temas.")
    data = request.get_json(silent=True) or {}
    company_id = _normalize_company_id(data.get("company_id") or "")
    if not company_id:
        return jsonify({"ok": False, "error": "Falta company_id."}), 400
    entries = chatbot.obtener_knowledge_empresa(company_id)
    if not entries:
        return jsonify({"ok": False, "error": "La empresa no tiene base de conocimiento cargada."}), 404
    temas = _auto_update_temas_from_knowledge(company_id, entries)
    _cache_del("companies_active", "companies_all")
    # Diagnóstico: qué pregunta tiene cada entrada
    entradas_debug = [{"pregunta": e.get("pregunta", "")[:80], "resp_preview": (e.get("respuesta") or "")[:60]} for e in entries]
    if temas:
        msg = f"Se generaron {len(temas)} temas: {', '.join(temas)}."
    else:
        preguntas_preview = "; ".join(e.get("pregunta", "")[:40] for e in entries[:3])
        msg = f"No se detectaron temas en {len(entries)} entradas. Preguntas: {preguntas_preview}"
    return jsonify({"ok": True, "temas": temas, "entradas_debug": entradas_debug, "message": msg})


@flask_app.post("/api/configuracion/conversaciones/autocierre/ejecutar")
@rrhh_auth_required
def configuracion_autocierre_ejecutar_api():
    if not _can_manage_general_config():
        return _forbidden_json_error("Sin permiso para ejecutar autocierre.")
    company = _current_company()
    result = _auto_close_expired_handoffs(company=company, force=True)
    return jsonify({"ok": True, **result})


@flask_app.get("/api/rrhh/conversaciones")
@rrhh_permission_required(
    auth_rrhh.PERM_CONVERSATIONS_VIEW,
    message="Sin permiso para ver conversaciones.",
)
def rrhh_conversaciones_api():
    _heartbeat_current_agent()
    include_closed = str(request.args.get("include_closed", "false")).lower() == "true"
    include_all_filters = str(request.args.get("include_all_filters", "false")).lower() == "true"
    branches_param = request.args.get("branches")
    if branches_param is not None:
        branches = [b.strip() for b in str(branches_param).split(",") if b.strip()]
    else:
        branches = None
    areas_param = request.args.get("areas")
    if areas_param is not None:
        areas = [a.strip() for a in str(areas_param).split(",") if a.strip()]
    else:
        areas = None
    company = _current_company()
    company_id_norm = _normalize_company_id(company.get("company_id"))
    default_branches = None
    default_areas = None
    if _auth_enabled() and not include_all_filters and branches is None and areas is None:
        user = _current_rrhh_user()
        if user:
            assignments = user.get("assignments") or []
            branch_set = set()
            for a in assignments:
                if isinstance(a, dict) and _normalize_company_id(a.get("company_id")) == company_id_norm:
                    b = str(a.get("branch") or "").strip()
                    if b:
                        branch_set.add(b)
            if branch_set:
                default_branches = list(branch_set)
            user_area = str(user.get("area") or "").strip()
            if user_area:
                default_areas = [user_area]
    if branches is None and default_branches is not None:
        branches = default_branches
    if areas is None and default_areas is not None:
        areas = default_areas
    auto_close_result = _auto_close_expired_handoffs(company=company)
    convs = _list_handoffs(
        include_closed=include_closed,
        limit=150,
        company_id=company.get("company_id"),
        branches=branches,
        areas=areas,
    )
    company_branches = []
    for b in (company.get("branches") or []):
        name = _branch_name(b)
        if name:
            company_branches.append({"name": name})
    company_areas = _get_all_areas_for_company(company)
    out = {
        "ok": True,
        "conversaciones": [_serialize_handoff(c) for c in convs],
        "agente_actual": _rrhh_agent_name() if _auth_enabled() else "",
        "agentes_activos": [
            _serialize_active_agent(agent)
            for agent in _list_active_agents(company_id=company.get("company_id"))
        ],
        "selected_company_id": company.get("company_id"),
        "selected_company_name": company.get("company_name"),
        "branches": company_branches,
        "areas": [str(a).strip() for a in (company_areas or []) if str(a).strip()],
        "autocierre": auto_close_result,
    }
    if default_branches is not None:
        out["filter_by_branches"] = default_branches
    if default_areas is not None:
        out["filter_by_areas"] = default_areas
    return jsonify(out)


@flask_app.get("/api/rrhh/contactos")
@rrhh_permission_required(
    auth_rrhh.PERM_CONVERSATIONS_VIEW,
    message="Sin permiso para ver contactos.",
)
def rrhh_contactos_api():
    """Lista contactos (teléfono + nombre) de conversaciones previas por WhatsApp para la empresa activa."""
    company = _current_company()
    company_id = _normalize_company_id(company.get("company_id"))
    if not company_id:
        return jsonify({"ok": True, "contactos": []})
    convs = _list_handoffs(include_closed=True, limit=500, company_id=company_id)
    by_phone = {}
    for c in convs:
        if str(c.get("channel") or "").strip().lower() != "whatsapp":
            continue
        phone = (c.get("whatsapp_to_phone") or "").strip()
        if not phone:
            continue
        norm = _normalize_phone_for_match(phone)
        if not norm:
            continue
        nombre = (c.get("colaborador_nombre") or "").strip() or phone
        updated = _as_utc_naive(c.get("updated_at")) or datetime.min
        if norm not in by_phone or (by_phone[norm]["updated_at"] or datetime.min) < updated:
            by_phone[norm] = {"phone": phone, "nombre": nombre, "updated_at": updated}
    contactos = [{"phone": v["phone"], "nombre": v["nombre"]} for v in by_phone.values()]
    contactos.sort(key=lambda x: (x.get("nombre") or "").lower())
    return jsonify({"ok": True, "contactos": contactos})


@flask_app.post("/api/rrhh/empresa/seleccionar")
@rrhh_permission_required(
    auth_rrhh.PERM_CONVERSATIONS_VIEW,
    message="Sin permiso para cambiar empresa activa.",
)
def rrhh_seleccionar_empresa_api():
    data = request.get_json(silent=True) or {}
    company_id = _normalize_company_id(data.get("company_id"))
    if not company_id:
        return jsonify({"ok": False, "error": "Seleccioná una empresa válida."}), 400

    company = _get_company(company_id, include_inactive=False)
    if not company:
        return jsonify({"ok": False, "error": "Empresa no encontrada o inactiva."}), 404

    current_user = _current_rrhh_user()
    if current_user and not _user_can_access_company(current_user, company_id):
        return _forbidden_json_error("No tenés acceso a la empresa seleccionada.")

    selected = _set_company_session(company.get("company_id"))
    _apply_company_branding(_read_general_settings())
    return jsonify(
        {
            "ok": True,
            "company": {
                "company_id": selected.get("company_id"),
                "company_name": selected.get("company_name"),
            },
        }
    )


@flask_app.get("/api/rrhh/usuarios")
@rrhh_permission_required(
    auth_rrhh.PERM_USERS_MANAGE,
    message="Sin permiso para gestionar usuarios.",
)
def rrhh_usuarios_api():
    users = auth_rrhh.list_file_users()
    companies = [_company_for_api(c) for c in _list_companies_for_current_rrhh_user(include_inactive=False)]
    return jsonify(
        {
            "ok": True,
            "users": users,
            "users_file": auth_rrhh.users_file_path(),
            "valid_roles": auth_rrhh.available_roles(),
            "permissions_catalog": auth_rrhh.permissions_catalog(),
            "companies": companies,
        }
    )


@flask_app.post("/api/rrhh/usuarios")
@rrhh_permission_required(
    auth_rrhh.PERM_USERS_MANAGE,
    message="Sin permiso para crear usuarios.",
)
def rrhh_crear_usuario_api():
    data = request.get_json(silent=True) or {}
    username = str(data.get("username") or "").strip()
    password = str(data.get("password") or "")
    display_name = str(data.get("display_name") or "").strip()
    role = str(data.get("role") or "rrhh").strip().lower()
    assignments = data.get("assignments")
    email = str(data.get("email") or "").strip()
    phone = str(data.get("phone") or "").strip()
    area = str(data.get("area") or "").strip()
    created_by = (_current_rrhh_user() or {}).get("username") or ""

    ok, user, error = auth_rrhh.create_user(
        username=username,
        password=password,
        display_name=display_name,
        role=role,
        created_by=created_by,
        assignments=assignments,
        email=email,
        phone=phone,
        area=area,
    )
    if not ok:
        status_code = 409 if "existe" in error.lower() else 400
        return jsonify({"ok": False, "error": error}), status_code

    return jsonify({"ok": True, "user": user})


@flask_app.get("/api/rrhh/roles")
@rrhh_permission_required(
    auth_rrhh.PERM_ROLES_MANAGE,
    message="Sin permiso para ver roles.",
)
def rrhh_roles_api():
    return jsonify(
        {
            "ok": True,
            "roles": auth_rrhh.list_roles(),
            "roles_file": auth_rrhh.roles_file_path(),
            "permissions_catalog": auth_rrhh.permissions_catalog(),
        }
    )


@flask_app.post("/api/rrhh/roles")
@rrhh_permission_required(
    auth_rrhh.PERM_ROLES_MANAGE,
    message="Sin permiso para crear roles.",
)
def rrhh_crear_rol_api():
    data = request.get_json(silent=True) or {}
    name = str(data.get("name") or "").strip().lower()
    display_name = str(data.get("display_name") or "").strip()
    permissions = data.get("permissions")
    company_ids = data.get("company_ids")
    if isinstance(company_ids, list):
        company_ids = [str(c or "").strip() for c in company_ids if str(c or "").strip()]
    else:
        company_ids = None
    ok, role, error = auth_rrhh.create_role(
        name=name,
        display_name=display_name,
        permissions=permissions,
        company_ids=company_ids,
    )
    if not ok:
        status = 409 if "existe" in error.lower() else 400
        return jsonify({"ok": False, "error": error}), status
    return jsonify({"ok": True, "role": role})


@flask_app.post("/api/rrhh/roles/<role_name>")
@rrhh_permission_required(
    auth_rrhh.PERM_ROLES_MANAGE,
    message="Sin permiso para editar roles.",
)
def rrhh_editar_rol_api(role_name):
    data = request.get_json(silent=True) or {}
    display_name = data.get("display_name")
    permissions = data.get("permissions")
    company_ids = data.get("company_ids")
    if company_ids is not None and isinstance(company_ids, list):
        company_ids = [str(c or "").strip() for c in company_ids if str(c or "").strip()]
    ok, role, error = auth_rrhh.update_role(
        name=role_name,
        display_name=display_name,
        permissions=permissions,
        company_ids=company_ids,
    )
    if not ok:
        msg = error.lower()
        if "no encontrado" in msg:
            status = 404
        elif "debe quedar al menos un usuario con permiso" in msg:
            status = 409
        else:
            status = 400
        return jsonify({"ok": False, "error": error}), status
    return jsonify({"ok": True, "role": role})


@flask_app.post("/api/rrhh/usuarios/<username>/rol")
@rrhh_permission_required(
    auth_rrhh.PERM_USERS_MANAGE,
    message="Sin permiso para editar roles de usuarios.",
)
def rrhh_actualizar_rol_api(username):
    data = request.get_json(silent=True) or {}
    role = str(data.get("role") or "").strip().lower()
    updated_by = (_current_rrhh_user() or {}).get("username") or ""

    ok, user, error = auth_rrhh.update_user_role(
        username=username,
        role=role,
        updated_by=updated_by,
    )
    if not ok:
        msg = error.lower()
        if "no encontrado" in msg:
            status_code = 404
        elif "debe quedar al menos un usuario con permiso" in msg:
            status_code = 409
        else:
            status_code = 400
        return jsonify({"ok": False, "error": error}), status_code
    return jsonify({"ok": True, "user": user})


@flask_app.post("/api/rrhh/usuarios/<username>/asignaciones")
@rrhh_permission_required(
    auth_rrhh.PERM_USERS_MANAGE,
    message="Sin permiso para editar asignaciones de usuarios.",
)
def rrhh_actualizar_asignaciones_api(username):
    data = request.get_json(silent=True) or {}
    assignments = data.get("assignments")
    updated_by = (_current_rrhh_user() or {}).get("username") or ""
    ok, user, error = auth_rrhh.update_user_assignments(
        username=username,
        assignments=assignments,
        updated_by=updated_by,
    )
    if not ok:
        msg = error.lower()
        status_code = 404 if "no encontrado" in msg else 400
        return jsonify({"ok": False, "error": error}), status_code
    return jsonify({"ok": True, "user": user})


@flask_app.post("/api/rrhh/usuarios/<username>/perfil")
@rrhh_permission_required(
    auth_rrhh.PERM_USERS_MANAGE,
    message="Sin permiso para editar usuarios.",
)
def rrhh_actualizar_perfil_usuario_api(username):
    data = request.get_json(silent=True) or {}
    updated_by = (_current_rrhh_user() or {}).get("username") or ""
    password_raw = (data.get("password") or "").strip() or None
    ok, user, error = auth_rrhh.update_user_profile(
        username=username,
        role=data.get("role"),
        assignments=data.get("assignments"),
        display_name=data.get("display_name"),
        email=data.get("email"),
        phone=data.get("phone"),
        area=data.get("area"),
        password=password_raw,
        updated_by=updated_by,
    )
    if not ok:
        msg = str(error or "").lower()
        if "no encontrado" in msg:
            status_code = 404
        elif "debe quedar al menos un usuario con permiso" in msg:
            status_code = 409
        else:
            status_code = 400
        return jsonify({"ok": False, "error": error}), status_code
    return jsonify({"ok": True, "user": user})


@flask_app.delete("/api/rrhh/usuarios/<username>")
@rrhh_permission_required(
    auth_rrhh.PERM_USERS_MANAGE,
    message="Sin permiso para eliminar usuarios.",
)
def rrhh_eliminar_usuario_api(username):
    deleted_by = (_current_rrhh_user() or {}).get("username") or ""
    ok, error = auth_rrhh.delete_user(username=username, deleted_by=deleted_by)
    if not ok:
        msg = error.lower()
        status_code = 404 if "no encontrado" in msg else 409 if "debe quedar" in msg else 400
        return jsonify({"ok": False, "error": error}), status_code
    return jsonify({"ok": True})


@flask_app.post("/api/rrhh/usuarios/<username>/enviar-reset")
@rrhh_permission_required(
    auth_rrhh.PERM_USERS_MANAGE,
    message="Sin permiso para gestionar usuarios.",
)
def rrhh_enviar_reset_password_api(username):
    data = request.get_json(silent=True) or {}
    requested_by = (_current_rrhh_user() or {}).get("username") or ""
    ok, payload, error = auth_rrhh.create_password_reset_token(
        username=username,
        ttl_minutes=data.get("ttl_minutes", 60),
        requested_by=requested_by,
    )
    if not ok:
        msg = str(error or "").lower()
        status_code = 404 if "no encontrado" in msg else 400
        return jsonify({"ok": False, "error": error}), status_code

    reset_url = url_for("password_reset_page", token=payload.get("token"), _external=True)
    mail_ok, mail_error = _send_password_reset_email(
        to_email=payload.get("email"),
        display_name=payload.get("display_name") or payload.get("username"),
        reset_url=reset_url,
        expires_at_iso=payload.get("expires_at"),
    )
    if not mail_ok:
        return jsonify(
            {
                "ok": True,
                "mail_sent": False,
                "warning": mail_error,
                "reset_url": reset_url,
                "email": payload.get("email"),
                "expires_at": payload.get("expires_at"),
            }
        )

    return jsonify(
        {
            "ok": True,
            "mail_sent": True,
            "email": payload.get("email"),
            "expires_at": payload.get("expires_at"),
        }
    )


@flask_app.delete("/api/rrhh/roles/<role_name>")
@rrhh_permission_required(
    auth_rrhh.PERM_ROLES_MANAGE,
    message="Sin permiso para eliminar roles.",
)
def rrhh_eliminar_rol_api(role_name):
    ok, error = auth_rrhh.delete_role(role_name)
    if not ok:
        msg = error.lower()
        status = 404 if "no encontrado" in msg else 409 if "no podés" in msg else 400
        return jsonify({"ok": False, "error": error}), status
    return jsonify({"ok": True})


@flask_app.get("/api/rrhh/conversaciones/<conversation_id>/mensajes")
@rrhh_permission_required(
    auth_rrhh.PERM_CONVERSATIONS_VIEW,
    message="Sin permiso para ver conversaciones.",
)
def rrhh_mensajes_api(conversation_id):
    conv = _fetch_handoff(conversation_id)
    if not conv:
        return jsonify({"ok": False, "error": "Conversación no encontrada"}), 404
    if not _conversation_matches_selected_company(conv):
        return jsonify({"ok": False, "error": "Conversación no disponible para esta empresa."}), 404
    mensajes = _list_handoff_messages(conversation_id)
    return jsonify(
        {
            "ok": True,
            "conversation_id": conversation_id,
            "estado": conv.get("estado") or HANDOFF_STATUS_PENDING,
            "rrhh_agente": conv.get("rrhh_agente") or "",
            "rrhh_agente_id": conv.get("rrhh_agente_id") or "",
            "rrhh_asignacion_automatica": bool(conv.get("rrhh_asignacion_automatica")),
            "channel": conv.get("channel") or "",
            "colaborador_nombre": conv.get("colaborador_nombre") or "",
            "colaborador_telefono": conv.get("colaborador_telefono") or "",
            "mensajes": _serialize_messages(mensajes),
        }
    )


@flask_app.post("/api/rrhh/conversaciones/<conversation_id>/tomar")
@rrhh_permission_required(
    auth_rrhh.PERM_CONVERSATIONS_MANAGE,
    message="Sin permiso para gestionar conversaciones.",
)
def rrhh_tomar_api(conversation_id):
    data = request.get_json(silent=True) or {}
    agente = _resolve_rrhh_agent(data)
    conv = _fetch_handoff(conversation_id)
    if not conv:
        return jsonify({"ok": False, "error": "Conversación no encontrada"}), 404
    if not _conversation_matches_selected_company(conv):
        return jsonify({"ok": False, "error": "Conversación no disponible para esta empresa."}), 404
    if not _take_handoff(conversation_id, agente):
        return jsonify({"ok": False, "error": "Conversación no encontrada"}), 404
    return jsonify({"ok": True, "conversation_id": conversation_id, "estado": HANDOFF_STATUS_ACTIVE})


@flask_app.post("/api/rrhh/conversaciones/<conversation_id>/reasignar")
@rrhh_permission_required(
    auth_rrhh.PERM_CONVERSATIONS_MANAGE,
    message="Sin permiso para gestionar conversaciones.",
)
def rrhh_reasignar_api(conversation_id):
    conv = _fetch_handoff(conversation_id)
    if not conv:
        return jsonify({"ok": False, "error": "Conversación no encontrada"}), 404
    if not _conversation_matches_selected_company(conv):
        return jsonify({"ok": False, "error": "Conversación no disponible para esta empresa."}), 404
    data = request.get_json(silent=True) or {}
    target_agent = _resolve_target_agent_for_reassignment(data)
    if not target_agent:
        return jsonify({"ok": False, "error": "Agente destino no disponible."}), 400
    actor = _resolve_rrhh_agent({})
    ok, error = _reassign_handoff(
        conversation_id,
        target_agent,
        reasignado_por=actor.get("display_name"),
    )
    if not ok:
        return jsonify({"ok": False, "error": error}), 404
    return jsonify(
        {
            "ok": True,
            "conversation_id": conversation_id,
            "estado": HANDOFF_STATUS_ACTIVE,
            "rrhh_agente": target_agent.get("display_name"),
            "rrhh_agente_id": target_agent.get("agent_id"),
        }
    )


@flask_app.post("/api/rrhh/conversaciones/<conversation_id>/cerrar")
@rrhh_permission_required(
    auth_rrhh.PERM_CONVERSATIONS_MANAGE,
    message="Sin permiso para gestionar conversaciones.",
)
def rrhh_cerrar_api(conversation_id):
    data = request.get_json(silent=True) or {}
    agente = _resolve_rrhh_agent(data)
    conv = _fetch_handoff(conversation_id)
    if not conv:
        return jsonify({"ok": False, "error": "Conversación no encontrada"}), 404
    if not _conversation_matches_selected_company(conv):
        return jsonify({"ok": False, "error": "Conversación no disponible para esta empresa."}), 404
    if not _close_handoff(conversation_id, agente.get("display_name")):
        return jsonify({"ok": False, "error": "Conversación no encontrada"}), 404
    return jsonify({"ok": True, "conversation_id": conversation_id, "estado": HANDOFF_STATUS_CLOSED})


@flask_app.post("/api/rrhh/conversaciones/<conversation_id>/reabrir")
@rrhh_permission_required(
    auth_rrhh.PERM_CONVERSATIONS_MANAGE,
    message="Sin permiso para gestionar conversaciones.",
)
def rrhh_reabrir_api(conversation_id):
    data = request.get_json(silent=True) or {}
    agente = _resolve_rrhh_agent(data)
    conv = _fetch_handoff(conversation_id)
    if not conv:
        return jsonify({"ok": False, "error": "Conversación no encontrada"}), 404
    if not _conversation_matches_selected_company(conv):
        return jsonify({"ok": False, "error": "Conversación no disponible para esta empresa."}), 404
    if not _reopen_handoff(conversation_id, agente.get("display_name")):
        return jsonify({"ok": False, "error": "No se pudo reabrir"}), 400
    return jsonify({"ok": True, "conversation_id": conversation_id, "estado": HANDOFF_STATUS_PENDING})


def _get_company_whatsapp_from(company):
    """Devuelve el número 'from' de WhatsApp de la empresa (primer número configurado) para Twilio."""
    if not company:
        return None
    nums = company.get("whatsapp_numbers") or []
    if not nums:
        return None
    phone = (nums[0].get("phone") or "").strip()
    return phone if phone else None


@flask_app.post("/api/rrhh/conversaciones/iniciar")
@rrhh_permission_required(
    auth_rrhh.PERM_CONVERSATIONS_MANAGE,
    message="Sin permiso para gestionar conversaciones.",
)
def rrhh_iniciar_conversacion_api():
    """Crea una conversación nueva e envía el primer mensaje por WhatsApp a esa persona (sin que te escriba antes)."""
    data = request.get_json(silent=True) or {}
    phone = (data.get("phone") or data.get("telefono") or "").strip()
    mensaje = (data.get("mensaje") or data.get("texto") or "").strip()
    nombre = (data.get("nombre") or "").strip()
    if not phone:
        return jsonify({"ok": False, "error": "Falta el número de teléfono (phone)."}), 400
    if not mensaje:
        return jsonify({"ok": False, "error": "Escribí el mensaje a enviar (mensaje)."}), 400
    company = _current_company()
    company_id = _normalize_company_id(company.get("company_id"))
    if not company_id:
        return jsonify({"ok": False, "error": "Seleccioná una empresa en el panel."}), 400
    from_number = _get_company_whatsapp_from(company)
    if not from_number:
        from_number = os.getenv("TWILIO_WHATSAPP_FROM", "").strip()
        if from_number and not from_number.startswith("whatsapp:"):
            try:
                from twilio_whatsapp import _format_to_whatsapp
                from_number = _format_to_whatsapp(from_number) or from_number
            except Exception:
                pass
    if not from_number:
        return jsonify({
            "ok": False,
            "error": "La empresa no tiene número de WhatsApp configurado. Agregá números en Configuración > Empresas.",
        }), 400
    if from_number.startswith("whatsapp:"):
        from_number_raw = from_number.replace("whatsapp:", "", 1)
    else:
        from_number_raw = from_number
    # Evitar crear duplicado si ya existe handoff abierto para este teléfono
    existing_id = _find_open_handoff_by_whatsapp_phone(phone)
    if existing_id:
        conv = _fetch_handoff(existing_id)
        if conv and _normalize_company_id(conv.get("company_id")) == company_id:
            _add_handoff_message(existing_id, remitente="rrhh", texto=mensaje, agente=_rrhh_agent_name())
            _send_whatsapp_to_collaborator(phone, mensaje, from_number=from_number_raw)
            _upsert_handoff(
                existing_id,
                {"updated_at": _utc_now(), "ultimo_mensaje": mensaje[:200], "ultimo_mensaje_fecha": _utc_now()},
                merge=True,
            )
            return jsonify({"ok": True, "conversation_id": existing_id, "existente": True})
    conversation_id = _new_conversation_id()
    now = _utc_now()
    agente = _current_rrhh_user() or {}
    display_name = str(agente.get("display_name") or agente.get("name") or _rrhh_agent_name() or "RRHH").strip()
    handoff_payload = {
        "conversation_id": conversation_id,
        "company_id": company_id,
        "company_name": company.get("company_name"),
        "branch": "",
        "area": "",
        "estado": HANDOFF_STATUS_PENDING,
        "created_at": now,
        "updated_at": now,
        "channel": "whatsapp",
        "whatsapp_to_phone": phone,
        "whatsapp_from_number": from_number_raw,
        "colaborador_nombre": nombre or phone,
        "colaborador_telefono": phone,
        "rrhh_agente": display_name,
        "rrhh_agente_id": str(agente.get("agent_id") or agente.get("user_id") or "").strip(),
        "ultima_consulta": "(Mensaje iniciado por RRHH)",
        "ultimo_mensaje": mensaje[:200],
        "ultimo_mensaje_fecha": now,
    }
    _upsert_handoff(conversation_id, handoff_payload, merge=False)
    _add_handoff_message(conversation_id, remitente="rrhh", texto=mensaje, agente=display_name)

    pid = _meta_phone_number_id() or None
    if not _send_meta_whatsapp(phone, mensaje, phone_number_id=pid):
        return jsonify({"ok": False, "error": "No se pudo enviar el mensaje por WhatsApp. Verificá que el token Meta esté vigente."}), 502
    return jsonify({"ok": True, "conversation_id": conversation_id})


# Límite de subida para adjuntos (10 MB)
UPLOAD_MAX_BYTES = 10 * 1024 * 1024
UPLOAD_ALLOWED_EXTENSIONS = {
    "jpg", "jpeg", "png", "gif", "webp",
    "pdf",
    "xls", "xlsx",
    "doc", "docx",
    "ppt", "pptx",
    "txt", "csv",
}
_EXT_TO_MIME = {
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
    "gif": "image/gif", "webp": "image/webp", "pdf": "application/pdf",
    "xls": "application/vnd.ms-excel",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "doc": "application/msword",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "ppt": "application/vnd.ms-powerpoint",
    "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    "txt": "text/plain",
    "csv": "text/csv",
}


def _get_storage_bucket():
    """Devuelve el bucket de Firebase Storage si está disponible."""
    try:
        from firebase_admin import storage
        bucket_name = os.getenv("FIREBASE_STORAGE_BUCKET", "").strip()
        if bucket_name:
            return storage.bucket(bucket_name)
        return storage.bucket()
    except Exception:
        return None


def _generate_signed_url(blob, expiration_minutes: int = 15) -> str:
    """Genera un signed URL compatible con Cloud Run (sin clave privada).

    En Cloud Run las credenciales de Compute Engine no tienen clave privada,
    por lo que se usa el token de acceso + service_account_email para que
    la biblioteca llame al API de IAM signBlob en su lugar.
    """
    from datetime import timedelta
    exp = timedelta(minutes=expiration_minutes)
    try:
        return blob.generate_signed_url(expiration=exp, method="GET", version="v4")
    except Exception:
        pass
    # Fallback IAM-based signing (Cloud Run / GCE)
    try:
        import google.auth
        import google.auth.transport.requests as _gtr
        import urllib.request as _ur
        credentials, _ = google.auth.default()
        auth_req = _gtr.Request()
        credentials.refresh(auth_req)
        # Intentar obtener el email de la service account
        sa_email = getattr(credentials, "service_account_email", None)
        if not sa_email:
            meta_req = _ur.Request(
                "http://metadata.google.internal/computeMetadata/v1/instance/service-accounts/default/email",
                headers={"Metadata-Flavor": "Google"},
            )
            sa_email = _ur.urlopen(meta_req, timeout=5).read().decode()
        return blob.generate_signed_url(
            expiration=exp,
            method="GET",
            version="v4",
            service_account_email=sa_email,
            access_token=credentials.token,
        )
    except Exception:
        # Último recurso: URL pública (sin expiración)
        blob.make_public()
        return blob.public_url


def _proxy_twilio_media(url_decoded):
    """Descarga un recurso de Twilio (MediaUrl) con auth y lo devuelve. Solo permite URLs de api.twilio.com."""
    url_decoded = (url_decoded or "").strip()
    if not url_decoded.startswith("https://api.twilio.com/"):
        return None
    sid = os.getenv("TWILIO_ACCOUNT_SID", "").strip()
    token = os.getenv("TWILIO_AUTH_TOKEN", "").strip()
    if not sid or not token:
        return None
    try:
        from base64 import b64encode
        from urllib.request import Request, urlopen
        auth = b64encode(f"{sid}:{token}".encode()).decode()
        req = Request(url_decoded, headers={"Authorization": f"Basic {auth}"})
        resp = urlopen(req, timeout=15)
        content = resp.read()
        content_type = resp.headers.get("Content-Type") or "application/octet-stream"
        from flask import Response
        return Response(content, mimetype=content_type)
    except Exception:
        return None


@flask_app.get("/api/rrhh/media")
@rrhh_permission_required(
    auth_rrhh.PERM_CONVERSATIONS_VIEW,
    message="Sin permiso para ver adjuntos.",
)
def rrhh_media_proxy():
    """Proxy para ver en el panel archivos/imágenes que el colaborador envió por WhatsApp (Twilio MediaUrl)."""
    from urllib.parse import unquote
    u = request.args.get("u", "").strip()
    if not u:
        return ("", 400)
    url_decoded = unquote(u)
    response = _proxy_twilio_media(url_decoded)
    if response is None:
        return ("", 404)
    return response


@flask_app.get("/api/rrhh/meta-diagnostico")
@rrhh_permission_required(auth_rrhh.PERM_CONVERSATIONS_VIEW, message="Sin permiso.")
def rrhh_meta_diagnostico():
    """Endpoint de diagnóstico: verifica token Meta y opcionalmente descarga un media_id."""
    import requests as _req
    token = _meta_access_token()
    pid = _meta_phone_number_id()
    result = {"token_configurado": bool(token), "phone_number_id": pid}
    if not token:
        return jsonify(result)
    # Verificar token con la API de Meta
    try:
        r = _req.get(
            f"https://graph.facebook.com/v18.0/{pid}",
            headers={"Authorization": f"Bearer {token}"},
            timeout=10,
        )
        result["meta_api_status"] = r.status_code
        result["meta_api_response"] = r.json() if r.headers.get("content-type", "").startswith("application/json") else r.text[:300]
    except Exception as e:
        result["meta_api_error"] = str(e)
    # Si se pasa ?media_id=... también prueba descargarlo (paso 1 y paso 2)
    media_id = request.args.get("media_id", "").strip()
    if media_id:
        try:
            r2 = _req.get(
                f"https://graph.facebook.com/v18.0/{media_id}",
                headers={"Authorization": f"Bearer {token}"},
                timeout=10,
            )
            result["media_status"] = r2.status_code
            result["media_response"] = r2.json() if r2.headers.get("content-type", "").startswith("application/json") else r2.text[:300]
            if r2.ok:
                dl_url = (r2.json() or {}).get("url")
                if dl_url:
                    r3 = _req.get(dl_url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
                    result["media_download_status"] = r3.status_code
                    result["media_download_bytes"] = len(r3.content) if r3.ok else 0
                    if not r3.ok:
                        result["media_download_error"] = r3.text[:300]
        except Exception as e:
            result["media_error"] = str(e)
    # Si se pasa ?send_to=NUMERO prueba enviar un mensaje de texto
    send_to = request.args.get("send_to", "").strip()
    if send_to:
        try:
            to_norm = re.sub(r"[^\d]", "", send_to)
            r_send = _req.post(
                f"https://graph.facebook.com/v18.0/{pid}/messages",
                json={"messaging_product": "whatsapp", "to": to_norm, "type": "text", "text": {"body": "Test diagnóstico"}},
                headers={"Authorization": f"Bearer {token}"},
                timeout=15,
            )
            result["send_status"] = r_send.status_code
            result["send_response"] = r_send.json() if r_send.headers.get("content-type", "").startswith("application/json") else r_send.text[:300]
        except Exception as e:
            result["send_error"] = str(e)
    return jsonify(result)


@flask_app.get("/api/rrhh/meta-media")
@rrhh_permission_required(
    auth_rrhh.PERM_CONVERSATIONS_VIEW,
    message="Sin permiso para ver adjuntos.",
)
def rrhh_meta_media_proxy():
    """Proxy para ver en el panel imágenes/archivos que el colaborador envió por Meta WhatsApp."""
    media_id = request.args.get("id", "").strip()
    if not media_id:
        return ("", 400)
    file_bytes, mime_type = _download_meta_media(media_id)
    if not file_bytes:
        return ("No se pudo descargar el archivo de Meta.", 404)
    from flask import Response
    return Response(file_bytes, content_type=mime_type or "application/octet-stream")


@flask_app.post("/api/rrhh/upload")
@rrhh_permission_required(
    auth_rrhh.PERM_CONVERSATIONS_MANAGE,
    message="Sin permiso para subir archivos.",
)
def rrhh_upload_api():
    """Sube un archivo (imagen o PDF) y devuelve una URL pública o firmada para usar como adjunto."""
    url, err = _upload_file_to_storage("handoff_uploads")
    if err:
        if "Storage no configurado" in err:
            return jsonify({"ok": False, "error": err}), 501
        if "supera el límite" in err:
            return jsonify({"ok": False, "error": err}), 400
        if "Tipo no permitido" in err:
            return jsonify({"ok": False, "error": err}), 400
        if "No se envió" in err:
            return jsonify({"ok": False, "error": err}), 400
        return jsonify({"ok": False, "error": err}), 500
    return jsonify({"ok": True, "url": url})


def _upload_file_to_storage(prefix_path: str):
    """Sube el archivo del request a Firebase Storage. prefix_path ej: 'handoff_uploads' o 'comunicados_uploads'."""
    bucket = _get_storage_bucket()
    if not bucket:
        return None, "Storage no configurado (Firebase Storage)."
    file_storage = request.files.get("file")
    if not file_storage or not file_storage.filename:
        return None, "No se envió ningún archivo."
    raw = file_storage.read()
    if len(raw) > UPLOAD_MAX_BYTES:
        return None, f"El archivo supera el límite de {UPLOAD_MAX_BYTES // (1024*1024)} MB."
    ext = (file_storage.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in UPLOAD_ALLOWED_EXTENSIONS:
        return None, f"Tipo no permitido. Permitidos: {', '.join(sorted(UPLOAD_ALLOWED_EXTENSIONS))}."
    content_type = (file_storage.content_type or "").split(";")[0].strip()
    if not content_type or content_type in ("application/octet-stream", "binary/octet-stream", "image/jpeg"):
        content_type = _EXT_TO_MIME.get(ext, "application/octet-stream")
    safe_name = "".join(c for c in file_storage.filename if c.isalnum() or c in "._- ").strip() or "archivo"
    path = f"{prefix_path}/{uuid.uuid4().hex}_{safe_name}"
    try:
        blob = bucket.blob(path)
        blob.upload_from_string(raw, content_type=content_type)
        url = _generate_signed_url(blob, expiration_minutes=60 * 24 * 7)
        return url, None
    except Exception as e:
        return None, str(e)


def _upload_one_legajo_filestorage(bucket, company_id: str, empleado_id: str, file_storage):
    """Sube un archivo (werkzeug FileStorage) a legajos_uploads/… Devuelve (meta dict | None, error str | None)."""
    if not file_storage or not file_storage.filename:
        return None, "Archivo vacío o sin nombre."
    raw = file_storage.read()
    if len(raw) > UPLOAD_MAX_BYTES:
        return None, f"El archivo supera el límite de {UPLOAD_MAX_BYTES // (1024*1024)} MB."
    ext = (file_storage.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in UPLOAD_ALLOWED_EXTENSIONS:
        return None, f"Tipo no permitido. Permitidos: {', '.join(sorted(UPLOAD_ALLOWED_EXTENSIONS))}."
    content_type = (file_storage.content_type or "").split(";")[0].strip()
    if not content_type or content_type in ("application/octet-stream", "binary/octet-stream", "image/jpeg"):
        content_type = _EXT_TO_MIME.get(ext, "application/octet-stream")
    safe_name = "".join(c for c in file_storage.filename if c.isalnum() or c in "._- ").strip() or "archivo"
    cid = _normalize_company_id(company_id) or str(company_id or "").strip().lower()
    eid = str(empleado_id or "").strip()
    path = f"legajos_uploads/{cid}/{eid}/{uuid.uuid4().hex}_{safe_name}"
    try:
        blob = bucket.blob(path)
        blob.upload_from_string(raw, content_type=content_type)
        return (
            {
                "storage_path": path,
                "filename": safe_name,
                "content_type": content_type,
                "size_bytes": len(raw),
            },
            None,
        )
    except Exception as e:
        return None, str(e)


def _upload_legajo_file_to_storage(company_id: str, empleado_id: str):
    """Sube un único archivo del request (campo ``file``)."""
    bucket = _get_storage_bucket()
    if not bucket:
        return None, "Storage no configurado (Firebase Storage)."
    file_storage = request.files.get("file")
    if not file_storage or not file_storage.filename:
        return None, "No se envió ningún archivo."
    return _upload_one_legajo_filestorage(bucket, company_id, empleado_id, file_storage)


@flask_app.post("/api/comunicados/upload")
@rrhh_permission_required(
    auth_rrhh.PERM_COMUNICADOS_SEND,
    message="No tenés permisos para enviar comunicados.",
)
def api_comunicados_upload():
    """Sube una imagen o PDF para adjuntar al comunicado; devuelve la URL pública o firmada."""
    url, err = _upload_file_to_storage("comunicados_uploads")
    if err:
        if "Storage no configurado" in err:
            return jsonify({"ok": False, "error": err + " Configurá FIREBASE_STORAGE_BUCKET (ej: it-analyzer.firebasestorage.app)."}), 501
        if "No se envió" in err or "Tipo no permitido" in err or "supera el límite" in err:
            return jsonify({"ok": False, "error": err}), 400
        return jsonify({"ok": False, "error": err}), 500
    return jsonify({"ok": True, "url": url})


@flask_app.post("/api/rrhh/conversaciones/<conversation_id>/mensajes")
@rrhh_permission_required(
    auth_rrhh.PERM_CONVERSATIONS_MANAGE,
    message="Sin permiso para gestionar conversaciones.",
)
def rrhh_responder_api(conversation_id):
    data = request.get_json(silent=True) or {}
    mensaje = str(data.get("mensaje") or "").strip()
    media_url_raw = data.get("media_url") or data.get("media_urls")
    media_urls = []
    if isinstance(media_url_raw, list):
        media_urls = [str(u).strip() for u in media_url_raw if u and str(u).strip().startswith("http")]
    elif isinstance(media_url_raw, str) and media_url_raw.strip().startswith("http"):
        media_urls = [media_url_raw.strip()]
    agente = _resolve_rrhh_agent(data)
    if not mensaje and not media_urls:
        return jsonify({"ok": False, "error": "Escribí un mensaje o agregá la URL de una imagen/archivo"}), 400

    conv = _fetch_handoff(conversation_id)
    if not conv:
        return jsonify({"ok": False, "error": "Conversación no encontrada"}), 404
    if not _conversation_matches_selected_company(conv):
        return jsonify({"ok": False, "error": "Conversación no disponible para esta empresa."}), 404

    estado = str(conv.get("estado") or "").strip().lower()
    if estado == HANDOFF_STATUS_CLOSED:
        return jsonify({"ok": False, "error": "La conversación está cerrada"}), 400

    if estado != HANDOFF_STATUS_ACTIVE:
        _take_handoff(conversation_id, agente)

    # Idempotencia básica para evitar duplicados por doble submit/click en pocos segundos.
    ultimo_texto_rrhh = str(conv.get("ultimo_rrhh_text") or "").strip()
    ultimo_agente_rrhh = str(conv.get("ultimo_rrhh_agente_id") or "").strip().lower()
    ultimo_rrhh_at = _as_utc_naive(conv.get("ultimo_rrhh_at"))
    ahora = _as_utc_naive(_utc_now())
    is_same_message = ultimo_texto_rrhh.lower() == mensaje.lower()
    same_agent = ultimo_agente_rrhh == str(agente.get("agent_id") or "").strip().lower()
    if ultimo_rrhh_at and ahora and is_same_message and same_agent and not media_urls:
        if (ahora - ultimo_rrhh_at).total_seconds() <= 8:
            return jsonify({"ok": True, "conversation_id": conversation_id, "duplicate_ignored": True})

    _add_handoff_message(
        conversation_id,
        remitente="rrhh",
        texto=mensaje or "(archivo adjunto)",
        agente=agente.get("display_name"),
        media_url=media_urls if media_urls else None,
    )
    _upsert_handoff(
        conversation_id,
        {
            "rrhh_agente": agente.get("display_name"),
            "rrhh_agente_id": agente.get("agent_id"),
            "updated_at": _utc_now(),
            "ultimo_rrhh_text": mensaje,
            "ultimo_rrhh_agente_id": str(agente.get("agent_id") or "").strip().lower(),
            "ultimo_rrhh_at": _utc_now(),
        },
        merge=True,
    )
    return jsonify({"ok": True, "conversation_id": conversation_id})


@flask_app.get("/api/filtros/contexto")
@rrhh_auth_required
def filtros_contexto_api():
    """Empresas con branches y areas para filtros en estadísticas e historial."""
    company = _current_company()
    ctx = _companies_for_filter_context()
    return jsonify({
        "ok": True,
        "companies": ctx,
        "selected_company_id": company.get("company_id") if company else None,
    })


@flask_app.get("/api/stats")
@rrhh_auth_required
def stats_api():
    if not _can_view_stats():
        return jsonify({"ok": False, "error": "Sin permiso para ver estadísticas."}), 403
    company_id_raw = request.args.get("company_id", "").strip() or None
    company_id = _normalize_company_id(company_id_raw) if company_id_raw else None
    branches_param = request.args.get("branches")
    branches = [b.strip() for b in (branches_param or "").split(",") if b.strip()] if branches_param else None
    areas_param = request.args.get("areas")
    areas = [a.strip() for a in (areas_param or "").split(",") if a.strip()] if areas_param else None
    if company_id:
        handoff_records = _list_handoffs(
            include_closed=True,
            limit=10000,
            company_id=company_id,
            branches=branches if branches else None,
            areas=areas if areas else None,
        )
    else:
        handoff_records = _all_handoff_records_for_stats()
    stats = stats_service.obtener_estadisticas(
        chatbot.db,
        rrhh_records=handoff_records,
        company_id=company_id,
    )
    filter_applied = {
        "company_id": company_id,
        "branches": branches or [],
        "areas": areas or [],
    }
    response = jsonify(
        {
            "ok": True,
            "source_project": _firebase_project_id(),
            "server_boot_at": SERVER_BOOT_AT,
            "filter_applied": filter_applied,
            **stats,
        }
    )
    response.headers["Cache-Control"] = "no-store"
    response.headers["X-Server-Boot-At"] = SERVER_BOOT_AT
    response.headers["X-Firebase-Project"] = _firebase_project_id()
    return response


@flask_app.post("/api/reset")
def reset_api():
    limpiar_estado_conversacion()
    _clear_handoff_session()
    _clear_chat_context()
    session["chat_session_id"] = _new_conversation_id()
    return jsonify(
        {
            "ok": True,
            "reply": "Sesión reiniciada. ¿De qué empresa me hablás? Elegí una opción o escribí el nombre.",
            "quick_actions": _construir_acciones_empresas(limite=8),
            "handoff_active": False,
        }
    )


if __name__ == "__main__":
    puerto = int(os.getenv("PORT", "5000"))
    flask_app.run(host="0.0.0.0", port=puerto, debug=False)
